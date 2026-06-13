"""
One-time migration to multi-tenant.

Assigns all existing single-shop data to a first business account, attaches
existing users to it, and creates the platform super-admin login.

Run from the project directory, e.g.:

    python3 migrate_to_multitenant.py \
        --business "Animal Nexus Pharmacy" \
        --superadmin-user root \
        --superadmin-pass 'choose-a-strong-pass'

Backend is taken from your .env (DB_BACKEND). Safe to re-run: it won't create
a second tenant or a second super-admin, and only fills in missing tenant_ids.
"""

import argparse
import sys
from werkzeug.security import generate_password_hash

import config

# Tables that carry tenant-owned data (MySQL backend).
_TENANT_TABLES = [
    "products", "customers", "invoices", "invoice_items", "credit_ledger",
    "suppliers", "purchase_invoices", "purchase_invoice_items", "supplier_ledger",
    "accounts", "journal_entries", "journal_lines", "partners", "partner_transactions",
]


def migrate_mysql(business, su_user, su_pass):
    import pymysql
    import mysql_db
    mysql_db.init_workbook()  # ensure tenants table + tenant_id columns exist

    conn = pymysql.connect(
        host=config.MYSQL_HOST, port=config.MYSQL_PORT, user=config.MYSQL_USER,
        password=config.MYSQL_PASSWORD, database=config.MYSQL_DATABASE,
        charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor,
    )
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT tenant_id FROM tenants ORDER BY tenant_id LIMIT 1")
            row = cur.fetchone()
            if row:
                tid = row["tenant_id"]
                print(f"Using existing tenant #{tid}.")
            else:
                cur.execute("INSERT INTO tenants (name, is_active) VALUES (%s, 1)", (business,))
                tid = cur.lastrowid
                print(f"Created tenant #{tid}: {business}")

            for tbl in _TENANT_TABLES:
                cur.execute(f"UPDATE {tbl} SET tenant_id=%s WHERE tenant_id IS NULL", (tid,))
                if cur.rowcount:
                    print(f"  {tbl}: assigned {cur.rowcount} rows to tenant #{tid}")

            cur.execute(
                "UPDATE users SET tenant_id=%s WHERE tenant_id IS NULL AND role <> 'super_admin'",
                (tid,))
            if cur.rowcount:
                print(f"  users: attached {cur.rowcount} existing user(s) to tenant #{tid}")

            cur.execute("SELECT COUNT(*) AS n FROM users WHERE role='super_admin'")
            if cur.fetchone()["n"] == 0:
                cur.execute(
                    """INSERT INTO users (tenant_id, username, password_hash, role, is_active, created_at)
                       VALUES (NULL, %s, %s, 'super_admin', 1, NOW())""",
                    (su_user.strip().lower(), generate_password_hash(su_pass)))
                print(f"  created super-admin '{su_user}'")
            else:
                print("  super-admin already exists — skipped.")
        conn.commit()
    finally:
        conn.close()
    print("MySQL migration complete.")


def migrate_excel(business, su_user, su_pass):
    import shutil
    import excel_db as db
    from openpyxl import load_workbook

    db.init_master()
    tenants = db.get_all_tenants()
    if tenants:
        tid = tenants[0]["tenant_id"]
        print(f"Using existing tenant #{tid}.")
    else:
        tid = db.create_tenant(business)
        print(f"Created tenant #{tid}: {business}")

    # Move the legacy data.xlsx to become this tenant's workbook.
    target = db._DATA_DIR / f"data_{tid}.xlsx"
    if db.DATA_FILE.exists() and not target.exists():
        shutil.copy2(db.DATA_FILE, target)
        print(f"  copied data.xlsx -> {target.name}")

    # Lift users out of the legacy workbook into the master registry.
    if db.DATA_FILE.exists():
        legacy = load_workbook(db.DATA_FILE)
        if "Users" in legacy.sheetnames:
            existing = {u["username"] for u in db.get_all_users()}
            moved = 0
            for row in legacy["Users"].iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                # legacy layout: user_id, username, password_hash, role, is_active, created_at
                username = str(row[1]).strip().lower()
                if username in existing:
                    continue
                db.add_user(username, row[2], role=row[3] or "staff", tenant_id=tid)
                moved += 1
            if moved:
                print(f"  moved {moved} user(s) into the master registry under tenant #{tid}")
        legacy.close()

    if not db.has_any_super_admin():
        db.add_user(su_user.strip().lower(), generate_password_hash(su_pass),
                    role="super_admin", tenant_id=None)
        print(f"  created super-admin '{su_user}'")
    else:
        print("  super-admin already exists — skipped.")
    print("Excel migration complete.")


def main():
    ap = argparse.ArgumentParser(description="Migrate the POS to multi-tenant.")
    ap.add_argument("--business", required=True, help="Name of the first business account")
    ap.add_argument("--superadmin-user", required=True, help="Platform super-admin username")
    ap.add_argument("--superadmin-pass", required=True, help="Platform super-admin password")
    args = ap.parse_args()

    if len(args.superadmin_pass) < 6:
        print("Super-admin password must be at least 6 characters.")
        sys.exit(1)

    backend = config.DB_BACKEND
    print(f"Backend: {backend}")
    if backend == "mysql":
        migrate_mysql(args.business, args.superadmin_user, args.superadmin_pass)
    else:
        migrate_excel(args.business, args.superadmin_user, args.superadmin_pass)


if __name__ == "__main__":
    main()
