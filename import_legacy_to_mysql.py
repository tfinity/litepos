"""
One-time importer: load a legacy single-shop Excel workbook (data.xlsx) into the
multi-tenant MySQL database as a new business account (tenant), and create the
platform super-admin.

Use this to move a client who has been running locally on Excel onto the MySQL
server. IDs, sales history and accounting are preserved.

Prerequisites:
- .env points at the target MySQL (DB_BACKEND=mysql, MYSQL_* set).
- The target database should be EMPTY of real data (the importer refuses to run
  if any tenant already exists, unless you pass --force).

Usage:
    python3 import_legacy_to_mysql.py \
        --excel data.xlsx \
        --business "Animal Nexus Pharmacy" \
        --superadmin-user root \
        --superadmin-pass 'strong-pass'
"""

import argparse
import sys
from datetime import datetime

import pymysql
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

import config

# Each entry: Excel sheet -> (MySQL table, [column names in order], pk column)
# The Excel header row must match these column names (it does, by construction).
_SPEC = [
    ("Products", "products",
     ["product_id", "name", "purchase_price", "counter_price", "retail_price",
      "quantity", "barcode", "expiry_date", "category", "created_at", "last_supplier_id"],
     "product_id"),
    ("Customers", "customers",
     ["customer_id", "name", "phone", "email", "address", "tax_id", "notes", "created_at"],
     "customer_id"),
    ("Invoices", "invoices",
     ["invoice_id", "created_at", "subtotal", "discount_total", "tax_rate", "tax_amount",
      "total", "payment_method", "customer_id", "status", "deleted_at", "deleted_by",
      "delete_reason"],
     "invoice_id"),
    ("InvoiceItems", "invoice_items",
     ["item_id", "invoice_id", "product_id", "product_name", "purchase_price",
      "counter_price", "quantity", "discount_amount", "line_total"],
     "item_id"),
    ("CreditLedger", "credit_ledger",
     ["entry_id", "customer_id", "invoice_id", "type", "amount", "note", "created_at"],
     "entry_id"),
    ("Suppliers", "suppliers",
     ["supplier_id", "name", "phone", "email", "address", "notes", "created_at"],
     "supplier_id"),
    ("Purchases", "purchase_invoices",
     ["purchase_id", "supplier_id", "created_at", "total_amount", "notes"],
     "purchase_id"),
    ("PurchaseItems", "purchase_invoice_items",
     ["item_id", "purchase_id", "product_id", "product_name", "quantity", "unit_cost",
      "line_total"],
     "item_id"),
    ("SupplierLedger", "supplier_ledger",
     ["entry_id", "supplier_id", "purchase_id", "type", "amount", "note", "created_at"],
     "entry_id"),
    ("Accounts", "accounts",
     ["account_id", "code", "name", "type", "is_system", "created_at"],
     "account_id"),
    ("JournalEntries", "journal_entries",
     ["entry_id", "date", "description", "source_type", "source_id", "created_by", "created_at"],
     "entry_id"),
    ("JournalLines", "journal_lines",
     ["line_id", "entry_id", "account_id", "debit", "credit"],
     "line_id"),
    ("Partners", "partners",
     ["partner_id", "name", "share_pct", "is_active", "created_at"],
     "partner_id"),
    ("PartnerTransactions", "partner_transactions",
     ["txn_id", "partner_id", "type", "amount", "note", "date", "created_at"],
     "txn_id"),
]


def _read_sheet(wb, sheet, columns):
    """Return list of row-dicts for a sheet, mapping by header name. Missing sheet -> []."""
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = {}
        for col in columns:
            i = idx.get(col)
            v = row[i] if i is not None and i < len(row) else None
            d[col] = None if v == "" else v
        rows.append(d)
    return rows


def main():
    ap = argparse.ArgumentParser(description="Import a legacy Excel shop into MySQL as a tenant.")
    ap.add_argument("--excel", default="data.xlsx", help="Path to the legacy data.xlsx")
    ap.add_argument("--business", required=True, help="Business/account name")
    ap.add_argument("--superadmin-user", required=True)
    ap.add_argument("--superadmin-pass", required=True)
    ap.add_argument("--force", action="store_true", help="Proceed even if tenants already exist")
    args = ap.parse_args()

    if config.DB_BACKEND != "mysql":
        print("ERROR: .env DB_BACKEND must be 'mysql' for this importer.")
        sys.exit(1)
    if len(args.superadmin_pass) < 6:
        print("ERROR: super-admin password must be at least 6 characters.")
        sys.exit(1)

    # Ensure the MySQL schema exists (tables + tenant_id columns).
    import mysql_db
    mysql_db.init_workbook()

    wb = load_workbook(args.excel, data_only=True)

    conn = pymysql.connect(
        host=config.MYSQL_HOST, port=config.MYSQL_PORT, user=config.MYSQL_USER,
        password=config.MYSQL_PASSWORD, database=config.MYSQL_DATABASE,
        charset="utf8mb4", cursorclass=pymysql.cursors.DictCursor, autocommit=False,
    )
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS n FROM tenants")
        if cur.fetchone()["n"] > 0 and not args.force:
            print("ERROR: tenants already exist in this database. Refusing to import.\n"
                  "Wipe the false data first (or pass --force if you know what you're doing).")
            sys.exit(1)

        # Create the business + its super-admin.
        cur.execute("INSERT INTO tenants (name, is_active) VALUES (%s, 1)", (args.business,))
        tid = cur.lastrowid
        cur.execute("SELECT COUNT(*) AS n FROM users WHERE role='super_admin'")
        if cur.fetchone()["n"] == 0:
            cur.execute(
                """INSERT INTO users (tenant_id, username, password_hash, role, is_active, created_at)
                   VALUES (NULL, %s, %s, 'super_admin', 1, NOW())""",
                (args.superadmin_user.strip().lower(), generate_password_hash(args.superadmin_pass)))
        print(f"Created tenant #{tid}: {args.business}")

        # Bring the shop's existing logins across (so the client keeps their
        # username/password). Legacy Users sheet: user_id, username, password_hash,
        # role, is_active, created_at. We re-key user_id (auto) and scope to tenant.
        for u in _read_sheet(wb, "Users",
                             ["username", "password_hash", "role", "is_active", "created_at"]):
            uname = (u.get("username") or "").strip().lower()
            if not uname or not u.get("password_hash"):
                continue
            role = u.get("role") or "staff"
            role = role if role in ("admin", "staff") else "staff"
            cur.execute("SELECT user_id FROM users WHERE username = %s", (uname,))
            if cur.fetchone():
                continue  # username already taken (e.g. super-admin)
            cur.execute(
                """INSERT INTO users (tenant_id, username, password_hash, role, is_active, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (tid, uname, u["password_hash"], role,
                 1 if (u.get("is_active") in (1, True, "1", "TRUE", "True", None)) else 0,
                 u.get("created_at") or datetime.now()))
            print(f"  imported login: {uname} ({role})")

        total = 0
        for sheet, table, columns, pk in _SPEC:
            rows = _read_sheet(wb, sheet, columns)
            if not rows:
                continue
            cols_sql = ", ".join(["tenant_id"] + columns)
            placeholders = ", ".join(["%s"] * (len(columns) + 1))
            sql = f"INSERT INTO {table} ({cols_sql}) VALUES ({placeholders})"
            for r in rows:
                cur.execute(sql, [tid] + [r[c] for c in columns])
            print(f"  {sheet} -> {table}: {len(rows)} rows")
            total += len(rows)

        # Commit the data atomically FIRST. AUTO_INCREMENT resets are DDL (they
        # implicitly commit), so they must run AFTER the transaction — otherwise a
        # later failure would leave a half-imported, already-committed database.
        conn.commit()

        for sheet, table, columns, pk in _SPEC:
            cur.execute(f"SELECT MAX({pk}) AS m FROM {table}")
            mx = cur.fetchone()["m"] or 0
            cur.execute(f"ALTER TABLE {table} AUTO_INCREMENT = {mx + 1}")

        print(f"\nImport complete: {total} rows into tenant #{tid}.")
        print("Log in as the super-admin, or as the shop's existing admin if you imported users.")
    except Exception as e:
        conn.rollback()
        print(f"\nERROR during import (rolled back): {e}")
        sys.exit(1)
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    main()
