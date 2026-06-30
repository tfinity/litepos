"""Excel-based data access layer for the POS system.

Multi-tenant: each business account has its own workbook ``data_<tenant>.xlsx``;
tenants and users live in a shared ``master.xlsx``. The active tenant comes from
the per-request context (tenant.py). When no tenant is set, operations fall back
to the legacy single-shop ``data.xlsx`` so older behaviour is preserved.
"""

import threading
from datetime import datetime, date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tenant import get_current_tenant

_DATA_DIR = Path(__file__).parent
DATA_FILE = _DATA_DIR / "data.xlsx"            # legacy / no-tenant fallback
MASTER_FILE = _DATA_DIR / "master.xlsx"        # tenants + users registry
_BACKUP_DIR = _DATA_DIR / "backups"
_MAX_BACKUPS = 7
_lock = threading.Lock()


def _data_file():
    """Path to the active tenant's workbook (or the legacy file when no tenant)."""
    tid = get_current_tenant()
    return _DATA_DIR / f"data_{tid}.xlsx" if tid is not None else DATA_FILE


def _atomic_save(wb, path):
    """Atomic save with rolling backup for the given workbook path."""
    if path.exists():
        _BACKUP_DIR.mkdir(exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        import shutil
        shutil.copy2(path, _BACKUP_DIR / f"{path.stem}_{ts}.xlsx")
        backups = sorted(_BACKUP_DIR.glob(f"{path.stem}_*.xlsx"))
        for old in backups[:-_MAX_BACKUPS]:
            old.unlink(missing_ok=True)
    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    # Windows doesn't allow replace() over an existing file — remove first
    if path.exists():
        path.unlink()
    tmp.replace(path)


def _save(wb):
    """Atomic save of the active tenant's workbook."""
    _atomic_save(wb, _data_file())


# ── Master registry (tenants + users), shared across all accounts ────

def _open_master():
    return load_workbook(MASTER_FILE)


def _save_master(wb):
    _atomic_save(wb, MASTER_FILE)


def init_master():
    """Ensure the master workbook (Tenants + Users sheets) exists."""
    if not MASTER_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Tenants"
        ws.append(TENANT_HEADERS)
        wb.create_sheet("Users").append(USER_HEADERS)
        _atomic_save(wb, MASTER_FILE)
        return
    # Migrate: ensure both sheets exist
    wb = _open_master()
    changed = False
    if "Tenants" not in wb.sheetnames:
        wb.create_sheet("Tenants").append(TENANT_HEADERS)
        changed = True
    if "Users" not in wb.sheetnames:
        wb.create_sheet("Users").append(USER_HEADERS)
        changed = True
    if changed:
        _atomic_save(wb, MASTER_FILE)
    wb.close()

PRODUCT_HEADERS = [
    "product_id", "name", "purchase_price", "counter_price", "retail_price",
    "quantity", "barcode", "expiry_date", "category", "created_at", "last_supplier_id",
]
CUSTOMER_HEADERS = [
    "customer_id", "name", "phone", "email", "address", "tax_id", "notes", "created_at",
]
INVOICE_HEADERS = [
    "invoice_id", "created_at", "subtotal", "discount_total",
    "tax_rate", "tax_amount", "total", "payment_method",
    "customer_id",
    "status", "deleted_at", "deleted_by", "delete_reason",
    "delivery_charges",  # must be last for backward compat with older workbooks (short rows → None)
]
ITEM_HEADERS = [
    "item_id", "invoice_id", "product_id", "product_name",
    "purchase_price", "counter_price", "quantity",
    "discount_amount", "line_total",
    "batch_id",  # must be last for backward compat with older workbooks (short rows -> None)
]
CREDIT_LEDGER_HEADERS = [
    "entry_id", "customer_id", "invoice_id", "type", "amount", "note", "created_at",
    "payment_method",  # must be last for backward compat with older workbooks (short rows -> None)
]
USER_HEADERS = [
    "user_id", "tenant_id", "username", "password_hash", "role", "is_active", "created_at",
]
TENANT_HEADERS = [
    "tenant_id", "name", "is_active", "created_at",
]
SUPPLIER_HEADERS = [
    "supplier_id", "name", "phone", "email", "address", "notes", "created_at",
]
PURCHASE_HEADERS = [
    "purchase_id", "supplier_id", "created_at", "total_amount", "notes", "payment_method",
]
PURCHASE_ITEM_HEADERS = [
    "item_id", "purchase_id", "product_id", "product_name", "quantity", "unit_cost", "line_total",
    "batch_id",  # must be last for backward compat with older workbooks (short rows -> None)
]
SUPPLIER_LEDGER_HEADERS = [
    "entry_id", "supplier_id", "purchase_id", "type", "amount", "note", "created_at", "payment_method",
]
BATCH_HEADERS = [
    "batch_id", "product_id", "batch_number", "supplier_id", "purchase_id",
    "unit_cost", "qty_received", "qty_remaining", "expiry_date", "received_at",
]

# ── Double-entry accounting ──────────────────────────────────────────
ACCOUNT_HEADERS = [
    "account_id", "code", "name", "type", "is_system", "created_at",
]
JOURNAL_HEADERS = [
    "entry_id", "date", "description", "source_type", "source_id", "created_by", "created_at",
]
JOURNAL_LINE_HEADERS = [
    "line_id", "entry_id", "account_id", "debit", "credit",
]

PARTNER_HEADERS = [
    "partner_id", "name", "share_pct", "is_active", "created_at",
]
PARTNER_TXN_HEADERS = [
    "txn_id", "partner_id", "type", "amount", "note", "date", "created_at",
]

# Account types: asset, liability, equity, income, expense
# Normal balance: assets & expenses are debit-normal; the rest credit-normal.
CHART_OF_ACCOUNTS = [
    # code, name, type, is_system
    ("1000", "Cash",                "asset",     True),
    ("1010", "Bank",                "asset",     True),
    ("1100", "Accounts Receivable", "asset",     True),
    ("1200", "Inventory",           "asset",     True),
    ("2000", "Accounts Payable",    "liability", True),
    ("2100", "Tax Payable",         "liability", True),
    ("3000", "Owner Capital",       "equity",    True),
    ("3100", "Owner Drawings",      "equity",    True),
    ("3900", "Retained Earnings",   "equity",    True),
    ("4000", "Sales Revenue",       "income",    True),
    ("5000", "Cost of Goods Sold",  "expense",   True),
    ("5100", "Rent",                "expense",   False),
    ("5200", "Salaries",            "expense",   False),
    ("5300", "Utilities",           "expense",   False),
    ("5400", "Transport",           "expense",   False),
    ("5900", "Miscellaneous",       "expense",   False),
]
_DEBIT_NORMAL_TYPES = ("asset", "expense")


def _is_valid_xlsx():
    from zipfile import BadZipFile
    try:
        wb = load_workbook(_data_file())
        wb.close()
        return True
    except (BadZipFile, Exception):
        return False


def init_workbook():
    """Create the active tenant's workbook with header rows if missing; migrate existing."""
    init_master()
    data_file = _data_file()
    if data_file.exists() and not _is_valid_xlsx():
        data_file.unlink()
    if not data_file.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(PRODUCT_HEADERS)
        ws2 = wb.create_sheet("Invoices")
        ws2.append(INVOICE_HEADERS)
        ws3 = wb.create_sheet("InvoiceItems")
        ws3.append(ITEM_HEADERS)
        ws4 = wb.create_sheet("Customers")
        ws4.append(CUSTOMER_HEADERS)
        ws5 = wb.create_sheet("CreditLedger")
        ws5.append(CREDIT_LEDGER_HEADERS)
        ws6 = wb.create_sheet("Users")
        ws6.append(USER_HEADERS)
        ws7 = wb.create_sheet("Suppliers")
        ws7.append(SUPPLIER_HEADERS)
        ws8 = wb.create_sheet("Purchases")
        ws8.append(PURCHASE_HEADERS)
        ws9 = wb.create_sheet("PurchaseItems")
        ws9.append(PURCHASE_ITEM_HEADERS)
        ws10 = wb.create_sheet("SupplierLedger")
        ws10.append(SUPPLIER_LEDGER_HEADERS)
        ws11 = wb.create_sheet("Accounts")
        ws11.append(ACCOUNT_HEADERS)
        ws12 = wb.create_sheet("JournalEntries")
        ws12.append(JOURNAL_HEADERS)
        ws13 = wb.create_sheet("JournalLines")
        ws13.append(JOURNAL_LINE_HEADERS)
        ws14 = wb.create_sheet("Partners")
        ws14.append(PARTNER_HEADERS)
        ws15 = wb.create_sheet("PartnerTransactions")
        ws15.append(PARTNER_TXN_HEADERS)
        ws16 = wb.create_sheet("ProductBatches")
        ws16.append(BATCH_HEADERS)
        _save(wb)
        wb.close()
    ensure_workbook_schema()
    seed_chart_of_accounts()


def ensure_workbook_schema():
    """Add Customers sheet and Invoices.customer_id column to legacy workbooks."""
    if not _data_file().exists():
        return
    with _lock:
        wb = _open()
        changed = False
        if "Customers" not in wb.sheetnames:
            ws = wb.create_sheet("Customers")
            ws.append(CUSTOMER_HEADERS)
            changed = True
        ws_inv = wb["Invoices"]
        last_h = ws_inv.cell(row=1, column=ws_inv.max_column).value
        if last_h != "customer_id":
            col = ws_inv.max_column + 1
            ws_inv.cell(row=1, column=col).value = "customer_id"
            changed = True
        if "CreditLedger" not in wb.sheetnames:
            ws_cl = wb.create_sheet("CreditLedger")
            ws_cl.append(CREDIT_LEDGER_HEADERS)
            changed = True
        if "Users" not in wb.sheetnames:
            ws_u = wb.create_sheet("Users")
            ws_u.append(USER_HEADERS)
            changed = True
        if "Suppliers" not in wb.sheetnames:
            wb.create_sheet("Suppliers").append(SUPPLIER_HEADERS)
            changed = True
        if "Purchases" not in wb.sheetnames:
            wb.create_sheet("Purchases").append(PURCHASE_HEADERS)
            changed = True
        if "PurchaseItems" not in wb.sheetnames:
            wb.create_sheet("PurchaseItems").append(PURCHASE_ITEM_HEADERS)
            changed = True
        if "SupplierLedger" not in wb.sheetnames:
            wb.create_sheet("SupplierLedger").append(SUPPLIER_LEDGER_HEADERS)
            changed = True
        if "Accounts" not in wb.sheetnames:
            wb.create_sheet("Accounts").append(ACCOUNT_HEADERS)
            changed = True
        if "JournalEntries" not in wb.sheetnames:
            wb.create_sheet("JournalEntries").append(JOURNAL_HEADERS)
            changed = True
        if "JournalLines" not in wb.sheetnames:
            wb.create_sheet("JournalLines").append(JOURNAL_LINE_HEADERS)
            changed = True
        if "Partners" not in wb.sheetnames:
            wb.create_sheet("Partners").append(PARTNER_HEADERS)
            changed = True
        if "PartnerTransactions" not in wb.sheetnames:
            wb.create_sheet("PartnerTransactions").append(PARTNER_TXN_HEADERS)
            changed = True
        if "ProductBatches" not in wb.sheetnames:
            wb.create_sheet("ProductBatches").append(BATCH_HEADERS)
            changed = True
        # Add last_supplier_id column to Products if missing
        ws_prod = wb["Products"]
        if ws_prod.cell(row=1, column=11).value != "last_supplier_id":
            ws_prod.cell(row=1, column=11).value = "last_supplier_id"
            changed = True
        # Add soft-delete columns to Invoices if missing
        ws_inv2 = wb["Invoices"]
        inv_headers = [ws_inv2.cell(row=1, column=c).value for c in range(1, ws_inv2.max_column + 1)]
        for col_name in ("status", "deleted_at", "deleted_by", "delete_reason"):
            if col_name not in inv_headers:
                ws_inv2.cell(row=1, column=ws_inv2.max_column + 1).value = col_name
                changed = True
                inv_headers.append(col_name)
        # Add payment_method column to Purchases if missing
        if "Purchases" in wb.sheetnames:
            ws_pur = wb["Purchases"]
            pur_headers = [ws_pur.cell(row=1, column=c).value for c in range(1, ws_pur.max_column + 1)]
            if "payment_method" not in pur_headers:
                ws_pur.cell(row=1, column=ws_pur.max_column + 1).value = "payment_method"
                changed = True
        # Add batch_id column to InvoiceItems if missing
        if "InvoiceItems" in wb.sheetnames:
            ws_ii = wb["InvoiceItems"]
            ii_headers = [ws_ii.cell(row=1, column=c).value for c in range(1, ws_ii.max_column + 1)]
            if "batch_id" not in ii_headers:
                ws_ii.cell(row=1, column=ws_ii.max_column + 1).value = "batch_id"
                changed = True
        # Add batch_id column to PurchaseItems if missing
        if "PurchaseItems" in wb.sheetnames:
            ws_pi = wb["PurchaseItems"]
            pi_headers = [ws_pi.cell(row=1, column=c).value for c in range(1, ws_pi.max_column + 1)]
            if "batch_id" not in pi_headers:
                ws_pi.cell(row=1, column=ws_pi.max_column + 1).value = "batch_id"
                changed = True
        if changed:
            _save(wb)
        wb.close()


def normalize_customer_id(val):
    """Return int customer id or None for blank / invalid."""
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _open():
    return load_workbook(_data_file())


def _next_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None and isinstance(row[0], (int, float)):
            max_id = max(max_id, int(row[0]))
    return max_id + 1


def _to_float(val, default=0.0):
    """Parse a form value to float; blank/None/invalid -> default."""
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _to_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _row_to_dict(headers, row):
    d = {}
    for i, h in enumerate(headers):
        val = row[i] if i < len(row) else None
        d[h] = val
    return d


def _normalize_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


# ── Products ──────────────────────────────────────────────────────────

def get_all_products():
    with _lock:
        wb = _open()
        ws = wb["Products"]
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            p = _row_to_dict(PRODUCT_HEADERS, row)
            p["expiry_date"] = _normalize_date(p["expiry_date"])
            products.append(p)
        wb.close()
    return products


def get_product(product_id):
    for p in get_all_products():
        if int(p["product_id"]) == int(product_id):
            return p
    return None


def get_product_by_barcode(barcode):
    for p in get_all_products():
        if str(p.get("barcode", "")).strip() == str(barcode).strip():
            return p
    return None


def add_product(data):
    with _lock:
        wb = _open()
        ws = wb["Products"]
        pid = _next_id(ws)
        expiry = data.get("expiry_date")
        if isinstance(expiry, str) and expiry:
            expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
        elif not isinstance(expiry, (date, datetime)):
            expiry = None
        ws.append([
            pid,
            data["name"],
            _to_float(data.get("purchase_price")),
            _to_float(data.get("counter_price")),
            _to_float(data.get("retail_price")),
            _to_int(data.get("quantity")),
            data.get("barcode", ""),
            expiry,
            data.get("category", ""),
            datetime.now(),
        ])
        _save(wb)
        wb.close()
    return pid


def update_product(product_id, data):
    with _lock:
        wb = _open()
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == int(product_id):
                row[1].value = data["name"]
                row[2].value = _to_float(data.get("purchase_price"))
                row[3].value = _to_float(data.get("counter_price"))
                row[4].value = _to_float(data.get("retail_price"))
                row[5].value = _to_int(data.get("quantity"))
                row[6].value = data.get("barcode", "")
                expiry = data.get("expiry_date")
                if isinstance(expiry, str) and expiry:
                    expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
                elif not isinstance(expiry, (date, datetime)):
                    expiry = None
                row[7].value = expiry
                row[8].value = data.get("category", "")
                break
        _save(wb)
        wb.close()


def delete_product(product_id):
    with _lock:
        wb = _open()
        ws = wb["Products"]
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[0].value) == int(product_id):
                ws.delete_rows(idx)
                break
        _save(wb)
        wb.close()


def _ensure_legacy_batches():
    """Auto-heal: give every product that has stock but no batch history a
    single legacy batch reflecting its current (blended) cost/quantity, so
    batch-aware code has something to read for pre-existing stock. Idempotent
    — only creates a batch for a product that doesn't already have one."""
    with _lock:
        wb = _open()
        if "ProductBatches" not in wb.sheetnames:
            wb.create_sheet("ProductBatches").append(BATCH_HEADERS)
        ws_b = wb["ProductBatches"]
        ws_p = wb["Products"]

        has_batch = set()
        for row in ws_b.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            has_batch.add(int(row[1]))

        batch_id = _next_id(ws_b)
        created = 0
        for row in ws_p.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            pid = int(row[0])
            qty = int(row[5] or 0)
            if qty <= 0 or pid in has_batch:
                continue
            purchase_price = float(row[2] or 0)
            expiry = row[7] if len(row) > 7 else None
            created_at = row[9] if len(row) > 9 else None
            last_supplier_id = row[10] if len(row) > 10 else None
            ws_b.append([
                batch_id, pid, f"P{pid}-LEGACY", last_supplier_id, None,
                purchase_price, qty, qty, expiry, created_at,
            ])
            batch_id += 1
            created += 1
        if created:
            _save(wb)
        wb.close()
    return created


def _recompute_product_cache(wb, product_id):
    """Recompute a product's cached quantity/purchase_price from its batches'
    remaining stock (weighted-average cost). Must run inside an already-open
    workbook (caller owns _lock/_save/close). No-op if the product has no
    batches at all (nothing to derive from)."""
    pid = int(product_id)
    if "ProductBatches" not in wb.sheetnames:
        return
    ws_b = wb["ProductBatches"]
    total_qty = 0
    total_value = 0.0
    for row in ws_b.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None:
            continue
        if int(row[1].value) != pid:
            continue
        qty_rem = int(row[7].value or 0)
        cost = float(row[5].value or 0)
        total_qty += qty_rem
        total_value += qty_rem * cost
    avg_cost = round(total_value / total_qty, 2) if total_qty > 0 else None
    ws_p = wb["Products"]
    for row in ws_p.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == pid:
            row[5].value = total_qty
            if avg_cost is not None:
                row[2].value = avg_cost
            break


def _deplete_batch_for_sale(wb, pid, qty, requested_batch_id, product_name):
    """Pick a batch to sell qty units of product pid from, decrement its
    qty_remaining, and return (batch_id, unit_cost). If requested_batch_id is
    given, sell from exactly that batch (staff picked it at the till). If not,
    auto-pick the oldest active batch that alone covers qty. Raises ValueError
    if there isn't enough stock in the selected (or any single) batch."""
    if "ProductBatches" not in wb.sheetnames:
        wb.create_sheet("ProductBatches").append(BATCH_HEADERS)
    ws_b = wb["ProductBatches"]

    if requested_batch_id is not None:
        target = None
        for brow in ws_b.iter_rows(min_row=2):
            if brow[0].value is not None and int(brow[0].value) == int(requested_batch_id):
                target = brow
                break
        if target is None or target[1].value is None or int(target[1].value) != pid:
            raise ValueError(f"Selected batch not found for '{product_name}'.")
        remaining = int(target[7].value or 0)
        if qty > remaining:
            raise ValueError(
                f"Not enough stock in batch '{target[2].value}' for '{product_name}': "
                f"requested {qty}, available {remaining}"
            )
        target[7].value = remaining - qty
        return int(target[0].value), float(target[5].value or 0)

    candidates = []
    for brow in ws_b.iter_rows(min_row=2):
        if brow[0].value is None or brow[1].value is None:
            continue
        if int(brow[1].value) != pid or int(brow[7].value or 0) <= 0:
            continue
        candidates.append(brow)
    candidates.sort(key=lambda r: (r[9].value or datetime.min, int(r[0].value)))
    for brow in candidates:
        remaining = int(brow[7].value or 0)
        if remaining >= qty:
            brow[7].value = remaining - qty
            return int(brow[0].value), float(brow[5].value or 0)

    total_available = sum(int(b[7].value or 0) for b in candidates)
    raise ValueError(
        f"Not enough stock for '{product_name}': requested {qty}, available {total_available}"
    )


def _restore_batch_for_sale(wb, pid, qty, batch_id=None):
    """Return qty units back to the batch they were sold from (or that
    product's legacy batch if batch_id is unknown — pre-batch invoices)."""
    if "ProductBatches" not in wb.sheetnames:
        return
    ws_b = wb["ProductBatches"]
    target = None
    if batch_id is not None:
        for brow in ws_b.iter_rows(min_row=2):
            if brow[0].value is not None and int(brow[0].value) == int(batch_id):
                target = brow
                break
    if target is None:
        for brow in ws_b.iter_rows(min_row=2):
            if brow[0].value is None or brow[1].value is None:
                continue
            if int(brow[1].value) == pid and brow[4].value is None:
                target = brow
                break
    if target is not None:
        target[7].value = int(target[7].value or 0) + qty


def get_product_batches(product_id=None, active_only=False):
    """Return ProductBatches rows, optionally for one product, optionally
    only those with stock remaining. Auto-heals legacy (pre-batch) stock
    into a synthetic batch on first read."""
    _ensure_legacy_batches()
    pid = int(product_id) if product_id is not None else None
    with _lock:
        wb = _open()
        if "ProductBatches" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["ProductBatches"]
        batches = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            b = _row_to_dict(BATCH_HEADERS, row)
            if pid is not None and (b.get("product_id") is None or int(b["product_id"]) != pid):
                continue
            if active_only and float(b.get("qty_remaining") or 0) <= 0:
                continue
            b["expiry_date"] = _normalize_date(b.get("expiry_date"))
            batches.append(b)
        wb.close()
    batches.sort(key=lambda x: (x["received_at"] or datetime.min, x["batch_id"]))
    return batches


def get_low_stock_products(threshold=10):
    return [p for p in get_all_products() if int(p["quantity"]) <= threshold]


def get_expiry_products(days_ahead=30):
    today = date.today()
    cutoff = today + timedelta(days=days_ahead)
    results = []
    for p in get_all_products():
        exp = p["expiry_date"]
        if exp is None:
            continue
        if exp <= cutoff:
            p["expired"] = exp < today
            p["days_left"] = (exp - today).days
            results.append(p)
    results.sort(key=lambda x: x["expiry_date"])
    return results


# ── Customers ─────────────────────────────────────────────────────────

def get_all_customers():
    with _lock:
        wb = _open()
        if "Customers" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Customers"]
        customers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            c = _row_to_dict(CUSTOMER_HEADERS, row)
            c["customer_id"] = int(c["customer_id"])
            customers.append(c)
        wb.close()
    customers.sort(key=lambda x: x["customer_id"])
    return customers


def get_customer(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return None
    for c in get_all_customers():
        if c["customer_id"] == cid:
            return c
    return None


def add_customer(data):
    with _lock:
        wb = _open()
        if "Customers" not in wb.sheetnames:
            ws_new = wb.create_sheet("Customers")
            ws_new.append(CUSTOMER_HEADERS)
        ws = wb["Customers"]
        cid = _next_id(ws)
        ws.append([
            cid,
            (data.get("name") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("email") or "").strip(),
            (data.get("address") or "").strip(),
            (data.get("tax_id") or "").strip(),
            (data.get("notes") or "").strip(),
            datetime.now(),
        ])
        _save(wb)
        wb.close()
    return cid


def update_customer(customer_id, data):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return
    with _lock:
        wb = _open()
        ws = wb["Customers"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            try:
                if int(float(row[0].value)) != cid:
                    continue
            except (ValueError, TypeError):
                continue
            row[1].value = (data.get("name") or "").strip()
            row[2].value = (data.get("phone") or "").strip()
            row[3].value = (data.get("email") or "").strip()
            row[4].value = (data.get("address") or "").strip()
            row[5].value = (data.get("tax_id") or "").strip()
            row[6].value = (data.get("notes") or "").strip()
            break
        _save(wb)
        wb.close()


def delete_customer(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return False
    with _lock:
        wb = _open()
        ws_inv = wb["Invoices"]
        for row in ws_inv.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            inv = _row_to_dict(INVOICE_HEADERS, row)
            if normalize_customer_id(inv.get("customer_id")) == cid:
                wb.close()
                raise ValueError("Cannot delete customer: invoices reference this profile.")
        ws = wb["Customers"]
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            try:
                if int(float(row[0].value)) != cid:
                    continue
            except (ValueError, TypeError):
                continue
            ws.delete_rows(idx)
            break
        _save(wb)
        wb.close()
    return True


def search_customers(query):
    q = str(query).lower().strip()
    if not q:
        return []
    results = []
    for c in get_all_customers():
        if (q in str(c.get("name", "")).lower()
                or q in str(c.get("phone", "")).lower()
                or q in str(c.get("email", "")).lower()):
            results.append(c)
    return results[:20]


def customer_lookup():
    """Map customer_id -> customer dict."""
    return {c["customer_id"]: c for c in get_all_customers()}


def get_sales_summary_by_customer():
    """Per-customer invoice count and total revenue (all time)."""
    cmap = customer_lookup()
    by_c = {cid: {"customer": c, "invoice_count": 0, "total_revenue": 0.0} for cid, c in cmap.items()}
    walk_in = {"customer": None, "invoice_count": 0, "total_revenue": 0.0}
    for inv in get_all_invoices():
        total = float(inv["total"] or 0)
        cid = normalize_customer_id(inv.get("customer_id"))
        if cid is None or cid not in by_c:
            walk_in["invoice_count"] += 1
            walk_in["total_revenue"] += total
        else:
            bucket = by_c[cid]
            bucket["invoice_count"] += 1
            bucket["total_revenue"] += total
    rows = [v for v in by_c.values() if v["invoice_count"]]
    rows.sort(key=lambda x: x["total_revenue"], reverse=True)
    walk_in["total_revenue"] = round(walk_in["total_revenue"], 2)
    for r in rows:
        r["total_revenue"] = round(r["total_revenue"], 2)
    walk_in["invoice_count"] = walk_in["invoice_count"]  # no round
    return rows, walk_in


def get_invoices_for_customer(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return []
    return [
        inv for inv in get_all_invoices()
        if normalize_customer_id(inv.get("customer_id")) == cid
    ]


def get_customer_product_aggregates(customer_id):
    """Total quantity and amount per product for all invoices of this customer."""
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return []
    inv_ids = {
        int(inv["invoice_id"])
        for inv in get_all_invoices()
        if normalize_customer_id(inv.get("customer_id")) == cid
    }
    if not inv_ids:
        return []
    with _lock:
        wb = _open()
        ws = wb["InvoiceItems"]
        by_pid = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(ITEM_HEADERS, row)
            iid = int(item["invoice_id"])
            if iid not in inv_ids:
                continue
            pid = int(item["product_id"])
            name = item.get("product_name") or ""
            if pid not in by_pid:
                by_pid[pid] = {
                    "product_id": pid,
                    "product_name": name,
                    "total_qty": 0,
                    "total_amount": 0.0,
                }
            by_pid[pid]["total_qty"] += int(item["quantity"] or 0)
            by_pid[pid]["total_amount"] += float(item["line_total"] or 0)
            if not by_pid[pid]["product_name"]:
                by_pid[pid]["product_name"] = name
        wb.close()
    rows = list(by_pid.values())
    rows.sort(key=lambda x: str(x["product_name"]).lower())
    for r in rows:
        r["total_amount"] = round(r["total_amount"], 2)
    return rows


def _customer_exists_in_workbook(wb, customer_id):
    cid = int(customer_id)
    if "Customers" not in wb.sheetnames:
        return False
    ws = wb["Customers"]
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] is None:
            continue
        try:
            if int(float(row[0])) == cid:
                return True
        except (ValueError, TypeError):
            continue
    return False


# ── Invoices ──────────────────────────────────────────────────────────

def get_all_invoices(include_deleted=False, start_date=None, end_date=None):
    with _lock:
        wb = _open()
        ws = wb["Invoices"]
        invoices = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            inv = _row_to_dict(INVOICE_HEADERS, row)
            if not include_deleted and inv.get("status") == "deleted":
                continue
            if start_date or end_date:
                dt = inv.get("created_at")
                inv_date = dt.date() if hasattr(dt, "date") else dt
                if inv_date:
                    if start_date and inv_date < start_date:
                        continue
                    if end_date and inv_date > end_date:
                        continue
            invoices.append(inv)
        wb.close()
    invoices.sort(key=lambda x: x["invoice_id"], reverse=True)
    return invoices


def get_deleted_invoices():
    with _lock:
        wb = _open()
        ws = wb["Invoices"]
        invoices = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            inv = _row_to_dict(INVOICE_HEADERS, row)
            if inv.get("status") == "deleted":
                invoices.append(inv)
        wb.close()
    invoices.sort(key=lambda x: x["invoice_id"], reverse=True)
    return invoices


def get_invoice(invoice_id):
    with _lock:
        wb = _open()
        ws = wb["Invoices"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            inv = _row_to_dict(INVOICE_HEADERS, row)
            if int(inv["invoice_id"]) == int(invoice_id):
                wb.close()
                return inv
        wb.close()
    return None


def delete_invoice(invoice_id, deleted_by, reason=""):
    """Soft-delete an invoice: mark deleted, reverse stock, reverse credit ledger."""
    iid = int(invoice_id)
    with _lock:
        wb = _open()
        ws_inv = wb["Invoices"]
        ws_items = wb["InvoiceItems"]
        ws_cl = wb["CreditLedger"] if "CreditLedger" in wb.sheetnames else None

        # Find and mark invoice row
        inv_row_ref = None
        for row in ws_inv.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == iid:
                inv_row_ref = row
                break
        if inv_row_ref is None:
            wb.close()
            raise ValueError("Invoice not found")

        inv = _row_to_dict(INVOICE_HEADERS, [c.value for c in inv_row_ref])
        if inv.get("status") == "deleted":
            wb.close()
            raise ValueError("Invoice already deleted")

        # Guard: deleting a credit sale that the customer has already paid
        # against would leave them overpaid (negative balance). Block it.
        cid = normalize_customer_id(inv.get("customer_id"))
        is_credit = str(inv.get("payment_method") or "").strip().lower() == "credit"
        if is_credit and cid is not None and ws_cl is not None:
            total_debit = 0.0
            total_paid = 0.0
            for row in ws_cl.iter_rows(min_row=2, values_only=True):
                if row[0] is None or normalize_customer_id(row[1]) != cid:
                    continue
                if row[3] == "debit":
                    total_debit += float(row[4] or 0)
                elif row[3] == "credit":
                    total_paid += float(row[4] or 0)
            inv_amount = float(inv.get("total") or 0)
            if round((total_debit - inv_amount) - total_paid, 2) < 0:
                wb.close()
                raise ValueError(
                    "This credit sale has payments against the customer's account. "
                    "Deleting it would leave the customer overpaid. Record a refund or "
                    "adjust their balance first, then delete.")

        # Reverse stock for each item — back to the specific batch it was
        # sold from (or that product's legacy batch for pre-batch invoices).
        touched_pids = set()
        for row in ws_items.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(ITEM_HEADERS, row)
            if int(item["invoice_id"]) != iid:
                continue
            pid = int(item["product_id"] or 0)
            qty = int(item["quantity"] or 0)
            _restore_batch_for_sale(wb, pid, qty, item.get("batch_id"))
            touched_pids.add(pid)
        for pid in touched_pids:
            _recompute_product_cache(wb, pid)

        # Reverse credit ledger entries for this invoice
        if ws_cl is not None:
            for row in ws_cl.iter_rows(min_row=2):
                if row[0].value is None:
                    continue
                if str(row[2].value) == str(iid):  # invoice_id column (index 2)
                    row[3].value = "deleted"  # mark entry type as deleted

        # Write soft-delete fields onto invoice row
        headers = [ws_inv.cell(row=1, column=c).value for c in range(1, ws_inv.max_column + 1)]
        col_status = headers.index("status") + 1 if "status" in headers else None
        col_dat = headers.index("deleted_at") + 1 if "deleted_at" in headers else None
        col_by = headers.index("deleted_by") + 1 if "deleted_by" in headers else None
        col_reason = headers.index("delete_reason") + 1 if "delete_reason" in headers else None

        inv_row_num = inv_row_ref[0].row
        if col_status:
            ws_inv.cell(row=inv_row_num, column=col_status).value = "deleted"
        if col_dat:
            ws_inv.cell(row=inv_row_num, column=col_dat).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if col_by:
            ws_inv.cell(row=inv_row_num, column=col_by).value = deleted_by
        if col_reason:
            ws_inv.cell(row=inv_row_num, column=col_reason).value = reason

        # Remove the auto-synced journal entry for this sale so the deleted
        # invoice drops out of the accounting reports too.
        if "JournalEntries" in wb.sheetnames:
            ws_je = wb["JournalEntries"]
            del_eids = set()
            del_je_rows = []
            for idx, row in enumerate(ws_je.iter_rows(min_row=2), start=2):
                if row[0].value is None:
                    continue
                if row[3].value == "sale" and row[4].value is not None \
                        and int(row[4].value) == iid:
                    del_eids.add(int(row[0].value))
                    del_je_rows.append(idx)
            for idx in sorted(del_je_rows, reverse=True):
                ws_je.delete_rows(idx, 1)
            if del_eids and "JournalLines" in wb.sheetnames:
                ws_jl = wb["JournalLines"]
                del_jl_rows = [idx for idx, row in enumerate(ws_jl.iter_rows(min_row=2), start=2)
                               if row[0].value is not None and int(row[1].value) in del_eids]
                for idx in sorted(del_jl_rows, reverse=True):
                    ws_jl.delete_rows(idx, 1)

        _save(wb)
        wb.close()


def get_invoice_items(invoice_id):
    with _lock:
        wb = _open()
        ws = wb["InvoiceItems"]
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(ITEM_HEADERS, row)
            if int(item["invoice_id"]) == int(invoice_id):
                items.append(item)
        wb.close()
    return items


def create_invoice(items, tax_rate, payment_method, customer_id=None, delivery_charges=0.0):
    """
    items: list of dicts with keys: product_id, quantity, discount_amount,
    optional unit_price (sale price per unit for this line; defaults to product counter_price).
    discount_amount is a direct amount per unit (not percentage).
    customer_id: optional profile id (must exist on Customers sheet).
    delivery_charges: optional flat delivery fee added to the total.
    Returns the new invoice_id, or raises ValueError on stock/discount issues.
    """
    with _lock:
        wb = _open()
        ws_products = wb["Products"]
        ws_invoices = wb["Invoices"]
        ws_items = wb["InvoiceItems"]

        cid = normalize_customer_id(customer_id)
        if payment_method == "Credit" and cid is None:
            wb.close()
            raise ValueError("Credit payment requires a customer to be selected.")
        if cid is not None and not _customer_exists_in_workbook(wb, cid):
            wb.close()
            raise ValueError("Customer not found.")

        invoice_id = _next_id(ws_invoices)
        item_id_start = _next_id(ws_items)

        product_rows = {}
        for row in ws_products.iter_rows(min_row=2):
            if row[0].value is not None:
                product_rows[int(row[0].value)] = row

        subtotal = 0.0
        discount_total = 0.0
        line_entries = []
        touched_pids = set()

        for i, item in enumerate(items):
            pid = int(item["product_id"])
            qty = int(item["quantity"])
            discount_per_unit = float(item.get("discount_amount", 0))
            requested_batch_id = item.get("batch_id")

            if pid not in product_rows:
                raise ValueError(f"Product ID {pid} not found")
            prow = product_rows[pid]
            catalog_counter = float(prow[3].value)  # product counter_price

            # Picks (or auto-picks) a batch and deducts qty from it; the
            # batch's own cost is the COGS for this line, not the blended
            # product-level cache.
            batch_id_used, purchase_price = _deplete_batch_for_sale(
                wb, pid, qty, requested_batch_id, prow[1].value)
            touched_pids.add(pid)

            raw_unit = item.get("unit_price")
            if raw_unit is None or raw_unit == "":
                unit_price = catalog_counter
            else:
                unit_price = float(raw_unit)

            if unit_price < 0:
                raise ValueError(
                    f"Invalid sale price for '{prow[1].value}': must be non-negative"
                )

            # Guard: price after per-unit discount must not go below purchase price
            discounted_price = unit_price - discount_per_unit
            if discounted_price < purchase_price:
                raise ValueError(
                    f"Discount too high for '{prow[1].value}': "
                    f"price after discount {discounted_price:.2f} is below "
                    f"purchase price {purchase_price:.2f}. "
                    f"Max discount: {unit_price - purchase_price:.2f}"
                )

            line_discount = discount_per_unit * qty
            line_total = discounted_price * qty
            subtotal += unit_price * qty  # pre-discount subtotal (uses line sale price)
            discount_total += line_discount

            line_entries.append([
                item_id_start + i,
                invoice_id,
                pid,
                prow[1].value,       # product_name
                purchase_price,
                unit_price,          # effective unit price for this line (receipt / history)
                qty,
                line_discount,
                round(line_total, 2),
                batch_id_used,
            ])

        for pid in touched_pids:
            _recompute_product_cache(wb, pid)

        net_subtotal = round(subtotal - discount_total, 2)
        tax_amount = round(net_subtotal * tax_rate, 2)
        delivery = round(float(delivery_charges or 0), 2)
        total = round(net_subtotal + tax_amount + delivery, 2)

        ws_invoices.append([
            invoice_id,
            datetime.now(),
            round(subtotal, 2),
            round(discount_total, 2),
            tax_rate,
            tax_amount,
            total,
            payment_method,
            cid,
            None,     # status
            None,     # deleted_at
            None,     # deleted_by
            None,     # delete_reason
            delivery, # delivery_charges
        ])

        for entry in line_entries:
            ws_items.append(entry)

        # Credit sale: record debit in ledger
        if payment_method == "Credit" and cid is not None:
            if "CreditLedger" not in wb.sheetnames:
                ws_cl = wb.create_sheet("CreditLedger")
                ws_cl.append(CREDIT_LEDGER_HEADERS)
            ws_cl = wb["CreditLedger"]
            cl_entry_id = _next_id(ws_cl)
            ws_cl.append([cl_entry_id, cid, invoice_id, "debit", total, "Credit sale", datetime.now()])

        _save(wb)
        wb.close()

    return invoice_id


def update_invoice(invoice_id, items, tax_rate, payment_method=None, customer_id=None):
    """
    Edit an existing invoice in-place.
    payment_method / customer_id: pass new values to change them; None keeps existing.
    Handles stock reconciliation and credit ledger (delete old debit, add new if Credit).
    Raises ValueError on validation errors.
    """
    with _lock:
        wb = _open()
        ws_products = wb["Products"]
        ws_invoices = wb["Invoices"]
        ws_items = wb["InvoiceItems"]

        # Locate invoice row
        inv_row_ref = None
        inv_row_idx = None
        for idx, row in enumerate(ws_invoices.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[0].value) == int(invoice_id):
                inv_row_ref = row
                inv_row_idx = idx
                break
        if inv_row_ref is None:
            wb.close()
            raise ValueError("Invoice not found.")

        invoice = _row_to_dict(INVOICE_HEADERS, [c.value for c in inv_row_ref])

        # Collect old item rows (indices in reverse so deletion is safe)
        old_items = []
        old_item_indices = []
        for idx, row in enumerate(ws_items.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[1].value) == int(invoice_id):
                old_items.append(_row_to_dict(ITEM_HEADERS, [c.value for c in row]))
                old_item_indices.append(idx)

        # Build product row lookup
        product_rows = {}
        for row in ws_products.iter_rows(min_row=2):
            if row[0].value is not None:
                product_rows[int(row[0].value)] = row

        # Return stock from old items to the exact batch each was sold from
        # (or that product's legacy batch for pre-batch invoices).
        for old in old_items:
            _restore_batch_for_sale(wb, int(old["product_id"]), int(old["quantity"]), old.get("batch_id"))

        # Validate + compute new items
        subtotal = 0.0
        discount_total = 0.0
        new_entries = []
        touched_pids = {int(old["product_id"]) for old in old_items}
        item_id_start = _next_id(ws_items)

        for i, item in enumerate(items):
            pid = int(item["product_id"])
            qty = int(item["quantity"])
            discount_per_unit = float(item.get("discount_amount", 0))
            requested_batch_id = item.get("batch_id")

            if pid not in product_rows:
                wb.close()
                raise ValueError(f"Product ID {pid} not found")
            prow = product_rows[pid]
            catalog_counter = float(prow[3].value)

            try:
                batch_id_used, purchase_price = _deplete_batch_for_sale(
                    wb, pid, qty, requested_batch_id, prow[1].value)
            except ValueError:
                wb.close()
                raise
            touched_pids.add(pid)

            raw_unit = item.get("unit_price")
            unit_price = float(raw_unit) if raw_unit not in (None, "") else catalog_counter

            if unit_price < 0:
                wb.close()
                raise ValueError(f"Invalid sale price for '{prow[1].value}'")

            discounted_price = unit_price - discount_per_unit
            if discounted_price < purchase_price:
                wb.close()
                raise ValueError(
                    f"Discount too high for '{prow[1].value}': "
                    f"price after discount {discounted_price:.2f} below purchase {purchase_price:.2f}"
                )

            line_discount = discount_per_unit * qty
            line_total = discounted_price * qty
            subtotal += unit_price * qty
            discount_total += line_discount

            new_entries.append([
                item_id_start + i,
                invoice_id,
                pid,
                prow[1].value,
                purchase_price,
                unit_price,
                qty,
                line_discount,
                round(line_total, 2),
                batch_id_used,
            ])

        for pid in touched_pids:
            _recompute_product_cache(wb, pid)

        net = round(subtotal - discount_total, 2)
        tax_amount = round(net * tax_rate, 2)
        total = round(net + tax_amount, 2)

        # Delete old item rows (reverse order)
        for idx in sorted(old_item_indices, reverse=True):
            ws_items.delete_rows(idx)

        for entry in new_entries:
            ws_items.append(entry)

        # Resolve new payment method / customer
        old_payment = invoice.get("payment_method")
        old_cid = normalize_customer_id(invoice.get("customer_id"))
        new_payment = payment_method if payment_method is not None else old_payment
        # customer_id sentinel: None means "keep old"; use the string "walkin" or 0 to clear
        if customer_id is None:
            new_cid = old_cid
        else:
            new_cid = normalize_customer_id(customer_id)

        if new_payment == "Credit" and new_cid is None:
            wb.close()
            raise ValueError("Credit payment requires a customer.")
        if new_cid is not None and not _customer_exists_in_workbook(wb, new_cid):
            wb.close()
            raise ValueError("Customer not found.")

        # Update invoice row
        inv_row_ref[2].value = round(subtotal, 2)
        inv_row_ref[3].value = round(discount_total, 2)
        inv_row_ref[4].value = tax_rate
        inv_row_ref[5].value = tax_amount
        inv_row_ref[6].value = total
        inv_row_ref[7].value = new_payment
        inv_row_ref[8].value = new_cid

        # Reconcile credit ledger:
        # Remove any existing debit entry for this invoice (regardless of old customer)
        if "CreditLedger" in wb.sheetnames:
            ws_cl = wb["CreditLedger"]
            to_del = []
            for idx, row in enumerate(ws_cl.iter_rows(min_row=2), start=2):
                if (row[0].value is not None
                        and row[2].value is not None
                        and int(row[2].value) == int(invoice_id)
                        and row[3].value == "debit"):
                    to_del.append(idx)
            for idx in sorted(to_del, reverse=True):
                ws_cl.delete_rows(idx)

        # Add fresh debit entry if new payment is Credit
        if new_payment == "Credit" and new_cid is not None:
            if "CreditLedger" not in wb.sheetnames:
                ws_cl = wb.create_sheet("CreditLedger")
                ws_cl.append(CREDIT_LEDGER_HEADERS)
            ws_cl = wb["CreditLedger"]
            cl_entry_id = _next_id(ws_cl)
            ws_cl.append([cl_entry_id, new_cid, invoice_id, "debit", total, "Credit sale", datetime.now()])

        _save(wb)
        wb.close()

    return invoice_id


def get_credit_ledger(customer_id=None):
    """Return all CreditLedger entries, optionally filtered by customer_id."""
    cid = normalize_customer_id(customer_id) if customer_id is not None else None
    with _lock:
        wb = _open()
        if "CreditLedger" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["CreditLedger"]
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            e = _row_to_dict(CREDIT_LEDGER_HEADERS, row)
            if cid is not None and normalize_customer_id(e.get("customer_id")) != cid:
                continue
            entries.append(e)
        wb.close()
    entries.sort(key=lambda x: x["entry_id"], reverse=True)
    return entries


def add_ledger_payment(customer_id, amount, note="", payment_method="cash"):
    """Record a cash payment from a customer (credit entry, reduces their balance)."""
    cid = normalize_customer_id(customer_id)
    if cid is None:
        raise ValueError("Invalid customer.")
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_customer(cid):
        raise ValueError("Customer not found.")
    pm = (payment_method or "cash").strip().lower()
    with _lock:
        wb = _open()
        if "CreditLedger" not in wb.sheetnames:
            ws_cl = wb.create_sheet("CreditLedger")
            ws_cl.append(CREDIT_LEDGER_HEADERS)
        ws_cl = wb["CreditLedger"]
        entry_id = _next_id(ws_cl)
        ws_cl.append([entry_id, cid, None, "credit", amount, (note or "").strip(), datetime.now(), pm])
        _save(wb)
        wb.close()
    return entry_id


def add_ledger_debit(customer_id, amount, note=""):
    """Record a manual debit against a customer (increases their outstanding balance)."""
    cid = normalize_customer_id(customer_id)
    if cid is None:
        raise ValueError("Invalid customer.")
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_customer(cid):
        raise ValueError("Customer not found.")
    with _lock:
        wb = _open()
        if "CreditLedger" not in wb.sheetnames:
            ws_cl = wb.create_sheet("CreditLedger")
            ws_cl.append(CREDIT_LEDGER_HEADERS)
        ws_cl = wb["CreditLedger"]
        entry_id = _next_id(ws_cl)
        ws_cl.append([entry_id, cid, None, "debit", amount, (note or "").strip(), datetime.now()])
        _save(wb)
        wb.close()
    return entry_id


def get_customer_balance(customer_id):
    """Return (total_debt, total_paid, balance). Positive balance = customer owes us."""
    entries = get_credit_ledger(customer_id=customer_id)
    total_debt = sum(float(e["amount"] or 0) for e in entries if e["type"] == "debit")
    total_paid = sum(float(e["amount"] or 0) for e in entries if e["type"] == "credit")
    balance = round(total_debt - total_paid, 2)
    return round(total_debt, 2), round(total_paid, 2), balance


def get_all_credit_balances():
    """Return list of {customer, customer_id, total_debt, total_paid, balance} for all customers with any ledger entry."""
    cmap = customer_lookup()
    entries = get_credit_ledger()
    by_cid = {}
    for e in entries:
        cid = normalize_customer_id(e.get("customer_id"))
        if cid is None:
            continue
        if cid not in by_cid:
            by_cid[cid] = {"total_debt": 0.0, "total_paid": 0.0}
        if e["type"] == "debit":
            by_cid[cid]["total_debt"] += float(e["amount"] or 0)
        else:
            by_cid[cid]["total_paid"] += float(e["amount"] or 0)
    result = []
    for cid, b in by_cid.items():
        balance = round(b["total_debt"] - b["total_paid"], 2)
        result.append({
            "customer": cmap.get(cid),
            "customer_id": cid,
            "total_debt": round(b["total_debt"], 2),
            "total_paid": round(b["total_paid"], 2),
            "balance": balance,
        })
    result.sort(key=lambda x: x["balance"], reverse=True)
    return result


def get_today_sales():
    today = date.today()
    count = 0
    total = 0.0
    for inv in get_all_invoices():
        created = inv["created_at"]
        if isinstance(created, datetime):
            inv_date = created.date()
        elif isinstance(created, date):
            inv_date = created
        else:
            continue
        if inv_date == today:
            count += 1
            total += float(inv["total"] or 0)
    return count, round(total, 2)


def search_products(query):
    q = str(query).lower().strip()
    results = []
    for p in get_all_products():
        if (q in str(p["name"]).lower()
                or q in str(p.get("barcode", "")).lower()):
            results.append(p)
    return results[:20]


def _parse_expiry(val):
    """Parse expiry from various formats: datetime, 'MM YYYY', 'M YYYY', date string."""
    if val is None or val == "" or val == "None":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    import re
    m = re.match(r'^(\d{1,2})\s+(\d{4})$', s)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            if month == 12:
                return date(year, 12, 31)
            return date(year, month + 1, 1) - timedelta(days=1)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def import_from_excel(filepath):
    """
    Import products from external Excel file.
    Columns: Sr.No, Product Name, Packing, Qty, Price(purchase), Company, Expiry, MRP, Sale Price(counter)
    Returns (imported_count, skipped_count, errors).
    """
    wb_src = load_workbook(filepath)
    ws = wb_src.active

    imported = 0
    skipped = 0
    errors = []

    # Detect if column I (index 8) exists (Sale Price / counter_price)
    has_counter_col = ws.max_column >= 9

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None and row[1] is None:
            continue
        name = str(row[1] or "").strip()
        if not name:
            skipped += 1
            continue
        packing = str(row[2] or "").strip()
        if packing:
            name = f"{name} ({packing})"

        qty = 0
        try:
            qty = int(row[3] or 0)
        except (ValueError, TypeError):
            pass

        purchase_price = 0.0
        try:
            purchase_price = float(row[4] or 0)
        except (ValueError, TypeError):
            pass

        company = str(row[5] or "").strip()
        expiry = _parse_expiry(row[6])

        mrp = 0.0
        try:
            mrp = float(row[7] or 0)
        except (ValueError, TypeError):
            pass

        # Counter price from column I if available, otherwise fall back to MRP
        counter_price = 0.0
        if has_counter_col and len(row) >= 9 and row[8] is not None:
            try:
                counter_price = float(row[8])
            except (ValueError, TypeError):
                counter_price = mrp if mrp > 0 else purchase_price
        else:
            counter_price = mrp if mrp > 0 else purchase_price

        try:
            add_product({
                "name": name,
                "purchase_price": purchase_price,
                "counter_price": counter_price,
                "retail_price": mrp,
                "quantity": qty,
                "barcode": "",
                "expiry_date": expiry,
                "category": company,
            })
            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
            skipped += 1

    wb_src.close()
    return imported, skipped, errors


# ── Users (stored in the shared master registry) ─────────────────────

def get_all_users(tenant_id=None):
    init_master()
    with _lock:
        wb = _open_master()
        ws = wb["Users"]
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            u = _row_to_dict(USER_HEADERS, row)
            u["user_id"] = int(u["user_id"])
            u["tenant_id"] = int(u["tenant_id"]) if u.get("tenant_id") not in (None, "") else None
            u["is_active"] = bool(u["is_active"])
            if tenant_id is not None and u["tenant_id"] != int(tenant_id):
                continue
            users.append(u)
        wb.close()
    return users


def get_user_by_id(user_id):
    for u in get_all_users():
        if u["user_id"] == int(user_id):
            return u
    return None


def get_user_by_username(username):
    username = (username or "").strip().lower()
    for u in get_all_users():
        if str(u["username"]).lower() == username:
            return u
    return None


def has_any_users():
    return len(get_all_users()) > 0


def has_any_super_admin():
    return any(u.get("role") == "super_admin" for u in get_all_users())


def add_user(username, password_hash, role="staff", tenant_id=None):
    init_master()
    with _lock:
        wb = _open_master()
        ws = wb["Users"]
        uid = _next_id(ws)
        ws.append([uid, int(tenant_id) if tenant_id is not None else None,
                   username.strip().lower(), password_hash, role, 1, datetime.now()])
        _save_master(wb)
        wb.close()
    return uid


def _update_user_field(user_id, col_index, value):
    with _lock:
        wb = _open_master()
        ws = wb["Users"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == int(user_id):
                row[col_index].value = value
                break
        _save_master(wb)
        wb.close()


def update_user_password(user_id, password_hash):
    _update_user_field(user_id, 3, password_hash)  # password_hash column


def set_user_role(user_id, role):
    _update_user_field(user_id, 4, role)  # role column


def toggle_user_active(user_id):
    u = get_user_by_id(user_id)
    _update_user_field(user_id, 5, 0 if (u and u["is_active"]) else 1)  # is_active column


def delete_user(user_id):
    with _lock:
        wb = _open_master()
        ws = wb["Users"]
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[0].value) == int(user_id):
                ws.delete_rows(idx)
                break
        _save_master(wb)
        wb.close()


# ── Tenants (business accounts) ──────────────────────────────────────

def create_tenant(name):
    name = str(name).strip()
    if not name:
        raise ValueError("Business name is required.")
    init_master()
    with _lock:
        wb = _open_master()
        ws = wb["Tenants"]
        tid = _next_id(ws)
        ws.append([tid, name, 1, datetime.now()])
        _save_master(wb)
        wb.close()
    return tid


def get_all_tenants(include_inactive=True):
    init_master()
    with _lock:
        wb = _open_master()
        ws = wb["Tenants"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            t = _row_to_dict(TENANT_HEADERS, row)
            t["tenant_id"] = int(t["tenant_id"])
            t["is_active"] = bool(t["is_active"])
            if not include_inactive and not t["is_active"]:
                continue
            rows.append(t)
        wb.close()
    rows.sort(key=lambda x: x["tenant_id"])
    return rows


def get_tenant(tenant_id):
    tid = int(tenant_id)
    for t in get_all_tenants():
        if t["tenant_id"] == tid:
            return t
    return None


def set_tenant_active(tenant_id, active):
    with _lock:
        wb = _open_master()
        ws = wb["Tenants"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == int(tenant_id):
                row[2].value = 1 if active else 0  # is_active column
                break
        _save_master(wb)
        wb.close()


# ── Suppliers ─────────────────────────────────────────────────────────

def get_all_suppliers():
    with _lock:
        wb = _open()
        ws = wb["Suppliers"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            s = _row_to_dict(SUPPLIER_HEADERS, row)
            s["supplier_id"] = int(s["supplier_id"])
            rows.append(s)
        wb.close()
    rows.sort(key=lambda x: x["supplier_id"])
    return rows


def get_supplier(supplier_id):
    sid = int(supplier_id)
    for s in get_all_suppliers():
        if s["supplier_id"] == sid:
            return s
    return None


def add_supplier(data):
    with _lock:
        wb = _open()
        ws = wb["Suppliers"]
        sid = _next_id(ws)
        ws.append([
            sid,
            (data.get("name") or "").strip(),
            (data.get("phone") or "").strip(),
            (data.get("email") or "").strip(),
            (data.get("address") or "").strip(),
            (data.get("notes") or "").strip(),
            datetime.now(),
        ])
        _save(wb)
        wb.close()
    return sid


def update_supplier(supplier_id, data):
    sid = int(supplier_id)
    with _lock:
        wb = _open()
        ws = wb["Suppliers"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == sid:
                row[1].value = (data.get("name") or "").strip()
                row[2].value = (data.get("phone") or "").strip()
                row[3].value = (data.get("email") or "").strip()
                row[4].value = (data.get("address") or "").strip()
                row[5].value = (data.get("notes") or "").strip()
                break
        _save(wb)
        wb.close()


def delete_supplier(supplier_id):
    sid = int(supplier_id)
    with _lock:
        wb = _open()
        ws_p = wb["Purchases"]
        for row in ws_p.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if row[1] is not None and int(row[1]) == sid:
                wb.close()
                raise ValueError("Cannot delete supplier: purchase invoices reference this supplier.")
        ws = wb["Suppliers"]
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[0].value) == sid:
                ws.delete_rows(idx)
                break
        _save(wb)
        wb.close()
    return True


def supplier_lookup():
    return {s["supplier_id"]: s for s in get_all_suppliers()}


# ── Purchase Invoices ─────────────────────────────────────────────────

def create_purchase_invoice(supplier_id, items, notes="", payment_method="credit",
                            purchase_date=None, direct_amount=None):
    """
    Receive stock from a supplier.
    items: list of {product_id, quantity, unit_cost}
    payment_method: 'credit' (unpaid -> owe the supplier), 'cash' or 'bank'
    (paid now -> reduces Cash/Bank, no payable).
    purchase_date: optional date (defaults to now) so past purchases can be backdated.
    direct_amount: when no items, record the purchase as a lump total only.
    """
    sid = int(supplier_id)
    payment_method = (payment_method or "credit").strip().lower()
    if payment_method not in ("credit", "cash", "bank"):
        payment_method = "credit"
    when = datetime.now()
    if purchase_date:
        if isinstance(purchase_date, str):
            when = datetime.combine(date.fromisoformat(purchase_date), datetime.min.time())
        elif isinstance(purchase_date, datetime):
            when = purchase_date
        elif isinstance(purchase_date, date):
            when = datetime.combine(purchase_date, datetime.min.time())
    with _lock:
        wb = _open()
        ws_prod = wb["Products"]
        ws_purch = wb["Purchases"]
        ws_pitems = wb["PurchaseItems"]

        if "SupplierLedger" not in wb.sheetnames:
            wb.create_sheet("SupplierLedger").append(SUPPLIER_LEDGER_HEADERS)

        ws_sl = wb["SupplierLedger"]

        # Build product row lookup
        product_rows = {}
        for row in ws_prod.iter_rows(min_row=2):
            if row[0].value is not None:
                product_rows[int(row[0].value)] = row

        if "ProductBatches" not in wb.sheetnames:
            wb.create_sheet("ProductBatches").append(BATCH_HEADERS)
        ws_b = wb["ProductBatches"]

        purchase_id = _next_id(ws_purch)
        item_id_start = _next_id(ws_pitems)
        total_amount = 0.0
        line_entries = []
        touched_pids = set()

        for i, item in enumerate(items):
            pid = int(item["product_id"])
            qty = int(item["quantity"])
            unit_cost = float(item["unit_cost"])
            if pid not in product_rows:
                wb.close()
                raise ValueError(f"Product ID {pid} not found")
            prow = product_rows[pid]
            line_total = round(unit_cost * qty, 2)
            total_amount += line_total

            batch_number = str(item.get("batch_number") or "").strip()
            expiry = item.get("expiry_date")
            if isinstance(expiry, str) and expiry:
                expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
            elif not isinstance(expiry, (date, datetime)):
                expiry = None

            batch_id_used = None
            if batch_number:
                # Reusing a lot number adds to that batch; reusing it at a
                # different cost is treated as a mistake, not a silent average.
                match_row = None
                for brow in ws_b.iter_rows(min_row=2):
                    if brow[0].value is None or brow[1].value is None:
                        continue
                    if int(brow[1].value) == pid and \
                       str(brow[2].value or "").strip().lower() == batch_number.lower():
                        match_row = brow
                        break
                if match_row is not None:
                    existing_cost = float(match_row[5].value or 0)
                    if round(existing_cost, 2) != round(unit_cost, 2):
                        wb.close()
                        raise ValueError(
                            f"Lot '{batch_number}' for {prow[1].value} already exists at cost "
                            f"{existing_cost:.2f} — use a different lot number if this is a different rate."
                        )
                    match_row[6].value = int(match_row[6].value or 0) + qty
                    match_row[7].value = int(match_row[7].value or 0) + qty
                    batch_id_used = int(match_row[0].value)
                else:
                    batch_id_used = _next_id(ws_b)
                    ws_b.append([batch_id_used, pid, batch_number, sid, purchase_id,
                                unit_cost, qty, qty, expiry, when])
            else:
                batch_id_used = _next_id(ws_b)
                ws_b.append([batch_id_used, pid, f"P{pid}-B{batch_id_used}", sid, purchase_id,
                            unit_cost, qty, qty, expiry, when])

            line_entries.append([
                item_id_start + i, purchase_id, pid,
                prow[1].value, qty, unit_cost, line_total, batch_id_used,
            ])
            prow[10].value = sid
            touched_pids.add(pid)

        for pid in touched_pids:
            _recompute_product_cache(wb, pid)

        if not items and direct_amount is not None:
            total_amount = float(direct_amount)
        total_amount = round(total_amount, 2)
        if total_amount <= 0:
            wb.close()
            raise ValueError("Purchase total must be positive.")
        ws_purch.append([purchase_id, sid, when, total_amount,
                         (notes or "").strip(), payment_method])
        for entry in line_entries:
            ws_pitems.append(entry)

        # Always record the purchase in the supplier ledger (relationship history).
        # Use the user's note if given, else a generic reference.
        debit_note = (notes or "").strip() or f"Purchase invoice #{purchase_id}"
        sl_id = _next_id(ws_sl)
        ws_sl.append([sl_id, sid, purchase_id, "debit", total_amount, debit_note, when, None])
        # If paid now, also record the payment so the supplier balance nets to zero.
        if payment_method in ("cash", "bank"):
            ws_sl.append([sl_id + 1, sid, purchase_id, "credit", total_amount,
                          f"Paid by {payment_method}", when, payment_method])

        _save(wb)
        wb.close()
    return purchase_id


def get_all_purchase_invoices():
    with _lock:
        wb = _open()
        ws = wb["Purchases"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            rows.append(_row_to_dict(PURCHASE_HEADERS, row))
        wb.close()
    rows.sort(key=lambda x: x["purchase_id"], reverse=True)
    return rows


def get_purchase_invoice(purchase_id):
    pid = int(purchase_id)
    for p in get_all_purchase_invoices():
        if int(p["purchase_id"]) == pid:
            return p
    return None


def get_purchase_invoice_items(purchase_id):
    pid = int(purchase_id)
    with _lock:
        wb = _open()
        ws = wb["PurchaseItems"]
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(PURCHASE_ITEM_HEADERS, row)
            if int(item["purchase_id"]) == pid:
                items.append(item)
        wb.close()
    return items


# ── Supplier Ledger ───────────────────────────────────────────────────

def add_supplier_payment(supplier_id, amount, note="", payment_method="cash"):
    sid = int(supplier_id)
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_supplier(sid):
        raise ValueError("Supplier not found.")
    pm = (payment_method or "cash").strip().lower()
    with _lock:
        wb = _open()
        if "SupplierLedger" not in wb.sheetnames:
            wb.create_sheet("SupplierLedger").append(SUPPLIER_LEDGER_HEADERS)
        ws_sl = wb["SupplierLedger"]
        entry_id = _next_id(ws_sl)
        ws_sl.append([entry_id, sid, None, "credit", amount,
                      (note or "").strip(), datetime.now(), pm])
        _save(wb)
        wb.close()
    return entry_id


def get_supplier_ledger_entries(supplier_id=None):
    sid = int(supplier_id) if supplier_id is not None else None
    with _lock:
        wb = _open()
        if "SupplierLedger" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["SupplierLedger"]
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            e = _row_to_dict(SUPPLIER_LEDGER_HEADERS, row)
            if sid is not None:
                if e.get("supplier_id") is None or int(e["supplier_id"]) != sid:
                    continue
            entries.append(e)
        wb.close()
    entries.sort(key=lambda x: x["entry_id"])
    return entries


def get_supplier_balance(supplier_id):
    entries = get_supplier_ledger_entries(supplier_id)
    total_debt = sum(float(e["amount"] or 0) for e in entries if e["type"] == "debit")
    total_paid = sum(float(e["amount"] or 0) for e in entries if e["type"] == "credit")
    return round(total_debt, 2), round(total_paid, 2), round(total_debt - total_paid, 2)


def delete_purchase_invoice(purchase_id):
    pid = int(purchase_id)
    with _lock:
        wb = _open()
        ws_purch = wb["Purchases"]
        ws_items = wb["PurchaseItems"]
        ws_sl = wb["SupplierLedger"] if "SupplierLedger" in wb.sheetnames else None
        ws_prod = wb["Products"]
        ws_b = wb["ProductBatches"] if "ProductBatches" in wb.sheetnames else None

        # Build product row lookup (col 5 = quantity)
        product_rows = {}
        for row in ws_prod.iter_rows(min_row=2):
            if row[0].value is not None:
                product_rows[int(row[0].value)] = row

        # Reverse stock: deduct from the specific batch this item created (or,
        # for purchases predating batch tracking, from that product's legacy
        # batch) and collect item rows to delete.
        item_rows_to_delete = []
        touched_pids = set()
        for idx, row in enumerate(ws_items.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if int(row[1].value) == pid:
                prod_id = int(row[2].value)
                qty = int(row[4].value or 0)
                batch_id = row[7].value if len(row) > 7 else None

                target = None
                if ws_b is not None:
                    if batch_id is not None:
                        for brow in ws_b.iter_rows(min_row=2):
                            if brow[0].value is not None and int(brow[0].value) == int(batch_id):
                                target = brow
                                break
                    else:
                        for brow in ws_b.iter_rows(min_row=2):
                            if brow[0].value is None or brow[1].value is None:
                                continue
                            if int(brow[1].value) == prod_id and brow[4].value is None:
                                target = brow
                                break
                if target is not None:
                    target[6].value = max(0, int(target[6].value or 0) - qty)
                    target[7].value = max(0, int(target[7].value or 0) - qty)
                    touched_pids.add(prod_id)
                elif prod_id in product_rows:
                    # No batch to reverse against at all — fall back to the
                    # pre-batch behaviour of decrementing stock directly.
                    prow = product_rows[prod_id]
                    prow[5].value = max(0, int(prow[5].value or 0) - qty)
                item_rows_to_delete.append(idx)
        for idx in sorted(item_rows_to_delete, reverse=True):
            ws_items.delete_rows(idx, 1)
        for prod_id in touched_pids:
            _recompute_product_cache(wb, prod_id)

        # Delete supplier ledger entries for this purchase
        if ws_sl:
            sl_rows_to_delete = []
            for idx, row in enumerate(ws_sl.iter_rows(min_row=2), start=2):
                if row[0].value is None:
                    continue
                if row[2].value is not None and int(row[2].value) == pid:
                    sl_rows_to_delete.append(idx)
            for idx in sorted(sl_rows_to_delete, reverse=True):
                ws_sl.delete_rows(idx, 1)

        # Delete journal entries and cascading lines
        if "JournalEntries" in wb.sheetnames:
            ws_je = wb["JournalEntries"]
            del_eids = set()
            del_je_rows = []
            for idx, row in enumerate(ws_je.iter_rows(min_row=2), start=2):
                if row[0].value is None:
                    continue
                if row[3].value == "purchase" and row[4].value is not None and int(row[4].value) == pid:
                    del_eids.add(int(row[0].value))
                    del_je_rows.append(idx)
            for idx in sorted(del_je_rows, reverse=True):
                ws_je.delete_rows(idx, 1)
            if del_eids and "JournalLines" in wb.sheetnames:
                ws_jl = wb["JournalLines"]
                del_jl_rows = [i for i, row in enumerate(ws_jl.iter_rows(min_row=2), start=2)
                               if row[0].value is not None and int(row[1].value) in del_eids]
                for idx in sorted(del_jl_rows, reverse=True):
                    ws_jl.delete_rows(idx, 1)

        # Delete the purchase invoice row
        purch_row = None
        for idx, row in enumerate(ws_purch.iter_rows(min_row=2), start=2):
            if row[0].value is not None and int(row[0].value) == pid:
                purch_row = idx
                break
        if purch_row:
            ws_purch.delete_rows(purch_row, 1)

        _save(wb)
        wb.close()


def delete_supplier_payment(entry_id):
    eid = int(entry_id)
    with _lock:
        wb = _open()
        if "SupplierLedger" not in wb.sheetnames:
            wb.close()
            raise ValueError("Entry not found.")
        ws_sl = wb["SupplierLedger"]

        row_to_delete = None
        for idx, row in enumerate(ws_sl.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if int(row[0].value) == eid:
                if row[3].value != "credit" or row[2].value is not None:
                    wb.close()
                    raise ValueError("Only standalone payment entries can be deleted.")
                row_to_delete = idx
                break

        if row_to_delete is None:
            wb.close()
            raise ValueError("Entry not found.")

        if "JournalEntries" in wb.sheetnames:
            ws_je = wb["JournalEntries"]
            del_eids = set()
            del_je_rows = []
            for idx, row in enumerate(ws_je.iter_rows(min_row=2), start=2):
                if row[0].value is None:
                    continue
                if row[3].value == "supplier_payment" and row[4].value is not None and int(row[4].value) == eid:
                    del_eids.add(int(row[0].value))
                    del_je_rows.append(idx)
            for idx in sorted(del_je_rows, reverse=True):
                ws_je.delete_rows(idx, 1)
            if del_eids and "JournalLines" in wb.sheetnames:
                ws_jl = wb["JournalLines"]
                del_jl_rows = [i for i, row in enumerate(ws_jl.iter_rows(min_row=2), start=2)
                               if row[0].value is not None and int(row[1].value) in del_eids]
                for idx in sorted(del_jl_rows, reverse=True):
                    ws_jl.delete_rows(idx, 1)

        ws_sl.delete_rows(row_to_delete, 1)
        _save(wb)
        wb.close()


def update_supplier_payment(entry_id, amount, note="", payment_method="cash"):
    eid = int(entry_id)
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    pm = (payment_method or "cash").strip().lower()

    with _lock:
        wb = _open()
        if "SupplierLedger" not in wb.sheetnames:
            wb.close()
            raise ValueError("Entry not found.")
        ws_sl = wb["SupplierLedger"]

        entry_created_at = None
        for row in ws_sl.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            if int(row[0].value) == eid:
                if row[3].value != "credit" or row[2].value is not None:
                    wb.close()
                    raise ValueError("Only standalone payment entries can be edited.")
                row[4].value = amount
                row[5].value = (note or "").strip()
                row[7].value = pm
                entry_created_at = row[6].value
                break
        else:
            wb.close()
            raise ValueError("Entry not found.")

        # Delete old journal entry for this payment
        if "JournalEntries" in wb.sheetnames:
            ws_je = wb["JournalEntries"]
            del_eids = set()
            del_je_rows = []
            for idx, row in enumerate(ws_je.iter_rows(min_row=2), start=2):
                if row[0].value is None:
                    continue
                if row[3].value == "supplier_payment" and row[4].value is not None and int(row[4].value) == eid:
                    del_eids.add(int(row[0].value))
                    del_je_rows.append(idx)
            for idx in sorted(del_je_rows, reverse=True):
                ws_je.delete_rows(idx, 1)
            if del_eids and "JournalLines" in wb.sheetnames:
                ws_jl = wb["JournalLines"]
                del_jl_rows = [i for i, row in enumerate(ws_jl.iter_rows(min_row=2), start=2)
                               if row[0].value is not None and int(row[1].value) in del_eids]
                for idx in sorted(del_jl_rows, reverse=True):
                    ws_jl.delete_rows(idx, 1)

        _save(wb)
        wb.close()

    # Re-post journal with updated values
    ap = get_account_by_code("2000")
    cash_acc = get_account_by_code("1010" if pm == "bank" else "1000")
    if ap and cash_acc:
        post_journal(
            "Supplier payment",
            [{"account_id": ap["account_id"], "debit": amount, "credit": 0},
             {"account_id": cash_acc["account_id"], "debit": 0, "credit": amount}],
            source_type="supplier_payment",
            source_id=eid,
            entry_date=entry_created_at,
        )


def get_all_supplier_balances():
    smap = supplier_lookup()
    entries = get_supplier_ledger_entries()
    by_sid = {}
    for e in entries:
        sid = e.get("supplier_id")
        if sid is None:
            continue
        sid = int(sid)
        if sid not in by_sid:
            by_sid[sid] = {"total_debt": 0.0, "total_paid": 0.0}
        if e["type"] == "debit":
            by_sid[sid]["total_debt"] += float(e["amount"] or 0)
        else:
            by_sid[sid]["total_paid"] += float(e["amount"] or 0)
    result = []
    for sid, b in by_sid.items():
        balance = round(b["total_debt"] - b["total_paid"], 2)
        result.append({
            "supplier": smap.get(sid),
            "supplier_id": sid,
            "total_debt": round(b["total_debt"], 2),
            "total_paid": round(b["total_paid"], 2),
            "balance": balance,
        })
    result.sort(key=lambda x: x["balance"], reverse=True)
    return result


# ── P&L Reports ───────────────────────────────────────────────────────

def get_sales_pl_report(start_date, end_date):
    """
    Returns per-item P&L for all sales in date range.
    Each row: date, invoice_id, product_name, qty, purchase_price, sale_price, profit
    Also returns aggregate totals.
    """
    start = start_date if isinstance(start_date, date) else datetime.strptime(str(start_date), "%Y-%m-%d").date()
    end = end_date if isinstance(end_date, date) else datetime.strptime(str(end_date), "%Y-%m-%d").date()

    # A credit invoice that has since been fully paid off (realized) should be
    # treated as "paid" here, not stuck in "credit" forever just because it was
    # originally booked on credit.
    realized, _ = _realized_credit_invoice_ids()

    # Collect invoice_ids in range + their dates and payment method
    inv_meta = {}
    for inv in get_all_invoices():
        created = inv["created_at"]
        if isinstance(created, datetime):
            inv_date = created.date()
        elif isinstance(created, date):
            inv_date = created
        else:
            continue
        if start <= inv_date <= end:
            iid = int(inv["invoice_id"])
            was_credit = str(inv.get("payment_method") or "").strip().lower() == "credit"
            inv_meta[iid] = {
                "date": inv_date,
                "is_credit": was_credit and iid not in realized,
            }

    if not inv_meta:
        return [], _empty_pl_totals()

    with _lock:
        wb = _open()
        ws = wb["InvoiceItems"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(ITEM_HEADERS, row)
            iid = int(item["invoice_id"])
            if iid not in inv_meta:
                continue
            qty = int(item["quantity"] or 0)
            purchase_price = float(item["purchase_price"] or 0)
            line_total = float(item["line_total"] or 0)
            cogs = round(purchase_price * qty, 2)
            profit = round(line_total - cogs, 2)
            rows.append({
                "date": inv_meta[iid]["date"],
                "invoice_id": iid,
                "is_credit": inv_meta[iid]["is_credit"],
                "product_name": item["product_name"],
                "quantity": qty,
                "purchase_price": purchase_price,
                "sale_price": round(line_total / qty, 2) if qty else 0,
                "line_total": line_total,
                "cogs": cogs,
                "profit": profit,
            })
        wb.close()

    rows.sort(key=lambda x: (x["date"], x["invoice_id"]))
    return rows, _split_pl_totals(rows)


def search_product_sales(product_query, start_date, end_date):
    """Return all invoice lines whose product_name contains product_query (case-insensitive)
    within the given date range. Includes customer name and invoice metadata."""
    q = (product_query or "").strip().lower()
    start = start_date if isinstance(start_date, date) else date.fromisoformat(str(start_date))
    end = end_date if isinstance(end_date, date) else date.fromisoformat(str(end_date))

    # Build invoice lookup with date + payment_method + customer_id
    inv_meta = {}
    for inv in get_all_invoices():
        created = inv.get("created_at")
        if isinstance(created, datetime):
            inv_date = created.date()
        elif isinstance(created, date):
            inv_date = created
        else:
            continue
        if start <= inv_date <= end:
            inv_meta[int(inv["invoice_id"])] = {
                "date": inv_date,
                "payment_method": inv.get("payment_method"),
                "customer_id": inv.get("customer_id"),
            }

    if not inv_meta:
        return []

    cmap = customer_lookup()

    with _lock:
        wb = _open()
        ws = wb["InvoiceItems"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(ITEM_HEADERS, row)
            iid = int(item["invoice_id"])
            if iid not in inv_meta:
                continue
            pname = (item.get("product_name") or "").lower()
            if q and q not in pname:
                continue
            meta = inv_meta[iid]
            qty = int(item.get("quantity") or 0)
            line_total = float(item.get("line_total") or 0)
            cust = cmap.get(int(meta["customer_id"])) if meta.get("customer_id") else None
            rows.append({
                "date": meta["date"],
                "invoice_id": iid,
                "payment_method": meta["payment_method"],
                "customer_name": cust["name"] if cust else None,
                "product_name": item.get("product_name"),
                "quantity": qty,
                "sale_price": round(line_total / qty, 2) if qty else 0,
                "line_total": line_total,
            })
        wb.close()

    rows.sort(key=lambda x: (x["date"], x["invoice_id"]), reverse=True)
    return rows


def _empty_pl_totals():
    zero = {"revenue": 0, "cogs": 0, "profit": 0}
    return {"revenue": 0, "cogs": 0, "profit": 0,
            "paid": dict(zero), "credit": dict(zero)}


def _split_pl_totals(rows):
    """Aggregate rows into grand totals plus paid vs credit buckets.
    Each row must have line_total, cogs, and is_credit."""
    buckets = {"paid": {"revenue": 0.0, "cogs": 0.0},
               "credit": {"revenue": 0.0, "cogs": 0.0}}
    for r in rows:
        b = buckets["credit"] if r.get("is_credit") else buckets["paid"]
        b["revenue"] += r["line_total"]
        b["cogs"] += r["cogs"]
    out = {}
    for key in ("paid", "credit"):
        rev = round(buckets[key]["revenue"], 2)
        cogs = round(buckets[key]["cogs"], 2)
        out[key] = {"revenue": rev, "cogs": cogs, "profit": round(rev - cogs, 2)}
    total_rev = round(out["paid"]["revenue"] + out["credit"]["revenue"], 2)
    total_cogs = round(out["paid"]["cogs"] + out["credit"]["cogs"], 2)
    out["revenue"] = total_rev
    out["cogs"] = total_cogs
    out["profit"] = round(total_rev - total_cogs, 2)
    return out


def get_supplier_sales_pl(supplier_id, start_date, end_date):
    """
    Returns sales P&L for products last supplied by a given supplier.
    Groups by product. Returns per-product rows + totals.
    """
    sid = int(supplier_id)
    start = start_date if isinstance(start_date, date) else datetime.strptime(str(start_date), "%Y-%m-%d").date()
    end = end_date if isinstance(end_date, date) else datetime.strptime(str(end_date), "%Y-%m-%d").date()

    # Products supplied by this supplier
    supplier_pids = set()
    for p in get_all_products():
        lsid = p.get("last_supplier_id")
        if lsid is not None:
            try:
                if int(lsid) == sid:
                    supplier_pids.add(int(p["product_id"]))
            except (ValueError, TypeError):
                pass

    if not supplier_pids:
        return [], {"revenue": 0, "cogs": 0, "profit": 0}

    # Invoice dates in range
    inv_dates = {}
    for inv in get_all_invoices():
        created = inv["created_at"]
        if isinstance(created, datetime):
            inv_date = created.date()
        elif isinstance(created, date):
            inv_date = created
        else:
            continue
        if start <= inv_date <= end:
            inv_dates[int(inv["invoice_id"])] = inv_date

    with _lock:
        wb = _open()
        ws = wb["InvoiceItems"]
        by_product = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            item = _row_to_dict(ITEM_HEADERS, row)
            iid = int(item["invoice_id"])
            if iid not in inv_dates:
                continue
            pid = int(item["product_id"] or 0)
            if pid not in supplier_pids:
                continue
            qty = int(item["quantity"] or 0)
            purchase_price = float(item["purchase_price"] or 0)
            line_total = float(item["line_total"] or 0)
            cogs = purchase_price * qty
            profit = line_total - cogs
            if pid not in by_product:
                by_product[pid] = {
                    "product_id": pid,
                    "product_name": item["product_name"] or "",
                    "total_qty": 0,
                    "total_cogs": 0.0,
                    "total_revenue": 0.0,
                    "total_profit": 0.0,
                }
            by_product[pid]["total_qty"] += qty
            by_product[pid]["total_cogs"] += cogs
            by_product[pid]["total_revenue"] += line_total
            by_product[pid]["total_profit"] += profit
        wb.close()

    rows = list(by_product.values())
    for r in rows:
        r["total_cogs"] = round(r["total_cogs"], 2)
        r["total_revenue"] = round(r["total_revenue"], 2)
        r["total_profit"] = round(r["total_profit"], 2)
    rows.sort(key=lambda x: x["product_name"].lower())
    total_revenue = round(sum(r["total_revenue"] for r in rows), 2)
    total_cogs = round(sum(r["total_cogs"] for r in rows), 2)
    total_profit = round(total_revenue - total_cogs, 2)
    return rows, {"revenue": total_revenue, "cogs": total_cogs, "profit": total_profit}


# ── Double-entry accounting: data layer ──────────────────────────────

def seed_chart_of_accounts():
    """Insert any standard accounts that are missing (idempotent, keyed by code)."""
    with _lock:
        wb = _open()
        if "Accounts" not in wb.sheetnames:
            wb.create_sheet("Accounts").append(ACCOUNT_HEADERS)
        ws = wb["Accounts"]
        existing_codes = set()
        max_id = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            existing_codes.add(str(row[1]))
            max_id = max(max_id, int(row[0]))
        added = False
        next_id = max_id + 1
        for code, name, atype, is_system in CHART_OF_ACCOUNTS:
            if str(code) in existing_codes:
                continue
            ws.append([next_id, code, name, atype, bool(is_system), datetime.now()])
            next_id += 1
            added = True
        if added:
            _save(wb)
        wb.close()


def get_all_accounts():
    with _lock:
        wb = _open()
        ws = wb["Accounts"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            a = _row_to_dict(ACCOUNT_HEADERS, row)
            a["account_id"] = int(a["account_id"])
            rows.append(a)
        wb.close()
    rows.sort(key=lambda x: str(x["code"]))
    return rows


def get_account(account_id):
    aid = int(account_id)
    for a in get_all_accounts():
        if a["account_id"] == aid:
            return a
    return None


def get_account_by_code(code):
    code = str(code).strip()
    for a in get_all_accounts():
        if str(a["code"]) == code:
            return a
    return None


def get_accounts_by_type(atype):
    return [a for a in get_all_accounts() if a["type"] == atype]


def add_account(code, name, atype, is_system=False):
    code = str(code).strip()
    if not code or not str(name).strip():
        raise ValueError("Account code and name are required.")
    if atype not in ("asset", "liability", "equity", "income", "expense"):
        raise ValueError("Invalid account type.")
    if get_account_by_code(code):
        raise ValueError(f"Account code {code} already exists.")
    with _lock:
        wb = _open()
        ws = wb["Accounts"]
        aid = _next_id(ws)
        ws.append([aid, code, str(name).strip(), atype, bool(is_system), datetime.now()])
        _save(wb)
        wb.close()
    return aid


def update_account_name(account_id, name):
    aid = int(account_id)
    name = str(name).strip()
    if not name:
        raise ValueError("Name is required.")
    with _lock:
        wb = _open()
        ws = wb["Accounts"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == aid:
                if bool(row[4].value):
                    wb.close()
                    raise ValueError("System accounts cannot be renamed.")
                row[2].value = name  # name column
                _save(wb)
                wb.close()
                return
        wb.close()
    raise ValueError("Account not found.")


def next_expense_code():
    """Next available expense account code in the 5xxx+ range."""
    codes = [int(a["code"]) for a in get_accounts_by_type("expense")
             if str(a["code"]).isdigit()]
    base = max(codes) if codes else 5000
    code = base + 10
    existing = {str(a["code"]) for a in get_all_accounts()}
    while str(code) in existing:
        code += 10
    return str(code)


def post_journal(description, lines, source_type="manual", source_id=None,
                 created_by="system", entry_date=None):
    """Post a balanced double-entry journal entry.
    lines: list of {account_id, debit, credit}. Sum(debit) must equal sum(credit).
    Returns the new entry_id. Raises ValueError if unbalanced or invalid.
    """
    if not lines or len(lines) < 2:
        raise ValueError("A journal entry needs at least two lines.")
    norm = []
    total_debit = 0.0
    total_credit = 0.0
    for ln in lines:
        aid = int(ln["account_id"])
        debit = round(float(ln.get("debit", 0) or 0), 2)
        credit = round(float(ln.get("credit", 0) or 0), 2)
        if debit < 0 or credit < 0:
            raise ValueError("Debit/credit cannot be negative.")
        if debit > 0 and credit > 0:
            raise ValueError("A line cannot have both debit and credit.")
        if debit == 0 and credit == 0:
            continue
        norm.append((aid, debit, credit))
        total_debit += debit
        total_credit += credit
    if not norm:
        raise ValueError("Journal entry has no non-zero lines.")
    if round(total_debit - total_credit, 2) != 0:
        raise ValueError(
            f"Entry not balanced: debits {total_debit:.2f} != credits {total_credit:.2f}")

    when = entry_date or datetime.now()
    with _lock:
        wb = _open()
        for sheet, headers in (("Accounts", ACCOUNT_HEADERS),
                               ("JournalEntries", JOURNAL_HEADERS),
                               ("JournalLines", JOURNAL_LINE_HEADERS)):
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet).append(headers)
        ws_je = wb["JournalEntries"]
        ws_jl = wb["JournalLines"]
        # validate account ids exist
        valid_ids = set()
        for row in wb["Accounts"].iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                valid_ids.add(int(row[0]))
        for aid, _, _ in norm:
            if aid not in valid_ids:
                wb.close()
                raise ValueError(f"Account id {aid} not found.")

        entry_id = _next_id(ws_je)
        ws_je.append([entry_id, when, str(description).strip(),
                      source_type, source_id, created_by, datetime.now()])
        line_id = _next_id(ws_jl)
        for aid, debit, credit in norm:
            ws_jl.append([line_id, entry_id, aid, debit, credit])
            line_id += 1
        _save(wb)
        wb.close()
    return entry_id


def get_journal_lines(entry_id=None):
    eid = int(entry_id) if entry_id is not None else None
    with _lock:
        wb = _open()
        if "JournalLines" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["JournalLines"]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = _row_to_dict(JOURNAL_LINE_HEADERS, row)
            if eid is not None and int(d["entry_id"]) != eid:
                continue
            rows.append(d)
        wb.close()
    return rows


def get_account_balance(account_id):
    """Signed balance in the account's normal direction.
    Debit-normal (asset/expense): debit - credit. Credit-normal: credit - debit."""
    aid = int(account_id)
    acct = get_account(aid)
    if not acct:
        raise ValueError("Account not found.")
    debit = 0.0
    credit = 0.0
    for ln in get_journal_lines():
        if int(ln["account_id"]) != aid:
            continue
        debit += float(ln["debit"] or 0)
        credit += float(ln["credit"] or 0)
    if acct["type"] in _DEBIT_NORMAL_TYPES:
        return round(debit - credit, 2)
    return round(credit - debit, 2)


def get_trial_balance():
    """Return per-account totals and the grand debit/credit totals.
    For a correct ledger, total_debit == total_credit."""
    accounts = {a["account_id"]: a for a in get_all_accounts()}
    sums = {aid: {"debit": 0.0, "credit": 0.0} for aid in accounts}
    for ln in get_journal_lines():
        aid = int(ln["account_id"])
        if aid not in sums:
            sums[aid] = {"debit": 0.0, "credit": 0.0}
        sums[aid]["debit"] += float(ln["debit"] or 0)
        sums[aid]["credit"] += float(ln["credit"] or 0)
    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for aid, acct in accounts.items():
        d = round(sums[aid]["debit"], 2)
        c = round(sums[aid]["credit"], 2)
        if d == 0 and c == 0:
            continue
        if acct["type"] in _DEBIT_NORMAL_TYPES:
            bal = round(d - c, 2)
        else:
            bal = round(c - d, 2)
        rows.append({"account": acct, "debit": d, "credit": c, "balance": bal})
        total_debit += d
        total_credit += c
    rows.sort(key=lambda r: str(r["account"]["code"]))
    return rows, {"debit": round(total_debit, 2), "credit": round(total_credit, 2)}


# ── Expenses (built on the journal) ──────────────────────────────────

def record_expense(expense_account_id, amount, paid_from_account_id,
                   description="", created_by="system", entry_date=None):
    """Record an operating expense paid from a cash/bank account.
    Posts: Dr Expense, Cr Cash/Bank."""
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    exp = get_account(int(expense_account_id))
    if not exp or exp["type"] != "expense":
        raise ValueError("Select a valid expense category.")
    src = get_account(int(paid_from_account_id))
    if not src or src["type"] != "asset":
        raise ValueError("Select a valid account to pay from (Cash/Bank).")
    desc = (description or "").strip() or f"Expense: {exp['name']}"
    return post_journal(
        desc,
        [{"account_id": exp["account_id"], "debit": amount},
         {"account_id": src["account_id"], "credit": amount}],
        source_type="expense", created_by=created_by, entry_date=entry_date)


def get_expense_entries(start_date=None, end_date=None):
    """List expense journal entries with amount, category, and paid-from account."""
    accounts = {a["account_id"]: a for a in get_all_accounts()}
    entries = []
    with _lock:
        wb = _open()
        if "JournalEntries" not in wb.sheetnames:
            wb.close()
            return []
        je = {}
        for row in wb["JournalEntries"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = _row_to_dict(JOURNAL_HEADERS, row)
            if d.get("source_type") != "expense":
                continue
            je[int(d["entry_id"])] = d
        lines_by_entry = {}
        for row in wb["JournalLines"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = _row_to_dict(JOURNAL_LINE_HEADERS, row)
            eid = int(d["entry_id"])
            if eid in je:
                lines_by_entry.setdefault(eid, []).append(d)
        wb.close()

    for eid, entry in je.items():
        exp_acct = None
        src_acct = None
        amount = 0.0
        for ln in lines_by_entry.get(eid, []):
            aid = int(ln["account_id"])
            debit = float(ln["debit"] or 0)
            credit = float(ln["credit"] or 0)
            if debit > 0:
                exp_acct = accounts.get(aid)
                amount = debit
            elif credit > 0:
                src_acct = accounts.get(aid)
        edate = entry.get("date")
        edd = edate.date() if isinstance(edate, datetime) else edate
        if start_date and edd and edd < start_date:
            continue
        if end_date and edd and edd > end_date:
            continue
        entries.append({
            "entry_id": eid,
            "date": edate,
            "description": entry.get("description"),
            "category": exp_acct,
            "paid_from": src_acct,
            "amount": round(amount, 2),
            "created_by": entry.get("created_by"),
        })
    entries.sort(key=lambda x: x["entry_id"], reverse=True)
    return entries


def delete_expense(entry_id):
    eid = int(entry_id)
    with _lock:
        wb = _open()
        if "JournalEntries" not in wb.sheetnames:
            wb.close()
            raise ValueError("Expense entry not found.")
        ws_je = wb["JournalEntries"]
        row_idx = None
        for idx, row in enumerate(ws_je.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if int(row[0].value) == eid and row[3].value == "expense":
                row_idx = idx
                break
        if row_idx is None:
            wb.close()
            raise ValueError("Expense entry not found.")
        if "JournalLines" in wb.sheetnames:
            ws_jl = wb["JournalLines"]
            del_rows = [i for i, row in enumerate(ws_jl.iter_rows(min_row=2), start=2)
                        if row[0].value is not None and int(row[1].value) == eid]
            for idx in sorted(del_rows, reverse=True):
                ws_jl.delete_rows(idx, 1)
        ws_je.delete_rows(row_idx, 1)
        _save(wb)
        wb.close()


# ── Journal sync: post accounting entries from operational data ───────

def _existing_journal_sources():
    """Set of (source_type, str(source_id)) already posted to the journal."""
    out = set()
    with _lock:
        wb = _open()
        if "JournalEntries" not in wb.sheetnames:
            wb.close()
            return out
        for row in wb["JournalEntries"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = _row_to_dict(JOURNAL_HEADERS, row)
            if d.get("source_id") is not None:
                out.add((str(d.get("source_type")), str(int(d["source_id"]))))
        wb.close()
    return out


def _realized_credit_invoice_ids():
    """Credit invoices considered paid: a customer's payments cover them oldest
    first. Returns (realized_set, breakdown) where breakdown maps invoice_id ->
    {"cash": x, "bank": y} showing which account(s) actually received the money,
    based on the payment_method recorded on the covering CreditLedger entries."""
    payments_by_cust = {}
    for e in get_credit_ledger():
        if e.get("type") == "credit":
            cid = normalize_customer_id(e.get("customer_id"))
            if cid is not None:
                pm = str(e.get("payment_method") or "cash").strip().lower()
                pm = pm if pm == "bank" else "cash"
                payments_by_cust.setdefault(cid, []).append({
                    "remaining": float(e.get("amount") or 0),
                    "method": pm,
                    "entry_id": int(e["entry_id"]),
                })
    credit_by_cust = {}
    for inv in get_all_invoices():
        if str(inv.get("payment_method") or "").strip().lower() == "credit":
            cid = normalize_customer_id(inv.get("customer_id"))
            if cid is not None:
                credit_by_cust.setdefault(cid, []).append(inv)

    realized = set()
    breakdown = {}
    for cid, invs in credit_by_cust.items():
        invs.sort(key=lambda x: int(x["invoice_id"]))
        payments = sorted(payments_by_cust.get(cid, []), key=lambda p: p["entry_id"])
        ppos = 0
        for inv in invs:
            need = round(float(inv.get("total") or 0), 2)
            alloc = {"cash": 0.0, "bank": 0.0}
            while need > 0.0001 and ppos < len(payments):
                pay = payments[ppos]
                take = min(need, pay["remaining"])
                alloc[pay["method"]] = round(alloc[pay["method"]] + take, 2)
                pay["remaining"] = round(pay["remaining"] - take, 2)
                need = round(need - take, 2)
                if pay["remaining"] <= 0.0001:
                    ppos += 1
            if need <= 0.0001:
                iid = int(inv["invoice_id"])
                realized.add(iid)
                breakdown[iid] = alloc
            else:
                break
    return realized, breakdown


def sync_journal_from_operations():
    """Generate journal entries for sales, purchases and supplier payments not yet
    posted. Credit sales are recorded only once the customer has paid for them
    (cash basis for credit). Idempotent. Returns count posted."""
    seed_chart_of_accounts()  # auto-heal a missing/partial chart of accounts
    acc = {str(a["code"]): a["account_id"] for a in get_all_accounts()}
    CASH, BANK, AR, INV, AP = acc["1000"], acc["1010"], acc["1100"], acc["1200"], acc["2000"]
    TAX, SALES, COGS = acc["2100"], acc["4000"], acc["5000"]
    existing = _existing_journal_sources()
    books_start = get_books_start()

    def _after_start(val):
        if books_start is None:
            return True
        d = val.date() if isinstance(val, datetime) else val
        return d is None or d >= books_start

    realized, realized_breakdown = _realized_credit_invoice_ids()
    pending = []  # (date, description, source_type, source_id, [lines])

    # Sales: cash sales post at once; a credit sale posts only once it is paid.
    for inv in get_all_invoices():
        sid = int(inv["invoice_id"])
        if ("sale", str(sid)) in existing:
            continue
        if not _after_start(inv.get("created_at")):
            continue
        is_credit = str(inv.get("payment_method") or "").strip().lower() == "credit"
        if is_credit and sid not in realized:
            continue  # unpaid credit sale -> stays off the balance sheet
        items = get_invoice_items(sid)
        cogs = round(sum(float(it["purchase_price"] or 0) * int(it["quantity"] or 0)
                         for it in items), 2)
        net = round(float(inv.get("subtotal") or 0) - float(inv.get("discount_total") or 0), 2)
        tax = round(float(inv.get("tax_amount") or 0), 2)
        delivery = round(float(inv.get("delivery_charges") or 0), 2)
        total = round(float(inv.get("total") or 0), 2)
        lines = []
        if is_credit:
            # Realized credit sale: split the receipt across whichever
            # account(s) the customer's covering payments actually went to.
            alloc = realized_breakdown.get(sid, {"cash": total, "bank": 0.0})
            if alloc["cash"] > 0:
                lines.append({"account_id": CASH, "debit": alloc["cash"]})
            if alloc["bank"] > 0:
                lines.append({"account_id": BANK, "debit": alloc["bank"]})
        else:
            pm = str(inv.get("payment_method") or "cash").strip().lower()
            recv_acct = BANK if pm == "bank" else CASH
            lines.append({"account_id": recv_acct, "debit": total})
        lines.append({"account_id": SALES, "credit": net + delivery})
        if tax > 0:
            lines.append({"account_id": TAX, "credit": tax})
        if cogs > 0:
            lines.append({"account_id": COGS, "debit": cogs})
            lines.append({"account_id": INV, "credit": cogs})
        pending.append((inv.get("created_at"), f"Sale INV-{sid}", "sale", sid, lines))

    # Purchases (stock received). Dr Inventory; credit AP (unpaid) or Cash/Bank (paid).
    _pay_acct = {"credit": AP, "cash": CASH, "bank": BANK}
    for p in get_all_purchase_invoices():
        pid = int(p["purchase_id"])
        if ("purchase", str(pid)) in existing:
            continue
        if not _after_start(p.get("created_at")):
            continue
        total = round(float(p.get("total_amount") or 0), 2)
        if total <= 0:
            continue
        pm = str(p.get("payment_method") or "credit").strip().lower()
        credit_acct = _pay_acct.get(pm, AP)
        pending.append((p.get("created_at"), f"Purchase PINV-{pid}", "purchase", pid,
                        [{"account_id": INV, "debit": total},
                         {"account_id": credit_acct, "credit": total}]))

    # Supplier payments -> Dr Accounts Payable, Cr Cash. Skip credits tied to a
    # purchase (a cash/bank purchase already credited Cash/Bank in its own entry).
    for e in get_supplier_ledger_entries():
        if e.get("type") != "credit":
            continue
        if e.get("purchase_id") not in (None, ""):
            continue
        eid = int(e["entry_id"])
        if ("supplier_payment", str(eid)) in existing:
            continue
        if not _after_start(e.get("created_at")):
            continue
        amt = round(float(e.get("amount") or 0), 2)
        if amt <= 0:
            continue
        pm = (e.get("payment_method") or "cash").strip().lower()
        cash_acct = BANK if pm == "bank" else CASH
        pending.append((e.get("created_at"), "Supplier payment", "supplier_payment", eid,
                        [{"account_id": AP, "debit": amt},
                         {"account_id": cash_acct, "credit": amt}]))

    # Customer credit repayments are NOT posted separately: the cash for a credit
    # sale enters the books when that sale is recognised (above), keeping the
    # balance sheet on a cash basis for credit.

    if not pending:
        return 0

    # Batch-append everything in one workbook open
    with _lock:
        wb = _open()
        for sheet, headers in (("JournalEntries", JOURNAL_HEADERS),
                               ("JournalLines", JOURNAL_LINE_HEADERS)):
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet).append(headers)
        ws_je = wb["JournalEntries"]
        ws_jl = wb["JournalLines"]
        entry_id = _next_id(ws_je)
        line_id = _next_id(ws_jl)
        now = datetime.now()
        for when, desc, stype, sid, lines in pending:
            # safety: only post balanced entries
            td = round(sum(float(l.get("debit", 0) or 0) for l in lines), 2)
            tc = round(sum(float(l.get("credit", 0) or 0) for l in lines), 2)
            if td != tc:
                continue
            ws_je.append([entry_id, when, desc, stype, sid, "system", now])
            for l in lines:
                ws_jl.append([line_id, entry_id, int(l["account_id"]),
                              round(float(l.get("debit", 0) or 0), 2),
                              round(float(l.get("credit", 0) or 0), 2)])
                line_id += 1
            entry_id += 1
        _save(wb)
        wb.close()
    return len(pending)


def get_balance_sheet():
    """Balance sheet derived from the journal. Always balances because the
    underlying ledger is double-entry. Returns dict of sections + totals."""
    rows, _ = get_trial_balance()
    sections = {"asset": [], "liability": [], "equity": []}
    total = {"asset": 0.0, "liability": 0.0, "equity": 0.0}
    income = 0.0
    expense = 0.0
    for r in rows:
        atype = r["account"]["type"]
        bal = r["balance"]
        if atype in sections:
            sections[atype].append({"name": r["account"]["name"], "balance": bal})
            total[atype] += bal
        elif atype == "income":
            income += bal
        elif atype == "expense":
            expense += bal
    net_income = round(income - expense, 2)
    # Net income rolls into equity
    equity_rows = list(sections["equity"])
    equity_rows.append({"name": "Net Income (current)", "balance": net_income})
    total_equity = round(total["equity"] + net_income, 2)
    return {
        "assets": sections["asset"],
        "liabilities": sections["liability"],
        "equity": equity_rows,
        "total_assets": round(total["asset"], 2),
        "total_liabilities": round(total["liability"], 2),
        "total_equity": total_equity,
        "total_liab_equity": round(total["liability"] + total_equity, 2),
        "net_income": net_income,
    }


# ── Opening balances & books start date ──────────────────────────────

def get_books_start():
    """Date of the opening-balance entry, or None if not set."""
    with _lock:
        wb = _open()
        if "JournalEntries" not in wb.sheetnames:
            wb.close()
            return None
        result = None
        for row in wb["JournalEntries"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = _row_to_dict(JOURNAL_HEADERS, row)
            if d.get("source_type") == "opening":
                dt = d.get("date")
                result = dt.date() if isinstance(dt, datetime) else dt
                break
        wb.close()
    return result


def get_opening_balances():
    """Return the opening amounts keyed by account code, or {} if not set."""
    with _lock:
        wb = _open()
        if "JournalEntries" not in wb.sheetnames:
            wb.close()
            return {}
        opening_ids = set()
        for row in wb["JournalEntries"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            d = _row_to_dict(JOURNAL_HEADERS, row)
            if d.get("source_type") == "opening":
                opening_ids.add(int(d["entry_id"]))
        by_acct = {}
        if opening_ids and "JournalLines" in wb.sheetnames:
            for row in wb["JournalLines"].iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                d = _row_to_dict(JOURNAL_LINE_HEADERS, row)
                if int(d["entry_id"]) in opening_ids:
                    by_acct[int(d["account_id"])] = (
                        float(d["debit"] or 0), float(d["credit"] or 0))
        wb.close()
    accts = {a["account_id"]: a for a in get_all_accounts()}
    out = {}
    for aid, (dr, cr) in by_acct.items():
        acct = accts.get(aid)
        if not acct:
            continue
        amount = dr if acct["type"] in _DEBIT_NORMAL_TYPES else cr
        out[str(acct["code"])] = round(amount, 2)
    return out


_AUTO_SYNCED_SOURCES = ("sale", "purchase", "supplier_payment", "customer_payment")


def _delete_opening_entries(wb, source_types=("opening",)):
    """Remove journal entries (+ their lines) of the given source types."""
    if "JournalEntries" not in wb.sheetnames:
        return
    ws_je = wb["JournalEntries"]
    target_ids = set()
    del_rows = []
    for idx, row in enumerate(ws_je.iter_rows(min_row=2), start=2):
        if row[0].value is None:
            continue
        if row[3].value in source_types:  # source_type column
            target_ids.add(int(row[0].value))
            del_rows.append(idx)
    for idx in sorted(del_rows, reverse=True):
        ws_je.delete_rows(idx, 1)
    if target_ids and "JournalLines" in wb.sheetnames:
        ws_jl = wb["JournalLines"]
        del_lines = []
        for idx, row in enumerate(ws_jl.iter_rows(min_row=2), start=2):
            if row[0].value is None:
                continue
            if int(row[1].value) in target_ids:  # entry_id column
                del_lines.append(idx)
        for idx in sorted(del_lines, reverse=True):
            ws_jl.delete_rows(idx, 1)


def clear_opening_balances():
    """Remove opening balances + auto-synced entries (next view re-syncs fresh)."""
    with _lock:
        wb = _open()
        _delete_opening_entries(wb, ("opening",) + _AUTO_SYNCED_SOURCES)
        _save(wb)
        wb.close()


def set_opening_balances(start_date, balances, created_by="system"):
    """Set/replace opening balances as of start_date.
    balances: dict of account code -> amount, e.g. {'1000': cash, '1200': inventory, ...}.
    Owner Capital (3000) is computed as the balancing figure."""
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    seed_chart_of_accounts()  # auto-heal a missing/partial chart of accounts
    acc = {str(a["code"]): a for a in get_all_accounts()}

    total_assets = 0.0
    total_liab = 0.0
    lines = []
    for code, amount in balances.items():
        amount = round(float(amount or 0), 2)
        if amount == 0:
            continue
        a = acc.get(str(code))
        if not a or str(code) == "3000":
            continue
        if a["type"] in _DEBIT_NORMAL_TYPES:
            lines.append({"account_id": a["account_id"], "debit": amount})
            total_assets += amount
        else:
            lines.append({"account_id": a["account_id"], "credit": amount})
            total_liab += amount

    capital = round(total_assets - total_liab, 2)
    cap_acct = acc["3000"]["account_id"]
    if capital >= 0:
        lines.append({"account_id": cap_acct, "credit": capital})
    else:
        lines.append({"account_id": cap_acct, "debit": -capital})

    if len([l for l in lines if l.get("debit") or l.get("credit")]) < 2:
        raise ValueError("Enter at least one opening balance.")

    with _lock:
        wb = _open()
        for sheet, headers in (("JournalEntries", JOURNAL_HEADERS),
                               ("JournalLines", JOURNAL_LINE_HEADERS)):
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet).append(headers)
        # Replace opening entry and clear auto-synced operational entries so the
        # next sync re-posts only transactions on/after the new start date.
        _delete_opening_entries(wb, ("opening",) + _AUTO_SYNCED_SOURCES)
        ws_je = wb["JournalEntries"]
        ws_jl = wb["JournalLines"]
        entry_id = _next_id(ws_je)
        line_id = _next_id(ws_jl)
        ws_je.append([entry_id, datetime.combine(start_date, datetime.min.time()),
                      "Opening balances", "opening", None, created_by, datetime.now()])
        for l in lines:
            ws_jl.append([line_id, entry_id, int(l["account_id"]),
                          round(float(l.get("debit", 0) or 0), 2),
                          round(float(l.get("credit", 0) or 0), 2)])
            line_id += 1
        _save(wb)
        wb.close()
    return entry_id


# ── Partners / investor equity ───────────────────────────────────────

def get_all_partners(include_inactive=False):
    with _lock:
        wb = _open()
        if "Partners" not in wb.sheetnames:
            wb.close()
            return []
        rows = []
        for row in wb["Partners"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            p = _row_to_dict(PARTNER_HEADERS, row)
            p["partner_id"] = int(p["partner_id"])
            p["share_pct"] = float(p["share_pct"] or 0)
            p["is_active"] = bool(p["is_active"])
            if not include_inactive and not p["is_active"]:
                continue
            rows.append(p)
        wb.close()
    rows.sort(key=lambda x: x["partner_id"])
    return rows


def get_partner(partner_id):
    pid = int(partner_id)
    for p in get_all_partners(include_inactive=True):
        if p["partner_id"] == pid:
            return p
    return None


def add_partner(name, share_pct):
    name = str(name).strip()
    if not name:
        raise ValueError("Partner name is required.")
    share = round(float(share_pct or 0), 2)
    if share < 0 or share > 100:
        raise ValueError("Share % must be between 0 and 100.")
    with _lock:
        wb = _open()
        if "Partners" not in wb.sheetnames:
            wb.create_sheet("Partners").append(PARTNER_HEADERS)
        ws = wb["Partners"]
        pid = _next_id(ws)
        ws.append([pid, name, share, True, datetime.now()])
        _save(wb)
        wb.close()
    return pid


def update_partner(partner_id, name, share_pct, is_active=True):
    pid = int(partner_id)
    name = str(name).strip()
    if not name:
        raise ValueError("Partner name is required.")
    share = round(float(share_pct or 0), 2)
    if share < 0 or share > 100:
        raise ValueError("Share % must be between 0 and 100.")
    with _lock:
        wb = _open()
        ws = wb["Partners"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value is not None and int(row[0].value) == pid:
                row[1].value = name
                row[2].value = share
                row[3].value = bool(is_active)
                _save(wb)
                wb.close()
                return
        wb.close()
    raise ValueError("Partner not found.")


def add_partner_transaction(partner_id, txn_type, amount, note="", txn_date=None,
                            created_by="system"):
    """Record a capital contribution or a drawing for a partner, and post the
    matching journal entry. txn_type: 'capital' or 'drawing'."""
    pid = int(partner_id)
    if txn_type not in ("capital", "drawing"):
        raise ValueError("Type must be 'capital' or 'drawing'.")
    amount = round(float(amount or 0), 2)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_partner(pid):
        raise ValueError("Partner not found.")
    when = txn_date or date.today()
    if isinstance(when, str):
        when = date.fromisoformat(when)

    with _lock:
        wb = _open()
        if "PartnerTransactions" not in wb.sheetnames:
            wb.create_sheet("PartnerTransactions").append(PARTNER_TXN_HEADERS)
        ws = wb["PartnerTransactions"]
        txn_id = _next_id(ws)
        ws.append([txn_id, pid, txn_type, amount, (note or "").strip(),
                   datetime.combine(when, datetime.min.time()), datetime.now()])
        _save(wb)
        wb.close()

    acc = {str(a["code"]): a["account_id"] for a in get_all_accounts()}
    if txn_type == "capital":
        lines = [{"account_id": acc["1000"], "debit": amount},
                 {"account_id": acc["3000"], "credit": amount}]
        desc = "Partner capital contribution"
        stype = "partner_capital"
    else:
        lines = [{"account_id": acc["3100"], "debit": amount},
                 {"account_id": acc["1000"], "credit": amount}]
        desc = "Partner drawing"
        stype = "partner_drawing"
    post_journal(desc, lines, source_type=stype, source_id=txn_id,
                 created_by=created_by,
                 entry_date=datetime.combine(when, datetime.min.time()))
    return txn_id


def get_partner_transactions(partner_id=None):
    pid = int(partner_id) if partner_id is not None else None
    with _lock:
        wb = _open()
        if "PartnerTransactions" not in wb.sheetnames:
            wb.close()
            return []
        rows = []
        for row in wb["PartnerTransactions"].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            t = _row_to_dict(PARTNER_TXN_HEADERS, row)
            if pid is not None and int(t["partner_id"]) != pid:
                continue
            rows.append(t)
        wb.close()
    rows.sort(key=lambda x: int(x["txn_id"]), reverse=True)
    return rows


def get_partner_equity():
    """Per-partner equity = capital contributed + share of net profit - drawings.
    Net profit (since books start) is taken from the balance sheet and split by share %."""
    partners = get_all_partners()
    net_income = get_balance_sheet()["net_income"]
    txns = get_partner_transactions()
    cap_by = {}
    draw_by = {}
    for t in txns:
        p = int(t["partner_id"])
        amt = float(t["amount"] or 0)
        if t["type"] == "capital":
            cap_by[p] = cap_by.get(p, 0.0) + amt
        elif t["type"] == "drawing":
            draw_by[p] = draw_by.get(p, 0.0) + amt
    rows = []
    total_share = round(sum(p["share_pct"] for p in partners), 2)
    for p in partners:
        pid = p["partner_id"]
        cap = round(cap_by.get(pid, 0.0), 2)
        draw = round(draw_by.get(pid, 0.0), 2)
        profit_share = round(net_income * (p["share_pct"] / 100.0), 2)
        equity = round(cap + profit_share - draw, 2)
        rows.append({
            "partner": p,
            "capital": cap,
            "share_pct": p["share_pct"],
            "profit_share": profit_share,
            "drawings": draw,
            "equity": equity,
        })
    totals = {
        "capital": round(sum(r["capital"] for r in rows), 2),
        "profit_share": round(sum(r["profit_share"] for r in rows), 2),
        "drawings": round(sum(r["drawings"] for r in rows), 2),
        "equity": round(sum(r["equity"] for r in rows), 2),
        "share_pct": total_share,
        "net_income": net_income,
    }
    return rows, totals
