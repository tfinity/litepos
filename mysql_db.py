"""MySQL data access layer for the POS system — identical interface to excel_db.py."""

from contextlib import contextmanager
from datetime import datetime, date, timedelta
import re

import pymysql
import pymysql.cursors

from config import MYSQL_HOST, MYSQL_PORT, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE
from tenant import get_current_tenant, require_tenant


def _tid():
    """Active tenant id for the current request; raises if none is set."""
    return require_tenant()

# Standard chart of accounts (code, name, type, is_system)
CHART_OF_ACCOUNTS = [
    ("1000", "Cash",                "asset",     True),
    ("1010", "Bank",                "asset",     True),
    ("1100", "Accounts Receivable", "asset",     True),
    ("1200", "Inventory",           "asset",     True),
    ("2000", "Accounts Payable",    "liability", True),
    ("2100", "Tax Payable",         "liability", True),
    ("3000", "Owner Capital",       "equity",    True),
    ("3100", "Owner Drawings",      "equity",    True),
    ("3900", "Retained Earnings",   "equity",    True),
    ("4000", "Sales Revenue",       "income",    True),
    ("5000", "Cost of Goods Sold",  "expense",   True),
    ("5100", "Rent",                "expense",   False),
    ("5200", "Salaries",            "expense",   False),
    ("5300", "Utilities",           "expense",   False),
    ("5400", "Transport",           "expense",   False),
    ("5900", "Miscellaneous",       "expense",   False),
]
_DEBIT_NORMAL_TYPES = ("asset", "expense")


@contextmanager
def _conn():
    conn = pymysql.connect(
        host=MYSQL_HOST,
        port=MYSQL_PORT,
        user=MYSQL_USER,
        password=MYSQL_PASSWORD,
        database=MYSQL_DATABASE,
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ── Schema init ───────────────────────────────────────────────────────

def init_workbook():
    """Create all tables if they don't exist."""
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS products (
                    product_id    INT AUTO_INCREMENT PRIMARY KEY,
                    name          VARCHAR(255) NOT NULL,
                    purchase_price DECIMAL(12,2) DEFAULT 0,
                    counter_price  DECIMAL(12,2) DEFAULT 0,
                    retail_price   DECIMAL(12,2) DEFAULT 0,
                    quantity       INT DEFAULT 0,
                    barcode        VARCHAR(100),
                    expiry_date    DATE,
                    category         VARCHAR(100),
                    created_at       DATETIME DEFAULT CURRENT_TIMESTAMP,
                    last_supplier_id INT
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS customers (
                    customer_id INT AUTO_INCREMENT PRIMARY KEY,
                    name        VARCHAR(255) NOT NULL,
                    phone       VARCHAR(50),
                    email       VARCHAR(100),
                    address     TEXT,
                    tax_id      VARCHAR(100),
                    notes       TEXT,
                    created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoices (
                    invoice_id     INT AUTO_INCREMENT PRIMARY KEY,
                    created_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                    subtotal       DECIMAL(12,2),
                    discount_total DECIMAL(12,2),
                    tax_rate       DECIMAL(6,4),
                    tax_amount     DECIMAL(12,2),
                    total          DECIMAL(12,2),
                    payment_method VARCHAR(50),
                    customer_id    INT,
                    status         VARCHAR(20) DEFAULT 'active',
                    deleted_at     DATETIME,
                    deleted_by     VARCHAR(100),
                    delete_reason  TEXT,
                    FOREIGN KEY (customer_id) REFERENCES customers(customer_id) ON DELETE SET NULL
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            for _col in [
                "ALTER TABLE invoices ADD COLUMN status VARCHAR(20) DEFAULT 'active'",
                "ALTER TABLE invoices ADD COLUMN deleted_at DATETIME",
                "ALTER TABLE invoices ADD COLUMN deleted_by VARCHAR(100)",
                "ALTER TABLE invoices ADD COLUMN delete_reason TEXT",
                "ALTER TABLE invoices ADD COLUMN delivery_charges DECIMAL(12,2) DEFAULT 0",
            ]:
                try:
                    cur.execute(_col)
                except Exception:
                    pass  # column already exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoice_items (
                    item_id         INT AUTO_INCREMENT PRIMARY KEY,
                    invoice_id      INT NOT NULL,
                    product_id      INT,
                    product_name    VARCHAR(255),
                    purchase_price  DECIMAL(12,2),
                    counter_price   DECIMAL(12,2),
                    quantity        INT,
                    discount_amount DECIMAL(12,2),
                    line_total      DECIMAL(12,2),
                    FOREIGN KEY (invoice_id) REFERENCES invoices(invoice_id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            try:
                cur.execute("ALTER TABLE invoice_items ADD COLUMN batch_id INT")
            except Exception:
                pass  # column already exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tenants (
                    tenant_id  INT AUTO_INCREMENT PRIMARY KEY,
                    name       VARCHAR(255) NOT NULL,
                    is_active  TINYINT(1) DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id       INT AUTO_INCREMENT PRIMARY KEY,
                    tenant_id     INT,
                    username      VARCHAR(100) NOT NULL UNIQUE,
                    password_hash VARCHAR(255) NOT NULL,
                    role          ENUM('super_admin','admin','staff') DEFAULT 'staff',
                    is_active     TINYINT(1) DEFAULT 1,
                    created_at    DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            # Upgrade existing users table for multi-tenant
            try:
                cur.execute("ALTER TABLE users ADD COLUMN tenant_id INT")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE users MODIFY role ENUM('super_admin','admin','staff') DEFAULT 'staff'")
            except Exception:
                pass
            cur.execute("""
                CREATE TABLE IF NOT EXISTS credit_ledger (
                    entry_id    INT AUTO_INCREMENT PRIMARY KEY,
                    customer_id INT,
                    invoice_id  INT,
                    type        VARCHAR(20) NOT NULL,
                    amount      DECIMAL(12,2),
                    note        TEXT,
                    created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            try:
                cur.execute("ALTER TABLE credit_ledger MODIFY type VARCHAR(20) NOT NULL")
            except Exception:
                pass  # already widened
            try:
                cur.execute("ALTER TABLE credit_ledger ADD COLUMN payment_method VARCHAR(20)")
            except Exception:
                pass  # column already exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS suppliers (
                    supplier_id INT AUTO_INCREMENT PRIMARY KEY,
                    name        VARCHAR(255) NOT NULL,
                    phone       VARCHAR(50),
                    email       VARCHAR(100),
                    address     TEXT,
                    notes       TEXT,
                    created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS purchase_invoices (
                    purchase_id    INT AUTO_INCREMENT PRIMARY KEY,
                    supplier_id    INT NOT NULL,
                    created_at     DATETIME DEFAULT CURRENT_TIMESTAMP,
                    total_amount   DECIMAL(12,2),
                    notes          TEXT,
                    payment_method VARCHAR(20) DEFAULT 'credit',
                    FOREIGN KEY (supplier_id) REFERENCES suppliers(supplier_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            try:
                cur.execute("ALTER TABLE purchase_invoices ADD COLUMN payment_method VARCHAR(20) DEFAULT 'credit'")
            except Exception:
                pass  # column already exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS purchase_invoice_items (
                    item_id      INT AUTO_INCREMENT PRIMARY KEY,
                    purchase_id  INT NOT NULL,
                    product_id   INT,
                    product_name VARCHAR(255),
                    quantity     INT,
                    unit_cost    DECIMAL(12,2),
                    line_total   DECIMAL(12,2),
                    FOREIGN KEY (purchase_id) REFERENCES purchase_invoices(purchase_id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            try:
                cur.execute("ALTER TABLE purchase_invoice_items ADD COLUMN batch_id INT")
            except Exception:
                pass  # column already exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS product_batches (
                    batch_id      INT AUTO_INCREMENT PRIMARY KEY,
                    tenant_id     INT,
                    product_id    INT NOT NULL,
                    batch_number  VARCHAR(100),
                    supplier_id   INT,
                    purchase_id   INT,
                    unit_cost     DECIMAL(12,2) DEFAULT 0,
                    qty_received  INT DEFAULT 0,
                    qty_remaining INT DEFAULT 0,
                    expiry_date   DATE,
                    received_at   DATETIME DEFAULT CURRENT_TIMESTAMP,
                    created_at    DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (product_id) REFERENCES products(product_id) ON DELETE CASCADE
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            try:
                cur.execute("CREATE INDEX idx_product_batches_tenant_product ON product_batches (tenant_id, product_id)")
            except Exception:
                pass  # index already exists
            cur.execute("""
                CREATE TABLE IF NOT EXISTS supplier_ledger (
                    entry_id    INT AUTO_INCREMENT PRIMARY KEY,
                    supplier_id INT NOT NULL,
                    purchase_id INT,
                    type        VARCHAR(20) NOT NULL,
                    amount      DECIMAL(12,2),
                    note        TEXT,
                    created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            try:
                cur.execute("ALTER TABLE supplier_ledger MODIFY type VARCHAR(20) NOT NULL")
            except Exception:
                pass  # already widened
            try:
                cur.execute("ALTER TABLE supplier_ledger ADD COLUMN payment_method VARCHAR(20) DEFAULT 'cash'")
            except Exception:
                pass  # column already exists
            # Add last_supplier_id to products if not exists
            try:
                cur.execute("ALTER TABLE products ADD COLUMN last_supplier_id INT DEFAULT NULL")
            except Exception:
                pass  # column already exists

            # ── Double-entry accounting ──
            cur.execute("""
                CREATE TABLE IF NOT EXISTS accounts (
                    account_id INT AUTO_INCREMENT PRIMARY KEY,
                    code       VARCHAR(20) NOT NULL UNIQUE,
                    name       VARCHAR(255) NOT NULL,
                    type       ENUM('asset','liability','equity','income','expense') NOT NULL,
                    is_system  TINYINT(1) DEFAULT 0,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS journal_entries (
                    entry_id    INT AUTO_INCREMENT PRIMARY KEY,
                    date        DATETIME NOT NULL,
                    description VARCHAR(500),
                    source_type VARCHAR(50),
                    source_id   INT,
                    created_by  VARCHAR(100),
                    created_at  DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS journal_lines (
                    line_id    INT AUTO_INCREMENT PRIMARY KEY,
                    entry_id   INT NOT NULL,
                    account_id INT NOT NULL,
                    debit      DECIMAL(14,2) DEFAULT 0,
                    credit     DECIMAL(14,2) DEFAULT 0,
                    FOREIGN KEY (entry_id) REFERENCES journal_entries(entry_id) ON DELETE CASCADE,
                    FOREIGN KEY (account_id) REFERENCES accounts(account_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS partners (
                    partner_id INT AUTO_INCREMENT PRIMARY KEY,
                    name       VARCHAR(255) NOT NULL,
                    share_pct  DECIMAL(6,2) DEFAULT 0,
                    is_active  TINYINT(1) DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS partner_transactions (
                    txn_id     INT AUTO_INCREMENT PRIMARY KEY,
                    partner_id INT NOT NULL,
                    type       ENUM('capital','drawing') NOT NULL,
                    amount     DECIMAL(14,2),
                    note       TEXT,
                    date       DATETIME,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (partner_id) REFERENCES partners(partner_id)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
            """)

            # Multi-tenant: every data table carries the owning tenant_id.
            for _tbl in ("products", "customers", "invoices", "invoice_items",
                         "credit_ledger", "suppliers", "purchase_invoices",
                         "purchase_invoice_items", "supplier_ledger", "accounts",
                         "journal_entries", "journal_lines", "partners",
                         "partner_transactions"):
                try:
                    cur.execute(f"ALTER TABLE {_tbl} ADD COLUMN tenant_id INT")
                except Exception:
                    pass  # column already exists
                try:
                    cur.execute(f"CREATE INDEX idx_{_tbl}_tenant ON {_tbl} (tenant_id)")
                except Exception:
                    pass  # index already exists
            # Chart-of-accounts code must be unique per tenant, not globally.
            try:
                cur.execute("ALTER TABLE accounts DROP INDEX code")
            except Exception:
                pass
            try:
                cur.execute("ALTER TABLE accounts ADD UNIQUE KEY uniq_tenant_code (tenant_id, code)")
            except Exception:
                pass


def normalize_customer_id(val):
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _to_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _normalize_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


# ── Products ──────────────────────────────────────────────────────────

def get_all_products():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM products WHERE tenant_id = %s ORDER BY product_id", (_tid(),))
            rows = cur.fetchall()
    for p in rows:
        p["expiry_date"] = _normalize_date(p["expiry_date"])
    return rows


def get_product(product_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM products WHERE product_id = %s AND tenant_id = %s",
                        (int(product_id), _tid()))
            row = cur.fetchone()
    if row:
        row["expiry_date"] = _normalize_date(row["expiry_date"])
    return row


def get_product_by_barcode(barcode):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM products WHERE barcode = %s AND tenant_id = %s",
                        (str(barcode).strip(), _tid()))
            row = cur.fetchone()
    if row:
        row["expiry_date"] = _normalize_date(row["expiry_date"])
    return row


def add_product(data):
    expiry = data.get("expiry_date")
    if isinstance(expiry, str) and expiry:
        expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
    elif not isinstance(expiry, (date, datetime)):
        expiry = None
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO products
                    (tenant_id, name, purchase_price, counter_price, retail_price, quantity, barcode, expiry_date, category, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                _tid(),
                data["name"],
                _to_float(data.get("purchase_price")),
                _to_float(data.get("counter_price")),
                _to_float(data.get("retail_price")),
                _to_int(data.get("quantity")),
                data.get("barcode", ""),
                expiry,
                data.get("category", ""),
                datetime.now(),
            ))
            return cur.lastrowid


def update_product(product_id, data):
    expiry = data.get("expiry_date")
    if isinstance(expiry, str) and expiry:
        expiry = datetime.strptime(expiry, "%Y-%m-%d").date()
    elif not isinstance(expiry, (date, datetime)):
        expiry = None
    pid = int(product_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE products SET
                    name = %s, purchase_price = %s, counter_price = %s,
                    retail_price = %s, quantity = %s, barcode = %s,
                    expiry_date = %s, category = %s
                WHERE product_id = %s AND tenant_id = %s
            """, (
                data["name"],
                _to_float(data.get("purchase_price")),
                _to_float(data.get("counter_price")),
                _to_float(data.get("retail_price")),
                _to_int(data.get("quantity")),
                data.get("barcode", ""),
                expiry,
                data.get("category", ""),
                pid,
                tid,
            ))
            # Quantity/cost are derived from batch stock once a product has
            # batch history -- the form's typed value would otherwise drift
            # out of sync with what invoicing actually sells from. Products
            # with no batches yet keep taking the typed quantity as-is.
            cur.execute(
                "SELECT 1 FROM product_batches WHERE product_id=%s AND tenant_id=%s LIMIT 1",
                (pid, tid)
            )
            if cur.fetchone():
                _recompute_product_cache(cur, tid, pid)


def delete_product(product_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM products WHERE product_id = %s AND tenant_id = %s",
                        (int(product_id), _tid()))


def _ensure_legacy_batches():
    """Auto-heal: give every product that has stock but no batch history a
    single legacy batch reflecting its current (blended) cost/quantity, so
    batch-aware code has something to read for pre-existing stock. Idempotent
    — only creates a batch for a product that doesn't already have one."""
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.product_id, p.purchase_price, p.quantity, p.last_supplier_id,
                       p.expiry_date, p.created_at
                FROM products p
                WHERE p.tenant_id = %s AND p.quantity > 0
                  AND NOT EXISTS (
                      SELECT 1 FROM product_batches b
                      WHERE b.product_id = p.product_id AND b.tenant_id = p.tenant_id
                  )
            """, (tid,))
            missing = cur.fetchall()
            for p in missing:
                cur.execute("""
                    INSERT INTO product_batches
                        (tenant_id, product_id, batch_number, supplier_id, purchase_id,
                         unit_cost, qty_received, qty_remaining, expiry_date, received_at)
                    VALUES (%s, %s, %s, %s, NULL, %s, %s, %s, %s, %s)
                """, (tid, p["product_id"], f"P{p['product_id']}-LEGACY", p.get("last_supplier_id"),
                      p.get("purchase_price") or 0, p.get("quantity") or 0, p.get("quantity") or 0,
                      p.get("expiry_date"), p.get("created_at")))
    return len(missing)


def _recompute_product_cache(cur, tid, product_id):
    """Recompute a product's cached quantity/purchase_price from its batches'
    remaining stock (weighted-average cost). Runs on the caller's open cursor
    so it's part of the same transaction as the batch changes."""
    cur.execute("""
        SELECT COALESCE(SUM(qty_remaining), 0) AS total_qty,
               COALESCE(SUM(qty_remaining * unit_cost), 0) AS total_value
        FROM product_batches WHERE product_id = %s AND tenant_id = %s
    """, (product_id, tid))
    row = cur.fetchone()
    total_qty = int(row["total_qty"] or 0)
    if total_qty > 0:
        avg_cost = round(float(row["total_value"] or 0) / total_qty, 2)
        cur.execute("UPDATE products SET quantity=%s, purchase_price=%s WHERE product_id=%s AND tenant_id=%s",
                    (total_qty, avg_cost, product_id, tid))
    else:
        cur.execute("UPDATE products SET quantity=%s WHERE product_id=%s AND tenant_id=%s",
                    (total_qty, product_id, tid))


def _deplete_batch_for_sale(cur, tid, pid, qty, requested_batch_id, product_name):
    """Pick a batch to sell qty units of product pid from, decrement its
    qty_remaining, and return (batch_id, unit_cost). If requested_batch_id is
    given, sell from exactly that batch (staff picked it at the till). If not,
    auto-pick the oldest active batch that alone covers qty. Raises ValueError
    if there isn't enough stock in the selected (or any single) batch."""
    if requested_batch_id is not None:
        cur.execute("""
            SELECT batch_id, batch_number, unit_cost, qty_remaining FROM product_batches
            WHERE batch_id=%s AND product_id=%s AND tenant_id=%s FOR UPDATE
        """, (requested_batch_id, pid, tid))
        target = cur.fetchone()
        if not target:
            raise ValueError(f"Selected batch not found for '{product_name}'.")
        remaining = int(target["qty_remaining"] or 0)
        if qty > remaining:
            raise ValueError(
                f"Not enough stock in batch '{target['batch_number']}' for '{product_name}': "
                f"requested {qty}, available {remaining}"
            )
        cur.execute("UPDATE product_batches SET qty_remaining = qty_remaining - %s WHERE batch_id=%s",
                    (qty, target["batch_id"]))
        return target["batch_id"], float(target["unit_cost"] or 0)

    cur.execute("""
        SELECT batch_id, unit_cost, qty_remaining FROM product_batches
        WHERE product_id=%s AND tenant_id=%s AND qty_remaining > 0
        ORDER BY received_at, batch_id FOR UPDATE
    """, (pid, tid))
    candidates = cur.fetchall()
    for c in candidates:
        remaining = int(c["qty_remaining"] or 0)
        if remaining >= qty:
            cur.execute("UPDATE product_batches SET qty_remaining = qty_remaining - %s WHERE batch_id=%s",
                        (qty, c["batch_id"]))
            return c["batch_id"], float(c["unit_cost"] or 0)

    total_available = sum(int(c["qty_remaining"] or 0) for c in candidates)
    raise ValueError(
        f"Not enough stock for '{product_name}': requested {qty}, available {total_available}"
    )


def _restore_batch_for_sale(cur, tid, pid, qty, batch_id=None):
    """Return qty units back to the batch they were sold from (or that
    product's legacy batch if batch_id is unknown — pre-batch invoices)."""
    target_id = None
    if batch_id is not None:
        cur.execute("SELECT batch_id FROM product_batches WHERE batch_id=%s AND tenant_id=%s",
                    (batch_id, tid))
        row = cur.fetchone()
        if row:
            target_id = row["batch_id"]
    if target_id is None:
        cur.execute("""
            SELECT batch_id FROM product_batches
            WHERE product_id=%s AND tenant_id=%s AND purchase_id IS NULL
            LIMIT 1
        """, (pid, tid))
        row = cur.fetchone()
        if row:
            target_id = row["batch_id"]
    if target_id is not None:
        cur.execute("UPDATE product_batches SET qty_remaining = qty_remaining + %s WHERE batch_id=%s",
                    (qty, target_id))


def get_product_batches(product_id=None, active_only=False):
    """Return product_batches rows, optionally for one product, optionally
    only those with stock remaining. Auto-heals legacy (pre-batch) stock
    into a synthetic batch on first read."""
    _ensure_legacy_batches()
    with _conn() as conn:
        with conn.cursor() as cur:
            q = "SELECT * FROM product_batches WHERE tenant_id = %s"
            params = [_tid()]
            if product_id is not None:
                q += " AND product_id = %s"
                params.append(int(product_id))
            if active_only:
                q += " AND qty_remaining > 0"
            q += " ORDER BY received_at, batch_id"
            cur.execute(q, params)
            rows = cur.fetchall()
    for b in rows:
        b["expiry_date"] = _normalize_date(b.get("expiry_date"))
    return rows


def get_low_stock_products(threshold=10):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM products WHERE tenant_id = %s AND quantity <= %s ORDER BY quantity",
                        (_tid(), threshold))
            rows = cur.fetchall()
    for p in rows:
        p["expiry_date"] = _normalize_date(p["expiry_date"])
    return rows


def get_expiry_products(days_ahead=30):
    """Batch-level expiry: each batch carries its own expiry date, so a
    product with stock from two different lots can show up twice here with
    different dates/quantities. Only batches with stock remaining are shown."""
    today = date.today()
    cutoff = today + timedelta(days=days_ahead)
    pmap = {p["product_id"]: p for p in get_all_products()}
    results = []
    for b in get_product_batches(active_only=True):
        exp = b.get("expiry_date")
        if exp is None or exp > cutoff:
            continue
        product = pmap.get(b.get("product_id"))
        results.append({
            "product_id": b.get("product_id"),
            "name": product["name"] if product else f"#{b.get('product_id')}",
            "category": product.get("category") if product else None,
            "batch_number": b.get("batch_number"),
            "quantity": b.get("qty_remaining"),
            "expiry_date": exp,
            "expired": exp < today,
            "days_left": (exp - today).days,
        })
    results.sort(key=lambda x: x["expiry_date"])
    return results


def search_products(query):
    q = f"%{str(query).strip()}%"
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM products WHERE tenant_id = %s AND (name LIKE %s OR barcode LIKE %s) LIMIT 20",
                (_tid(), q, q)
            )
            rows = cur.fetchall()
    for p in rows:
        p["expiry_date"] = _normalize_date(p["expiry_date"])
    return rows


# ── Customers ─────────────────────────────────────────────────────────

def get_all_customers():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM customers WHERE tenant_id = %s ORDER BY customer_id", (_tid(),))
            return cur.fetchall()


def get_customer(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return None
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM customers WHERE customer_id = %s AND tenant_id = %s",
                        (cid, _tid()))
            return cur.fetchone()


def add_customer(data):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO customers (tenant_id, name, phone, email, address, tax_id, notes, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                _tid(),
                (data.get("name") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
                (data.get("address") or "").strip(),
                (data.get("tax_id") or "").strip(),
                (data.get("notes") or "").strip(),
                datetime.now(),
            ))
            return cur.lastrowid


def update_customer(customer_id, data):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE customers SET
                    name = %s, phone = %s, email = %s,
                    address = %s, tax_id = %s, notes = %s
                WHERE customer_id = %s AND tenant_id = %s
            """, (
                (data.get("name") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
                (data.get("address") or "").strip(),
                (data.get("tax_id") or "").strip(),
                (data.get("notes") or "").strip(),
                cid,
                _tid(),
            ))


def delete_customer(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return False
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM invoices WHERE customer_id = %s AND tenant_id = %s",
                        (cid, _tid()))
            if cur.fetchone()["cnt"] > 0:
                raise ValueError("Cannot delete customer: invoices reference this profile.")
            cur.execute("DELETE FROM customers WHERE customer_id = %s AND tenant_id = %s",
                        (cid, _tid()))
    return True


def search_customers(query):
    q = f"%{str(query).strip()}%"
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM customers WHERE tenant_id = %s AND "
                "(name LIKE %s OR phone LIKE %s OR email LIKE %s) LIMIT 20",
                (_tid(), q, q, q)
            )
            return cur.fetchall()


def customer_lookup():
    return {c["customer_id"]: c for c in get_all_customers()}


def get_sales_summary_by_customer(start_date=None, end_date=None):
    cmap = customer_lookup()
    by_c = {cid: {"customer": c, "invoice_count": 0, "total_revenue": 0.0} for cid, c in cmap.items()}
    walk_in = {"customer": None, "invoice_count": 0, "total_revenue": 0.0}
    for inv in get_all_invoices(start_date=start_date, end_date=end_date):
        total = float(inv["total"] or 0)
        cid = normalize_customer_id(inv.get("customer_id"))
        if cid is None or cid not in by_c:
            walk_in["invoice_count"] += 1
            walk_in["total_revenue"] += total
        else:
            by_c[cid]["invoice_count"] += 1
            by_c[cid]["total_revenue"] += total
    rows = [v for v in by_c.values() if v["invoice_count"]]
    rows.sort(key=lambda x: x["total_revenue"], reverse=True)
    for r in rows:
        r["total_revenue"] = round(r["total_revenue"], 2)
    walk_in["total_revenue"] = round(walk_in["total_revenue"], 2)
    return rows, walk_in


def get_invoices_for_customer(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return []
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM invoices WHERE customer_id = %s AND tenant_id = %s ORDER BY invoice_id DESC",
                (cid, _tid())
            )
            return cur.fetchall()


def get_customer_product_aggregates(customer_id):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        return []
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT ii.product_id, ii.product_name,
                       SUM(ii.quantity) AS total_qty,
                       ROUND(SUM(ii.line_total), 2) AS total_amount
                FROM invoice_items ii
                JOIN invoices i ON i.invoice_id = ii.invoice_id
                WHERE i.customer_id = %s AND i.tenant_id = %s
                GROUP BY ii.product_id, ii.product_name
                ORDER BY ii.product_name
            """, (cid, _tid()))
            return cur.fetchall()


# ── Invoices ──────────────────────────────────────────────────────────

def get_all_invoices(include_deleted=False, start_date=None, end_date=None):
    with _conn() as conn:
        with conn.cursor() as cur:
            where = "tenant_id = %s"
            params = [_tid()]
            if not include_deleted:
                where += " AND (status != 'deleted' OR status IS NULL)"
            if start_date:
                where += " AND DATE(created_at) >= %s"
                params.append(start_date)
            if end_date:
                where += " AND DATE(created_at) <= %s"
                params.append(end_date)
            cur.execute(f"SELECT * FROM invoices WHERE {where} ORDER BY invoice_id DESC", params)
            return cur.fetchall()


def get_deleted_invoices():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM invoices WHERE tenant_id = %s AND status = 'deleted' "
                        "ORDER BY invoice_id DESC", (_tid(),))
            return cur.fetchall()


def get_invoice(invoice_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM invoices WHERE invoice_id = %s AND tenant_id = %s",
                        (int(invoice_id), _tid()))
            return cur.fetchone()


def delete_invoice(invoice_id, deleted_by, reason=""):
    """Soft-delete: mark deleted, reverse stock, reverse credit ledger."""
    iid = int(invoice_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM invoices WHERE invoice_id = %s AND tenant_id = %s", (iid, tid))
            inv = cur.fetchone()
            if not inv:
                raise ValueError("Invoice not found")
            if inv.get("status") == "deleted":
                raise ValueError("Invoice already deleted")

            # Guard: deleting a credit sale the customer has already paid against
            # would leave them overpaid (negative balance). Block it.
            cid = normalize_customer_id(inv.get("customer_id"))
            is_credit = str(inv.get("payment_method") or "").strip().lower() == "credit"
            if is_credit and cid is not None:
                cur.execute("""
                    SELECT
                        COALESCE(SUM(CASE WHEN type='debit'  THEN amount ELSE 0 END),0) AS debit,
                        COALESCE(SUM(CASE WHEN type='credit' THEN amount ELSE 0 END),0) AS paid
                    FROM credit_ledger WHERE customer_id = %s AND tenant_id = %s
                """, (cid, tid))
                bal = cur.fetchone()
                inv_amount = float(inv.get("total") or 0)
                if round((float(bal["debit"]) - inv_amount) - float(bal["paid"]), 2) < 0:
                    raise ValueError(
                        "This credit sale has payments against the customer's account. "
                        "Deleting it would leave the customer overpaid. Record a refund or "
                        "adjust their balance first, then delete.")

            # Reverse stock — back to the specific batch each item was sold
            # from (or that product's legacy batch for pre-batch invoices).
            cur.execute("SELECT * FROM invoice_items WHERE invoice_id = %s AND tenant_id = %s", (iid, tid))
            touched_pids = set()
            for item in cur.fetchall():
                pid = int(item["product_id"])
                _restore_batch_for_sale(cur, tid, pid, int(item["quantity"]), item.get("batch_id"))
                touched_pids.add(pid)
            for pid in touched_pids:
                _recompute_product_cache(cur, tid, pid)
            # Remove the credit-sale debit for this invoice (payments have NULL invoice_id)
            cur.execute("DELETE FROM credit_ledger WHERE invoice_id = %s AND type = 'debit' AND tenant_id = %s",
                        (iid, tid))
            # Remove the auto-synced journal entry for this sale (cascade clears lines)
            cur.execute(
                "DELETE FROM journal_entries WHERE source_type = 'sale' AND source_id = %s AND tenant_id = %s",
                (iid, tid))
            # Soft-delete
            from datetime import datetime as _dt
            cur.execute(
                """UPDATE invoices SET status='deleted', deleted_at=%s, deleted_by=%s, delete_reason=%s
                   WHERE invoice_id = %s AND tenant_id = %s""",
                (_dt.now(), deleted_by, reason, iid, tid)
            )


def get_invoice_items(invoice_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM invoice_items WHERE invoice_id = %s AND tenant_id = %s",
                (int(invoice_id), _tid())
            )
            return cur.fetchall()


def create_invoice(items, tax_rate, payment_method, customer_id=None, delivery_charges=0.0):
    cid = normalize_customer_id(customer_id)
    if payment_method == "Credit" and cid is None:
        raise ValueError("Credit payment requires a customer to be selected.")
    tid = _tid()

    with _conn() as conn:
        with conn.cursor() as cur:
            if cid is not None:
                cur.execute("SELECT customer_id FROM customers WHERE customer_id = %s AND tenant_id = %s",
                            (cid, tid))
                if not cur.fetchone():
                    raise ValueError("Customer not found.")

            subtotal = 0.0
            discount_total = 0.0
            line_entries = []
            touched_pids = set()

            for item in items:
                pid = int(item["product_id"])
                qty = int(item["quantity"])
                discount_per_unit = float(item.get("discount_amount", 0))
                requested_batch_id = item.get("batch_id")

                cur.execute("SELECT * FROM products WHERE product_id = %s AND tenant_id = %s FOR UPDATE",
                            (pid, tid))
                prod = cur.fetchone()
                if not prod:
                    raise ValueError(f"Product ID {pid} not found")
                catalog_counter = float(prod["counter_price"])

                # Picks (or auto-picks) a batch and deducts qty from it; the
                # batch's own cost is the COGS for this line, not the blended
                # product-level cache.
                batch_id_used, purchase_price = _deplete_batch_for_sale(
                    cur, tid, pid, qty, requested_batch_id, prod["name"])
                touched_pids.add(pid)

                raw_unit = item.get("unit_price")
                unit_price = float(raw_unit) if raw_unit not in (None, "") else catalog_counter

                if unit_price < 0:
                    raise ValueError(f"Invalid sale price for '{prod['name']}'")

                discounted_price = unit_price - discount_per_unit
                if discounted_price < purchase_price:
                    raise ValueError(
                        f"Discount too high for '{prod['name']}': "
                        f"price after discount {discounted_price:.2f} below "
                        f"purchase price {purchase_price:.2f}. "
                        f"Max discount: {unit_price - purchase_price:.2f}"
                    )

                line_discount = discount_per_unit * qty
                line_total = discounted_price * qty
                subtotal += unit_price * qty
                discount_total += line_discount

                line_entries.append((pid, prod["name"], purchase_price, unit_price, qty, line_discount,
                                     round(line_total, 2), batch_id_used))

            for pid in touched_pids:
                _recompute_product_cache(cur, tid, pid)

            net = round(subtotal - discount_total, 2)
            tax_amount = round(net * tax_rate, 2)
            delivery = round(float(delivery_charges or 0), 2)
            total = round(net + tax_amount + delivery, 2)

            cur.execute("""
                INSERT INTO invoices (tenant_id, created_at, subtotal, discount_total, tax_rate, tax_amount, total, payment_method, customer_id, delivery_charges)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (tid, datetime.now(), round(subtotal, 2), round(discount_total, 2), tax_rate, tax_amount, total, payment_method, cid, delivery))
            invoice_id = cur.lastrowid

            for pid, pname, pp, up, qty, ld, lt, batch_id in line_entries:
                cur.execute("""
                    INSERT INTO invoice_items (tenant_id, invoice_id, product_id, product_name, purchase_price, counter_price, quantity, discount_amount, line_total, batch_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (tid, invoice_id, pid, pname, pp, up, qty, ld, lt, batch_id))

            if payment_method == "Credit" and cid is not None:
                cur.execute("""
                    INSERT INTO credit_ledger (tenant_id, customer_id, invoice_id, type, amount, note, created_at)
                    VALUES (%s, %s, %s, 'debit', %s, 'Credit sale', %s)
                """, (tid, cid, invoice_id, total, datetime.now()))

    return invoice_id


def update_invoice(invoice_id, items, tax_rate, payment_method=None, customer_id=None):
    iid = int(invoice_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM invoices WHERE invoice_id = %s AND tenant_id = %s FOR UPDATE",
                        (iid, tid))
            invoice = cur.fetchone()
            if not invoice:
                raise ValueError("Invoice not found.")

            cur.execute("SELECT * FROM invoice_items WHERE invoice_id = %s AND tenant_id = %s", (iid, tid))
            old_items = cur.fetchall()

            # Return stock from old items to the exact batch each was sold
            # from (or that product's legacy batch for pre-batch invoices).
            touched_pids = {int(old["product_id"]) for old in old_items}
            for old in old_items:
                _restore_batch_for_sale(cur, tid, int(old["product_id"]), int(old["quantity"]), old.get("batch_id"))

            old_payment = invoice.get("payment_method")
            old_cid = normalize_customer_id(invoice.get("customer_id"))
            new_payment = payment_method if payment_method is not None else old_payment
            new_cid = normalize_customer_id(customer_id) if customer_id is not None else old_cid

            if new_payment == "Credit" and new_cid is None:
                raise ValueError("Credit payment requires a customer.")
            if new_cid is not None:
                cur.execute("SELECT customer_id FROM customers WHERE customer_id = %s AND tenant_id = %s",
                            (new_cid, tid))
                if not cur.fetchone():
                    raise ValueError("Customer not found.")

            subtotal = 0.0
            discount_total = 0.0
            line_entries = []

            for item in items:
                pid = int(item["product_id"])
                qty = int(item["quantity"])
                discount_per_unit = float(item.get("discount_amount", 0))
                requested_batch_id = item.get("batch_id")

                cur.execute("SELECT * FROM products WHERE product_id = %s AND tenant_id = %s FOR UPDATE",
                            (pid, tid))
                prod = cur.fetchone()
                if not prod:
                    raise ValueError(f"Product ID {pid} not found")
                catalog_counter = float(prod["counter_price"])

                batch_id_used, purchase_price = _deplete_batch_for_sale(
                    cur, tid, pid, qty, requested_batch_id, prod["name"])
                touched_pids.add(pid)

                raw_unit = item.get("unit_price")
                unit_price = float(raw_unit) if raw_unit not in (None, "") else catalog_counter

                if unit_price < 0:
                    raise ValueError(f"Invalid sale price for '{prod['name']}'")

                discounted_price = unit_price - discount_per_unit
                if discounted_price < purchase_price:
                    raise ValueError(
                        f"Discount too high for '{prod['name']}': "
                        f"price after discount {discounted_price:.2f} below purchase {purchase_price:.2f}"
                    )

                line_discount = discount_per_unit * qty
                line_total = discounted_price * qty
                subtotal += unit_price * qty
                discount_total += line_discount

                line_entries.append((pid, prod["name"], purchase_price, unit_price, qty, line_discount,
                                     round(line_total, 2), batch_id_used))

            for pid in touched_pids:
                _recompute_product_cache(cur, tid, pid)

            net = round(subtotal - discount_total, 2)
            tax_amount = round(net * tax_rate, 2)
            total = round(net + tax_amount, 2)

            cur.execute("DELETE FROM invoice_items WHERE invoice_id = %s AND tenant_id = %s", (iid, tid))
            for pid, pname, pp, up, qty, ld, lt, batch_id in line_entries:
                cur.execute("""
                    INSERT INTO invoice_items (tenant_id, invoice_id, product_id, product_name, purchase_price, counter_price, quantity, discount_amount, line_total, batch_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (tid, iid, pid, pname, pp, up, qty, ld, lt, batch_id))

            cur.execute("""
                UPDATE invoices SET
                    subtotal = %s, discount_total = %s, tax_rate = %s,
                    tax_amount = %s, total = %s, payment_method = %s, customer_id = %s
                WHERE invoice_id = %s AND tenant_id = %s
            """, (round(subtotal, 2), round(discount_total, 2), tax_rate, tax_amount, total, new_payment, new_cid, iid, tid))

            # Reconcile ledger: delete old debit for this invoice, re-add if Credit
            cur.execute(
                "DELETE FROM credit_ledger WHERE invoice_id = %s AND type = 'debit' AND tenant_id = %s",
                (iid, tid)
            )
            if new_payment == "Credit" and new_cid is not None:
                cur.execute("""
                    INSERT INTO credit_ledger (tenant_id, customer_id, invoice_id, type, amount, note, created_at)
                    VALUES (%s, %s, %s, 'debit', %s, 'Credit sale', %s)
                """, (tid, new_cid, iid, total, datetime.now()))

    return iid


# ── Credit Ledger ─────────────────────────────────────────────────────

def get_credit_ledger(customer_id=None):
    cid = normalize_customer_id(customer_id) if customer_id is not None else None
    with _conn() as conn:
        with conn.cursor() as cur:
            if cid is not None:
                cur.execute(
                    "SELECT * FROM credit_ledger WHERE customer_id = %s AND tenant_id = %s ORDER BY entry_id DESC",
                    (cid, _tid())
                )
            else:
                cur.execute("SELECT * FROM credit_ledger WHERE tenant_id = %s ORDER BY entry_id DESC",
                            (_tid(),))
            return cur.fetchall()


def add_ledger_payment(customer_id, amount, note="", payment_method="cash"):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        raise ValueError("Invalid customer.")
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_customer(cid):
        raise ValueError("Customer not found.")
    pm = (payment_method or "cash").strip().lower()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO credit_ledger (tenant_id, customer_id, invoice_id, type, amount, note, created_at, payment_method)
                VALUES (%s, %s, NULL, 'credit', %s, %s, %s, %s)
            """, (_tid(), cid, amount, (note or "").strip(), datetime.now(), pm))
            return cur.lastrowid


def delete_ledger_payment(entry_id):
    eid = int(entry_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM credit_ledger WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
            row = cur.fetchone()
            if not row:
                raise ValueError("Entry not found.")
            if row["type"] != "credit" or row.get("invoice_id"):
                raise ValueError("Only standalone payment entries can be deleted.")
            cur.execute("DELETE FROM journal_entries WHERE source_type='customer_payment' AND source_id=%s AND tenant_id=%s", (eid, tid))
            cur.execute("DELETE FROM credit_ledger WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
        conn.commit()


def update_ledger_payment(entry_id, amount, note="", payment_method="cash"):
    eid = int(entry_id)
    tid = _tid()
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    pm = (payment_method or "cash").strip().lower()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM credit_ledger WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
            row = cur.fetchone()
            if not row:
                raise ValueError("Entry not found.")
            if row["type"] != "credit" or row.get("invoice_id"):
                raise ValueError("Only standalone payment entries can be edited.")
            cur.execute(
                "UPDATE credit_ledger SET amount=%s, note=%s, payment_method=%s WHERE entry_id=%s AND tenant_id=%s",
                (amount, (note or "").strip(), pm, eid, tid)
            )
            cur.execute("DELETE FROM journal_entries WHERE source_type='customer_payment' AND source_id=%s AND tenant_id=%s", (eid, tid))
            cur.execute("SELECT account_id, code FROM accounts WHERE tenant_id=%s AND code IN ('1000','1010','4000')", (tid,))
            acc = {r["code"]: r["account_id"] for r in cur.fetchall()}
            CASH, BANK, SALES = acc["1000"], acc["1010"], acc["4000"]
            recv_acct = BANK if pm == "bank" else CASH
            cur.execute(
                "INSERT INTO journal_entries (tenant_id, date, description, source_type, source_id, created_by) VALUES (%s,%s,'Customer payment','customer_payment',%s,'system')",
                (tid, row["created_at"], eid)
            )
            je_id = cur.lastrowid
            cur.executemany(
                "INSERT INTO journal_lines (tenant_id, entry_id, account_id, debit, credit) VALUES (%s,%s,%s,%s,%s)",
                [(tid, je_id, recv_acct, amount, 0.0), (tid, je_id, SALES, 0.0, amount)]
            )
        conn.commit()


def add_ledger_debit(customer_id, amount, note=""):
    cid = normalize_customer_id(customer_id)
    if cid is None:
        raise ValueError("Invalid customer.")
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_customer(cid):
        raise ValueError("Customer not found.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO credit_ledger (tenant_id, customer_id, invoice_id, type, amount, note, created_at)
                VALUES (%s, %s, NULL, 'debit', %s, %s, %s)
            """, (_tid(), cid, amount, (note or "").strip(), datetime.now()))
            return cur.lastrowid


def get_customer_balance(customer_id):
    entries = get_credit_ledger(customer_id=customer_id)
    total_debt = sum(float(e["amount"] or 0) for e in entries if e["type"] == "debit")
    total_paid = sum(float(e["amount"] or 0) for e in entries if e["type"] == "credit")
    return round(total_debt, 2), round(total_paid, 2), round(total_debt - total_paid, 2)


def get_all_credit_balances():
    cmap = customer_lookup()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT customer_id,
                       SUM(CASE WHEN type='debit'  THEN amount ELSE 0 END) AS total_debt,
                       SUM(CASE WHEN type='credit' THEN amount ELSE 0 END) AS total_paid
                FROM credit_ledger
                WHERE tenant_id = %s
                GROUP BY customer_id
            """, (_tid(),))
            rows = cur.fetchall()
    result = []
    for row in rows:
        cid = normalize_customer_id(row["customer_id"])
        if cid is None:
            continue
        balance = round(float(row["total_debt"]) - float(row["total_paid"]), 2)
        result.append({
            "customer": cmap.get(cid),
            "customer_id": cid,
            "total_debt": round(float(row["total_debt"]), 2),
            "total_paid": round(float(row["total_paid"]), 2),
            "balance": balance,
        })
    result.sort(key=lambda x: x["balance"], reverse=True)
    return result


# ── Analytics ─────────────────────────────────────────────────────────

def get_today_sales():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) AS cnt, COALESCE(SUM(total), 0) AS total
                FROM invoices
                WHERE tenant_id = %s AND DATE(created_at) = CURDATE()
                  AND (status != 'deleted' OR status IS NULL)
            """, (_tid(),))
            row = cur.fetchone()
    return int(row["cnt"]), round(float(row["total"]), 2)


# ── Import from Excel ─────────────────────────────────────────────────

def _parse_expiry(val):
    if val is None or val == "" or val == "None":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})\s+(\d{4})$', s)
    if m:
        month, year = int(m.group(1)), int(m.group(2))
        if 1 <= month <= 12:
            if month == 12:
                return date(year, 12, 31)
            return date(year, month + 1, 1) - timedelta(days=1)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def import_from_excel(filepath):
    from openpyxl import load_workbook as lw
    wb_src = lw(filepath)
    ws = wb_src.active
    imported = 0
    skipped = 0
    errors = []
    has_counter_col = ws.max_column >= 9

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None and row[1] is None:
            continue
        name = str(row[1] or "").strip()
        if not name:
            skipped += 1
            continue
        packing = str(row[2] or "").strip()
        if packing:
            name = f"{name} ({packing})"
        try:
            qty = int(row[3] or 0)
        except (ValueError, TypeError):
            qty = 0
        try:
            purchase_price = float(row[4] or 0)
        except (ValueError, TypeError):
            purchase_price = 0.0
        company = str(row[5] or "").strip()
        expiry = _parse_expiry(row[6])
        try:
            mrp = float(row[7] or 0)
        except (ValueError, TypeError):
            mrp = 0.0
        if has_counter_col and len(row) >= 9 and row[8] is not None:
            try:
                counter_price = float(row[8])
            except (ValueError, TypeError):
                counter_price = mrp if mrp > 0 else purchase_price
        else:
            counter_price = mrp if mrp > 0 else purchase_price
        try:
            add_product({
                "name": name,
                "purchase_price": purchase_price,
                "counter_price": counter_price,
                "retail_price": mrp,
                "quantity": qty,
                "barcode": "",
                "expiry_date": expiry,
                "category": company,
            })
            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")
            skipped += 1

    wb_src.close()
    return imported, skipped, errors


# ── Users ─────────────────────────────────────────────────────────────

def get_all_users(tenant_id=None):
    with _conn() as conn:
        with conn.cursor() as cur:
            if tenant_id is not None:
                cur.execute("SELECT * FROM users WHERE tenant_id = %s ORDER BY user_id",
                            (int(tenant_id),))
            else:
                cur.execute("SELECT * FROM users ORDER BY user_id")
            rows = cur.fetchall()
    for u in rows:
        u["is_active"] = bool(u["is_active"])
    return rows


def has_any_super_admin():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM users WHERE role = 'super_admin'")
            return cur.fetchone()["cnt"] > 0


def get_user_by_id(user_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM users WHERE user_id = %s", (int(user_id),))
            row = cur.fetchone()
    if row:
        row["is_active"] = bool(row["is_active"])
    return row


def get_user_by_username(username):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM users WHERE username = %s", (username.strip().lower(),))
            row = cur.fetchone()
    if row:
        row["is_active"] = bool(row["is_active"])
    return row


def has_any_users():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM users")
            return cur.fetchone()["cnt"] > 0


def add_user(username, password_hash, role="staff", tenant_id=None):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO users (tenant_id, username, password_hash, role, is_active, created_at)
                VALUES (%s, %s, %s, %s, 1, %s)
            """, (int(tenant_id) if tenant_id is not None else None,
                  username.strip().lower(), password_hash, role, datetime.now()))
            return cur.lastrowid


# ── Tenants (business accounts) ──────────────────────────────────────

def create_tenant(name):
    name = str(name).strip()
    if not name:
        raise ValueError("Business name is required.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO tenants (name, is_active) VALUES (%s, 1)", (name,))
            return cur.lastrowid


def get_all_tenants(include_inactive=True):
    with _conn() as conn:
        with conn.cursor() as cur:
            if include_inactive:
                cur.execute("SELECT * FROM tenants ORDER BY tenant_id")
            else:
                cur.execute("SELECT * FROM tenants WHERE is_active = 1 ORDER BY tenant_id")
            rows = cur.fetchall()
    for t in rows:
        t["is_active"] = bool(t["is_active"])
    return rows


def get_tenant(tenant_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM tenants WHERE tenant_id = %s", (int(tenant_id),))
            t = cur.fetchone()
    if t:
        t["is_active"] = bool(t["is_active"])
    return t


def set_tenant_active(tenant_id, active):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE tenants SET is_active = %s WHERE tenant_id = %s",
                        (1 if active else 0, int(tenant_id)))


def update_user_password(user_id, password_hash):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE users SET password_hash = %s WHERE user_id = %s",
                (password_hash, int(user_id))
            )


def set_user_role(user_id, role):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE users SET role = %s WHERE user_id = %s",
                (role, int(user_id))
            )


def toggle_user_active(user_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE users SET is_active = NOT is_active WHERE user_id = %s",
                (int(user_id),)
            )


def delete_user(user_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM users WHERE user_id = %s", (int(user_id),))


# ── Suppliers ─────────────────────────────────────────────────────────

def get_all_suppliers():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM suppliers WHERE tenant_id = %s ORDER BY supplier_id", (_tid(),))
            return cur.fetchall()


def get_supplier(supplier_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM suppliers WHERE supplier_id = %s AND tenant_id = %s",
                        (int(supplier_id), _tid()))
            return cur.fetchone()


def add_supplier(data):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO suppliers (tenant_id, name, phone, email, address, notes, created_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (
                _tid(),
                (data.get("name") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
                (data.get("address") or "").strip(),
                (data.get("notes") or "").strip(),
                datetime.now(),
            ))
            return cur.lastrowid


def update_supplier(supplier_id, data):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE suppliers SET name=%s, phone=%s, email=%s, address=%s, notes=%s
                WHERE supplier_id=%s AND tenant_id=%s
            """, (
                (data.get("name") or "").strip(),
                (data.get("phone") or "").strip(),
                (data.get("email") or "").strip(),
                (data.get("address") or "").strip(),
                (data.get("notes") or "").strip(),
                int(supplier_id),
                _tid(),
            ))


def delete_supplier(supplier_id):
    sid = int(supplier_id)
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM purchase_invoices WHERE supplier_id=%s AND tenant_id=%s",
                        (sid, _tid()))
            if cur.fetchone()["cnt"] > 0:
                raise ValueError("Cannot delete supplier: purchase invoices reference this supplier.")
            cur.execute("DELETE FROM suppliers WHERE supplier_id=%s AND tenant_id=%s", (sid, _tid()))
    return True


def supplier_lookup():
    return {s["supplier_id"]: s for s in get_all_suppliers()}


# ── Purchase Invoices ─────────────────────────────────────────────────

def create_purchase_invoice(supplier_id, items, notes="", payment_method="credit",
                            purchase_date=None, direct_amount=None):
    """direct_amount: when no items, record the purchase as a lump total only
    (old purchase where products aren't remembered)."""
    sid = int(supplier_id)
    tid = _tid()
    payment_method = (payment_method or "credit").strip().lower()
    if payment_method not in ("credit", "cash", "bank"):
        payment_method = "credit"
    when = datetime.now()
    if purchase_date:
        if isinstance(purchase_date, str):
            when = datetime.combine(date.fromisoformat(purchase_date), datetime.min.time())
        elif isinstance(purchase_date, datetime):
            when = purchase_date
        elif isinstance(purchase_date, date):
            when = datetime.combine(purchase_date, datetime.min.time())
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT supplier_id FROM suppliers WHERE supplier_id=%s AND tenant_id=%s", (sid, tid))
            if not cur.fetchone():
                raise ValueError("Supplier not found.")

            total_amount = 0.0
            prepared = []  # (pid, pname, qty, unit_cost, line_total, batch_number, expiry, counter_price, retail_price)

            for item in items:
                pid = int(item["product_id"])
                qty = int(item["quantity"])
                unit_cost = float(item["unit_cost"])
                cur.execute("SELECT * FROM products WHERE product_id=%s AND tenant_id=%s FOR UPDATE", (pid, tid))
                prod = cur.fetchone()
                if not prod:
                    raise ValueError(f"Product ID {pid} not found")
                line_total = round(unit_cost * qty, 2)
                total_amount += line_total
                batch_number = str(item.get("batch_number") or "").strip()
                expiry = item.get("expiry_date") or None
                counter_price = item.get("counter_price")
                retail_price = item.get("retail_price")
                prepared.append((pid, prod["name"], qty, unit_cost, line_total, batch_number, expiry,
                                  counter_price, retail_price))

            if not items and direct_amount is not None:
                total_amount = float(direct_amount)
            total_amount = round(total_amount, 2)
            if total_amount <= 0:
                raise ValueError("Purchase total must be positive.")
            cur.execute("""
                INSERT INTO purchase_invoices (tenant_id, supplier_id, created_at, total_amount, notes, payment_method)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (tid, sid, when, total_amount, (notes or "").strip(), payment_method))
            purchase_id = cur.lastrowid

            line_entries = []
            touched_pids = set()
            for pid, pname, qty, unit_cost, line_total, batch_number, expiry, counter_price, retail_price in prepared:
                batch_id_used = None
                if batch_number:
                    # Reusing a lot number at the same cost tops up that batch.
                    # Reusing it at a different cost is a legitimate second
                    # arrival of the same lot (supplier re-priced mid-batch) --
                    # it becomes its own batch entry sharing the same lot
                    # label, so both cost tiers stay visible instead of one
                    # silently overwriting or blocking the other.
                    cur.execute("""
                        SELECT batch_id, unit_cost, qty_received, qty_remaining FROM product_batches
                        WHERE product_id=%s AND tenant_id=%s AND LOWER(batch_number)=LOWER(%s)
                        FOR UPDATE
                    """, (pid, tid, batch_number))
                    candidates = cur.fetchall()
                    match = next((c for c in candidates
                                  if round(float(c["unit_cost"] or 0), 2) == round(unit_cost, 2)), None)
                    if match:
                        cur.execute("""
                            UPDATE product_batches SET qty_received = qty_received + %s,
                                   qty_remaining = qty_remaining + %s
                            WHERE batch_id = %s
                        """, (qty, qty, match["batch_id"]))
                        batch_id_used = match["batch_id"]
                    else:
                        cur.execute("""
                            INSERT INTO product_batches
                                (tenant_id, product_id, batch_number, supplier_id, purchase_id,
                                 unit_cost, qty_received, qty_remaining, expiry_date, received_at)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (tid, pid, batch_number, sid, purchase_id, unit_cost, qty, qty, expiry, when))
                        batch_id_used = cur.lastrowid
                else:
                    cur.execute("""
                        INSERT INTO product_batches
                            (tenant_id, product_id, batch_number, supplier_id, purchase_id,
                             unit_cost, qty_received, qty_remaining, expiry_date, received_at)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (tid, pid, None, sid, purchase_id, unit_cost, qty, qty, expiry, when))
                    batch_id_used = cur.lastrowid
                    cur.execute("UPDATE product_batches SET batch_number=%s WHERE batch_id=%s",
                                (f"P{pid}-B{batch_id_used}", batch_id_used))
                line_entries.append((pid, pname, qty, unit_cost, line_total, batch_id_used))
                cur.execute("UPDATE products SET last_supplier_id=%s WHERE product_id=%s AND tenant_id=%s",
                            (sid, pid, tid))
                if counter_price not in (None, ""):
                    cur.execute("UPDATE products SET counter_price=%s WHERE product_id=%s AND tenant_id=%s",
                                (round(float(counter_price), 2), pid, tid))
                if retail_price not in (None, ""):
                    cur.execute("UPDATE products SET retail_price=%s WHERE product_id=%s AND tenant_id=%s",
                                (round(float(retail_price), 2), pid, tid))
                touched_pids.add(pid)

            for pid in touched_pids:
                _recompute_product_cache(cur, tid, pid)

            for pid, pname, qty, uc, lt, batch_id in line_entries:
                cur.execute("""
                    INSERT INTO purchase_invoice_items
                        (tenant_id, purchase_id, product_id, product_name, quantity, unit_cost, line_total, batch_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """, (tid, purchase_id, pid, pname, qty, uc, lt, batch_id))

            # Always record the purchase in the supplier ledger (relationship history).
            # Use the user's note if given, else a generic reference.
            debit_note = (notes or "").strip() or f"Purchase invoice #{purchase_id}"
            cur.execute("""
                INSERT INTO supplier_ledger (tenant_id, supplier_id, purchase_id, type, amount, note, created_at, payment_method)
                VALUES (%s,%s,%s,'debit',%s,%s,%s,NULL)
            """, (tid, sid, purchase_id, total_amount, debit_note, when))
            # If paid now, also record the payment so the supplier balance nets to zero.
            if payment_method in ("cash", "bank"):
                cur.execute("""
                    INSERT INTO supplier_ledger (tenant_id, supplier_id, purchase_id, type, amount, note, created_at, payment_method)
                    VALUES (%s,%s,%s,'credit',%s,%s,%s,%s)
                """, (tid, sid, purchase_id, total_amount, f"Paid by {payment_method}", when, payment_method))

    return purchase_id


def get_all_purchase_invoices():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM purchase_invoices WHERE tenant_id=%s ORDER BY purchase_id DESC", (_tid(),))
            return cur.fetchall()


def get_purchase_invoice(purchase_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM purchase_invoices WHERE purchase_id=%s AND tenant_id=%s",
                        (int(purchase_id), _tid()))
            return cur.fetchone()


def get_purchase_invoice_items(purchase_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM purchase_invoice_items WHERE purchase_id=%s AND tenant_id=%s",
                (int(purchase_id), _tid())
            )
            return cur.fetchall()


# ── Supplier Ledger ───────────────────────────────────────────────────

def add_supplier_payment(supplier_id, amount, note="", payment_method="cash"):
    sid = int(supplier_id)
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_supplier(sid):
        raise ValueError("Supplier not found.")
    pm = (payment_method or "cash").strip().lower()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO supplier_ledger (tenant_id, supplier_id, purchase_id, type, amount, note, created_at, payment_method)
                VALUES (%s, %s, NULL, 'credit', %s, %s, %s, %s)
            """, (_tid(), sid, amount, (note or "").strip(), datetime.now(), pm))
            return cur.lastrowid


def get_supplier_ledger_entries(supplier_id=None):
    with _conn() as conn:
        with conn.cursor() as cur:
            if supplier_id is not None:
                cur.execute(
                    "SELECT * FROM supplier_ledger WHERE supplier_id=%s AND tenant_id=%s ORDER BY entry_id",
                    (int(supplier_id), _tid())
                )
            else:
                cur.execute("SELECT * FROM supplier_ledger WHERE tenant_id=%s ORDER BY entry_id", (_tid(),))
            return cur.fetchall()


def get_supplier_balance(supplier_id):
    entries = get_supplier_ledger_entries(supplier_id)
    total_debt = sum(float(e["amount"] or 0) for e in entries if e["type"] == "debit")
    total_paid = sum(float(e["amount"] or 0) for e in entries if e["type"] == "credit")
    return round(total_debt, 2), round(total_paid, 2), round(total_debt - total_paid, 2)


def delete_purchase_invoice(purchase_id):
    pid = int(purchase_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT product_id, quantity, batch_id FROM purchase_invoice_items WHERE purchase_id=%s AND tenant_id=%s",
                (pid, tid)
            )
            items = cur.fetchall()
            touched_pids = set()
            for item in items:
                qty = int(item["quantity"])
                prod_id = item["product_id"]
                batch_id = item.get("batch_id")

                target_batch_id = None
                if batch_id is not None:
                    target_batch_id = batch_id
                else:
                    # Purchase predates batch tracking — reverse against that
                    # product's legacy batch (purchase_id IS NULL) if one exists.
                    cur.execute("""
                        SELECT batch_id FROM product_batches
                        WHERE product_id=%s AND tenant_id=%s AND purchase_id IS NULL
                        LIMIT 1
                    """, (prod_id, tid))
                    legacy = cur.fetchone()
                    if legacy:
                        target_batch_id = legacy["batch_id"]

                if target_batch_id is not None:
                    cur.execute("""
                        UPDATE product_batches
                        SET qty_received = GREATEST(0, qty_received - %s),
                            qty_remaining = GREATEST(0, qty_remaining - %s)
                        WHERE batch_id = %s AND tenant_id = %s
                    """, (qty, qty, target_batch_id, tid))
                    touched_pids.add(prod_id)
                else:
                    # No batch to reverse against at all — fall back to the
                    # pre-batch behaviour of decrementing stock directly.
                    cur.execute(
                        "UPDATE products SET quantity = GREATEST(0, quantity - %s) WHERE product_id=%s AND tenant_id=%s",
                        (qty, prod_id, tid)
                    )
            for prod_id in touched_pids:
                _recompute_product_cache(cur, tid, prod_id)

            cur.execute("DELETE FROM journal_entries WHERE source_type='purchase' AND source_id=%s AND tenant_id=%s", (pid, tid))
            cur.execute("DELETE FROM supplier_ledger WHERE purchase_id=%s AND tenant_id=%s", (pid, tid))
            cur.execute("DELETE FROM purchase_invoice_items WHERE purchase_id=%s AND tenant_id=%s", (pid, tid))
            cur.execute("DELETE FROM purchase_invoices WHERE purchase_id=%s AND tenant_id=%s", (pid, tid))


def delete_supplier_payment(entry_id):
    eid = int(entry_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM supplier_ledger WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
            row = cur.fetchone()
            if not row:
                raise ValueError("Entry not found.")
            if row["type"] != "credit" or row.get("purchase_id"):
                raise ValueError("Only standalone payment entries can be deleted.")
            cur.execute("DELETE FROM journal_entries WHERE source_type='supplier_payment' AND source_id=%s AND tenant_id=%s", (eid, tid))
            cur.execute("DELETE FROM supplier_ledger WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
        conn.commit()


def update_supplier_payment(entry_id, amount, note="", payment_method="cash"):
    eid = int(entry_id)
    tid = _tid()
    amount = float(amount)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    pm = (payment_method or "cash").strip().lower()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM supplier_ledger WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
            row = cur.fetchone()
            if not row:
                raise ValueError("Entry not found.")
            if row["type"] != "credit" or row.get("purchase_id"):
                raise ValueError("Only standalone payment entries can be edited.")
            cur.execute(
                "UPDATE supplier_ledger SET amount=%s, note=%s, payment_method=%s WHERE entry_id=%s AND tenant_id=%s",
                (amount, (note or "").strip(), pm, eid, tid)
            )
            cur.execute("DELETE FROM journal_entries WHERE source_type='supplier_payment' AND source_id=%s AND tenant_id=%s", (eid, tid))
            cur.execute("SELECT account_id, code FROM accounts WHERE tenant_id=%s AND code IN ('1000','1010','2000')", (tid,))
            acc = {r["code"]: r["account_id"] for r in cur.fetchall()}
            AP, CASH, BANK = acc["2000"], acc["1000"], acc["1010"]
            cash_acct = BANK if pm == "bank" else CASH
            cur.execute(
                "INSERT INTO journal_entries (tenant_id, date, description, source_type, source_id, created_by) VALUES (%s,%s,'Supplier payment','supplier_payment',%s,'system')",
                (tid, row["created_at"], eid)
            )
            je_id = cur.lastrowid
            cur.executemany(
                "INSERT INTO journal_lines (tenant_id, entry_id, account_id, debit, credit) VALUES (%s,%s,%s,%s,%s)",
                [(tid, je_id, AP, amount, 0.0), (tid, je_id, cash_acct, 0.0, amount)]
            )
        conn.commit()


def get_all_supplier_balances():
    smap = supplier_lookup()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT supplier_id,
                       SUM(CASE WHEN type='debit'  THEN amount ELSE 0 END) AS total_debt,
                       SUM(CASE WHEN type='credit' THEN amount ELSE 0 END) AS total_paid
                FROM supplier_ledger WHERE tenant_id = %s GROUP BY supplier_id
            """, (_tid(),))
            rows = cur.fetchall()
    result = []
    for row in rows:
        sid = int(row["supplier_id"])
        balance = round(float(row["total_debt"]) - float(row["total_paid"]), 2)
        result.append({
            "supplier": smap.get(sid),
            "supplier_id": sid,
            "total_debt": round(float(row["total_debt"]), 2),
            "total_paid": round(float(row["total_paid"]), 2),
            "balance": balance,
        })
    result.sort(key=lambda x: x["balance"], reverse=True)
    return result


# ── P&L Reports ───────────────────────────────────────────────────────

def _empty_pl_totals():
    zero = {"revenue": 0, "cogs": 0, "profit": 0}
    return {"revenue": 0, "cogs": 0, "profit": 0,
            "paid": dict(zero), "credit": dict(zero)}


def _split_pl_totals(rows):
    """Aggregate rows into grand totals plus paid vs credit buckets.
    Each row must have line_total, cogs, and is_credit."""
    buckets = {"paid": {"revenue": 0.0, "cogs": 0.0},
               "credit": {"revenue": 0.0, "cogs": 0.0}}
    for r in rows:
        b = buckets["credit"] if r.get("is_credit") else buckets["paid"]
        b["revenue"] += float(r["line_total"])
        b["cogs"] += float(r["cogs"])
    out = {}
    for key in ("paid", "credit"):
        rev = round(buckets[key]["revenue"], 2)
        cogs = round(buckets[key]["cogs"], 2)
        out[key] = {"revenue": rev, "cogs": cogs, "profit": round(rev - cogs, 2)}
    total_rev = round(out["paid"]["revenue"] + out["credit"]["revenue"], 2)
    total_cogs = round(out["paid"]["cogs"] + out["credit"]["cogs"], 2)
    out["revenue"] = total_rev
    out["cogs"] = total_cogs
    out["profit"] = round(total_rev - total_cogs, 2)
    return out


def get_sales_pl_report(start_date, end_date):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    DATE(i.created_at) AS date,
                    i.invoice_id,
                    i.payment_method,
                    ii.product_name,
                    ii.quantity,
                    ii.purchase_price,
                    ROUND(ii.line_total / ii.quantity, 2) AS sale_price,
                    ii.line_total,
                    ROUND(ii.purchase_price * ii.quantity, 2) AS cogs,
                    ROUND(ii.line_total - ii.purchase_price * ii.quantity, 2) AS profit
                FROM invoice_items ii
                JOIN invoices i ON i.invoice_id = ii.invoice_id
                WHERE i.tenant_id = %s AND DATE(i.created_at) BETWEEN %s AND %s
                  AND (i.status IS NULL OR i.status <> 'deleted')
                ORDER BY i.created_at, i.invoice_id
            """, (_tid(), start_date, end_date))
            rows = cur.fetchall()
    if not rows:
        return [], _empty_pl_totals()
    # A credit invoice that has since been fully paid off (realized) should be
    # treated as "paid" here, not stuck in "credit" forever just because it was
    # originally booked on credit.
    realized, _ = _realized_credit_invoice_ids()
    for r in rows:
        was_credit = str(r.get("payment_method") or "").strip().lower() == "credit"
        r["is_credit"] = was_credit and int(r["invoice_id"]) not in realized
    return rows, _split_pl_totals(rows)


def search_product_sales(product_query, start_date, end_date):
    """Return invoice lines whose product_name contains product_query within the date range."""
    q = f"%{(product_query or '').strip()}%"
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DATE(i.created_at) AS date,
                       i.invoice_id, i.payment_method,
                       c.name AS customer_name,
                       ii.product_name, ii.quantity,
                       ROUND(ii.line_total / ii.quantity, 2) AS sale_price,
                       ii.line_total
                FROM invoice_items ii
                JOIN invoices i ON i.invoice_id = ii.invoice_id AND i.tenant_id = ii.tenant_id
                LEFT JOIN customers c ON c.customer_id = i.customer_id AND c.tenant_id = i.tenant_id
                WHERE i.tenant_id = %s
                  AND (i.status IS NULL OR i.status != 'deleted')
                  AND DATE(i.created_at) BETWEEN %s AND %s
                  AND ii.product_name LIKE %s
                ORDER BY i.created_at DESC, i.invoice_id
            """, (_tid(), start_date, end_date, q))
            return cur.fetchall()


def get_supplier_sales_pl(supplier_id, start_date, end_date):
    """Sales P&L for units actually sourced from a given supplier. Each line
    is attributed by the batch it was sold from (batch.supplier_id) — exact,
    since different batches of the same product can come from different
    suppliers. Lines from invoices predating batch tracking (no batch_id)
    fall back to the product's current last_supplier_id as a best-effort
    approximation."""
    sid = int(supplier_id)
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    ii.product_id,
                    ii.product_name,
                    SUM(ii.quantity) AS total_qty,
                    ROUND(SUM(ii.purchase_price * ii.quantity), 2) AS total_cogs,
                    ROUND(SUM(ii.line_total), 2) AS total_revenue,
                    ROUND(SUM(ii.line_total - ii.purchase_price * ii.quantity), 2) AS total_profit
                FROM invoice_items ii
                JOIN invoices i ON i.invoice_id = ii.invoice_id AND i.tenant_id = ii.tenant_id
                JOIN products p ON p.product_id = ii.product_id AND p.tenant_id = ii.tenant_id
                LEFT JOIN product_batches b ON b.batch_id = ii.batch_id AND b.tenant_id = ii.tenant_id
                WHERE i.tenant_id = %s
                  AND DATE(i.created_at) BETWEEN %s AND %s
                  AND (i.status IS NULL OR i.status <> 'deleted')
                  AND (CASE WHEN ii.batch_id IS NOT NULL THEN b.supplier_id ELSE p.last_supplier_id END) = %s
                GROUP BY ii.product_id, ii.product_name
                ORDER BY ii.product_name
            """, (_tid(), start_date, end_date, sid))
            rows = cur.fetchall()
    total_revenue = round(sum(float(r["total_revenue"]) for r in rows), 2)
    total_cogs = round(sum(float(r["total_cogs"]) for r in rows), 2)
    total_profit = round(total_revenue - total_cogs, 2)
    return rows, {"revenue": total_revenue, "cogs": total_cogs, "profit": total_profit}


# ── Double-entry accounting: data layer ──────────────────────────────

def seed_chart_of_accounts():
    """Insert any standard accounts missing for the tenant (idempotent by code)."""
    tid = get_current_tenant()
    if tid is None:
        return  # no tenant context -> nothing to seed
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT code FROM accounts WHERE tenant_id = %s", (tid,))
            existing = {str(r["code"]) for r in cur.fetchall()}
            to_add = [(tid, c, n, t, 1 if sys else 0)
                      for c, n, t, sys in CHART_OF_ACCOUNTS if str(c) not in existing]
            if to_add:
                cur.executemany(
                    "INSERT INTO accounts (tenant_id, code, name, type, is_system) VALUES (%s,%s,%s,%s,%s)",
                    to_add)


def get_all_accounts():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM accounts WHERE tenant_id = %s ORDER BY code", (_tid(),))
            return cur.fetchall()


def get_account(account_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM accounts WHERE account_id = %s AND tenant_id = %s",
                        (int(account_id), _tid()))
            return cur.fetchone()


def get_account_by_code(code):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM accounts WHERE code = %s AND tenant_id = %s",
                        (str(code).strip(), _tid()))
            return cur.fetchone()


def get_accounts_by_type(atype):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM accounts WHERE type = %s AND tenant_id = %s ORDER BY code",
                        (atype, _tid()))
            return cur.fetchall()


def add_account(code, name, atype, is_system=False):
    code = str(code).strip()
    if not code or not str(name).strip():
        raise ValueError("Account code and name are required.")
    if atype not in ("asset", "liability", "equity", "income", "expense"):
        raise ValueError("Invalid account type.")
    if get_account_by_code(code):
        raise ValueError(f"Account code {code} already exists.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO accounts (tenant_id, code, name, type, is_system) VALUES (%s,%s,%s,%s,%s)",
                (_tid(), code, str(name).strip(), atype, 1 if is_system else 0),
            )
            return cur.lastrowid


def post_journal(description, lines, source_type="manual", source_id=None,
                 created_by="system", entry_date=None):
    """Post a balanced double-entry journal entry. lines: list of
    {account_id, debit, credit}. Returns entry_id; raises ValueError if unbalanced."""
    if not lines or len(lines) < 2:
        raise ValueError("A journal entry needs at least two lines.")
    norm = []
    total_debit = 0.0
    total_credit = 0.0
    for ln in lines:
        aid = int(ln["account_id"])
        debit = round(float(ln.get("debit", 0) or 0), 2)
        credit = round(float(ln.get("credit", 0) or 0), 2)
        if debit < 0 or credit < 0:
            raise ValueError("Debit/credit cannot be negative.")
        if debit > 0 and credit > 0:
            raise ValueError("A line cannot have both debit and credit.")
        if debit == 0 and credit == 0:
            continue
        norm.append((aid, debit, credit))
        total_debit += debit
        total_credit += credit
    if not norm:
        raise ValueError("Journal entry has no non-zero lines.")
    if round(total_debit - total_credit, 2) != 0:
        raise ValueError(
            f"Entry not balanced: debits {total_debit:.2f} != credits {total_credit:.2f}")

    when = entry_date or datetime.now()
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            ids = {a["account_id"] for a in get_all_accounts()}
            for aid, _, _ in norm:
                if aid not in ids:
                    raise ValueError(f"Account id {aid} not found.")
            cur.execute(
                """INSERT INTO journal_entries (tenant_id, date, description, source_type, source_id, created_by)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (tid, when, str(description).strip(), source_type, source_id, created_by),
            )
            entry_id = cur.lastrowid
            cur.executemany(
                "INSERT INTO journal_lines (tenant_id, entry_id, account_id, debit, credit) VALUES (%s,%s,%s,%s,%s)",
                [(tid, entry_id, aid, debit, credit) for aid, debit, credit in norm],
            )
            return entry_id


def get_journal_lines(entry_id=None):
    with _conn() as conn:
        with conn.cursor() as cur:
            if entry_id is not None:
                cur.execute("SELECT * FROM journal_lines WHERE entry_id = %s AND tenant_id = %s",
                            (int(entry_id), _tid()))
            else:
                cur.execute("SELECT * FROM journal_lines WHERE tenant_id = %s", (_tid(),))
            return cur.fetchall()


def get_account_balance(account_id):
    aid = int(account_id)
    acct = get_account(aid)
    if not acct:
        raise ValueError("Account not found.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COALESCE(SUM(debit),0) AS d, COALESCE(SUM(credit),0) AS c "
                "FROM journal_lines WHERE account_id = %s AND tenant_id = %s", (aid, _tid()))
            r = cur.fetchone()
    debit = float(r["d"])
    credit = float(r["c"])
    if acct["type"] in _DEBIT_NORMAL_TYPES:
        return round(debit - credit, 2)
    return round(credit - debit, 2)


def get_trial_balance():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT a.account_id, a.code, a.name, a.type,
                       COALESCE(SUM(jl.debit),0) AS debit,
                       COALESCE(SUM(jl.credit),0) AS credit
                FROM accounts a
                LEFT JOIN journal_lines jl ON jl.account_id = a.account_id
                WHERE a.tenant_id = %s
                GROUP BY a.account_id, a.code, a.name, a.type
                ORDER BY a.code
            """, (_tid(),))
            raw = cur.fetchall()
    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for r in raw:
        d = round(float(r["debit"]), 2)
        c = round(float(r["credit"]), 2)
        if d == 0 and c == 0:
            continue
        acct = {"account_id": r["account_id"], "code": r["code"],
                "name": r["name"], "type": r["type"]}
        if r["type"] in _DEBIT_NORMAL_TYPES:
            bal = round(d - c, 2)
        else:
            bal = round(c - d, 2)
        rows.append({"account": acct, "debit": d, "credit": c, "balance": bal})
        total_debit += d
        total_credit += c
    return rows, {"debit": round(total_debit, 2), "credit": round(total_credit, 2)}


def get_account_ledger(account_id, start_date=None, end_date=None):
    """Every journal line posted to one account, oldest first, with a running
    balance in the account's normal direction (debit-normal for assets/
    expenses, credit-normal otherwise). This is the full audit trail behind
    any Balance Sheet or Income Statement figure for that account.

    If start_date is given, activity before it is folded into an "opening
    balance" instead of dropped, so the running balance still ties out to
    the true all-time total (matches the Balance Sheet, which is not
    date-filtered)."""
    aid = int(account_id)
    acct = get_account(aid)
    if not acct:
        raise ValueError("Account not found.")
    is_debit_normal = acct["type"] in _DEBIT_NORMAL_TYPES

    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT je.entry_id, je.date, je.description, je.source_type, je.source_id,
                       jl.debit, jl.credit
                FROM journal_lines jl
                JOIN journal_entries je ON je.entry_id = jl.entry_id AND je.tenant_id = jl.tenant_id
                WHERE jl.account_id = %s AND jl.tenant_id = %s
                ORDER BY je.date, je.entry_id
            """, (aid, _tid()))
            raw = cur.fetchall()

    opening_balance = 0.0
    running = 0.0
    out_rows = []
    for r in raw:
        d = r["date"]
        if hasattr(d, "date"):
            d = d.date()
        debit = round(float(r["debit"] or 0), 2)
        credit = round(float(r["credit"] or 0), 2)
        delta = (debit - credit) if is_debit_normal else (credit - debit)
        if start_date is not None and d is not None and d < start_date:
            opening_balance = round(opening_balance + delta, 2)
            continue
        if end_date is not None and d is not None and d > end_date:
            continue
        running = round(running + delta, 2)
        out_rows.append({
            "entry_id": r["entry_id"], "date": d, "description": r["description"],
            "source_type": r["source_type"], "source_id": r["source_id"],
            "debit": debit, "credit": credit,
            "running_balance": round(opening_balance + running, 2),
        })

    return {
        "account": acct,
        "opening_balance": round(opening_balance, 2),
        "rows": out_rows,
        "closing_balance": out_rows[-1]["running_balance"] if out_rows else round(opening_balance, 2),
    }


# ── Expenses (built on the journal) ──────────────────────────────────

def record_expense(expense_account_id, amount, paid_from_account_id,
                   description="", created_by="system", entry_date=None):
    """Record an operating expense paid from a cash/bank account.
    Posts: Dr Expense, Cr Cash/Bank."""
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    exp = get_account(int(expense_account_id))
    if not exp or exp["type"] != "expense":
        raise ValueError("Select a valid expense category.")
    src = get_account(int(paid_from_account_id))
    if not src or src["type"] != "asset":
        raise ValueError("Select a valid account to pay from (Cash/Bank).")
    desc = (description or "").strip() or f"Expense: {exp['name']}"
    return post_journal(
        desc,
        [{"account_id": exp["account_id"], "debit": amount},
         {"account_id": src["account_id"], "credit": amount}],
        source_type="expense", created_by=created_by, entry_date=entry_date)


def get_expense_entries(start_date=None, end_date=None):
    """List expense journal entries with amount, category, and paid-from account."""
    where = ["je.source_type = 'expense'", "je.tenant_id = %s"]
    params = [_tid()]
    if start_date:
        where.append("DATE(je.date) >= %s")
        params.append(start_date)
    if end_date:
        where.append("DATE(je.date) <= %s")
        params.append(end_date)
    clause = " AND ".join(where)
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT je.entry_id, je.date, je.description, je.created_by,
                       da.account_id AS exp_id, da.code AS exp_code, da.name AS exp_name, da.type AS exp_type,
                       ca.account_id AS src_id, ca.code AS src_code, ca.name AS src_name, ca.type AS src_type,
                       dl.debit AS amount
                FROM journal_entries je
                JOIN journal_lines dl ON dl.entry_id = je.entry_id AND dl.debit > 0
                JOIN accounts da ON da.account_id = dl.account_id
                LEFT JOIN journal_lines cl ON cl.entry_id = je.entry_id AND cl.credit > 0
                LEFT JOIN accounts ca ON ca.account_id = cl.account_id
                WHERE {clause}
                ORDER BY je.entry_id DESC
            """, params)
            raw = cur.fetchall()
    entries = []
    for r in raw:
        entries.append({
            "entry_id": r["entry_id"],
            "date": r["date"],
            "description": r["description"],
            "category": {"account_id": r["exp_id"], "code": r["exp_code"],
                         "name": r["exp_name"], "type": r["exp_type"]} if r["exp_id"] else None,
            "paid_from": {"account_id": r["src_id"], "code": r["src_code"],
                          "name": r["src_name"], "type": r["src_type"]} if r["src_id"] else None,
            "amount": round(float(r["amount"]), 2),
            "created_by": r["created_by"],
        })
    return entries


def delete_expense(entry_id):
    eid = int(entry_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT entry_id FROM journal_entries WHERE entry_id=%s AND tenant_id=%s AND source_type='expense'",
                (eid, tid)
            )
            if not cur.fetchone():
                raise ValueError("Expense entry not found.")
            cur.execute("DELETE FROM journal_entries WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
        conn.commit()


def record_capital_injection(account_id, amount, description="", created_by="system",
                             entry_date=None):
    """Record extra owner funds added to Cash/Bank, outside the opening balance.
    Posts: Dr Cash/Bank, Cr Owner Capital."""
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    dest = get_account(int(account_id))
    if not dest or dest["type"] != "asset" or str(dest["code"]) not in ("1000", "1010"):
        raise ValueError("Select a valid account to add funds to (Cash/Bank).")
    accts = {str(a["code"]): a for a in get_all_accounts()}
    cap_acct = accts["3000"]
    desc = (description or "").strip() or f"Funds added to {dest['name']}"
    return post_journal(
        desc,
        [{"account_id": dest["account_id"], "debit": amount},
         {"account_id": cap_acct["account_id"], "credit": amount}],
        source_type="capital_injection", created_by=created_by, entry_date=entry_date)


def get_capital_injections(start_date=None, end_date=None):
    """List capital-injection journal entries with amount and destination account."""
    where = ["je.source_type = 'capital_injection'", "je.tenant_id = %s"]
    params = [_tid()]
    if start_date:
        where.append("DATE(je.date) >= %s")
        params.append(start_date)
    if end_date:
        where.append("DATE(je.date) <= %s")
        params.append(end_date)
    clause = " AND ".join(where)
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT je.entry_id, je.date, je.description, je.created_by,
                       da.account_id AS dest_id, da.code AS dest_code, da.name AS dest_name, da.type AS dest_type,
                       dl.debit AS amount
                FROM journal_entries je
                JOIN journal_lines dl ON dl.entry_id = je.entry_id AND dl.debit > 0
                JOIN accounts da ON da.account_id = dl.account_id
                WHERE {clause}
                ORDER BY je.entry_id DESC
            """, params)
            raw = cur.fetchall()
    entries = []
    for r in raw:
        entries.append({
            "entry_id": r["entry_id"],
            "date": r["date"],
            "description": r["description"],
            "account": {"account_id": r["dest_id"], "code": r["dest_code"],
                        "name": r["dest_name"], "type": r["dest_type"]} if r["dest_id"] else None,
            "amount": round(float(r["amount"]), 2),
            "created_by": r["created_by"],
        })
    return entries


def delete_capital_injection(entry_id):
    eid = int(entry_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT entry_id FROM journal_entries WHERE entry_id=%s AND tenant_id=%s AND source_type='capital_injection'",
                (eid, tid)
            )
            if not cur.fetchone():
                raise ValueError("Entry not found.")
            cur.execute("DELETE FROM journal_entries WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
        conn.commit()


def record_capital_withdrawal(account_id, amount, description="", created_by="system",
                              entry_date=None):
    """Record funds pulled out of Cash/Bank (owner draw, transfer out, etc),
    outside the opening balance. Posts: Dr Owner Capital, Cr Cash/Bank."""
    amount = round(float(amount), 2)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    src = get_account(int(account_id))
    if not src or src["type"] != "asset" or str(src["code"]) not in ("1000", "1010"):
        raise ValueError("Select a valid account to remove funds from (Cash/Bank).")
    bal = get_account_balance(src["account_id"])
    if amount > bal:
        raise ValueError(f"Insufficient balance in {src['name']} ({bal:,.2f} available).")
    accts = {str(a["code"]): a for a in get_all_accounts()}
    cap_acct = accts["3000"]
    desc = (description or "").strip() or f"Funds removed from {src['name']}"
    return post_journal(
        desc,
        [{"account_id": cap_acct["account_id"], "debit": amount},
         {"account_id": src["account_id"], "credit": amount}],
        source_type="capital_withdrawal", created_by=created_by, entry_date=entry_date)


def get_capital_movements(start_date=None, end_date=None):
    """List both capital_injection and capital_withdrawal entries together,
    each tagged with kind 'add'/'remove', newest first."""
    where = ["je.source_type IN ('capital_injection', 'capital_withdrawal')", "je.tenant_id = %s"]
    params = [_tid()]
    if start_date:
        where.append("DATE(je.date) >= %s")
        params.append(start_date)
    if end_date:
        where.append("DATE(je.date) <= %s")
        params.append(end_date)
    clause = " AND ".join(where)
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT je.entry_id, je.source_type, je.date, je.description, je.created_by,
                       da.account_id AS dest_id, da.code AS dest_code, da.name AS dest_name, da.type AS dest_type,
                       dl.debit AS debit_amount,
                       ca.account_id AS src_id, ca.code AS src_code, ca.name AS src_name, ca.type AS src_type,
                       cl.credit AS credit_amount
                FROM journal_entries je
                LEFT JOIN journal_lines dl ON dl.entry_id = je.entry_id AND dl.debit > 0
                LEFT JOIN accounts da ON da.account_id = dl.account_id
                LEFT JOIN journal_lines cl ON cl.entry_id = je.entry_id AND cl.credit > 0
                LEFT JOIN accounts ca ON ca.account_id = cl.account_id
                WHERE {clause}
                ORDER BY je.entry_id DESC
            """, params)
            raw = cur.fetchall()
    entries = []
    for r in raw:
        is_add = r["source_type"] == "capital_injection"
        kind = "add" if is_add else "remove"
        if is_add:
            acct = {"account_id": r["dest_id"], "code": r["dest_code"],
                    "name": r["dest_name"], "type": r["dest_type"]} if r["dest_id"] else None
            amount = r["debit_amount"]
        else:
            acct = {"account_id": r["src_id"], "code": r["src_code"],
                    "name": r["src_name"], "type": r["src_type"]} if r["src_id"] else None
            amount = r["credit_amount"]
        entries.append({
            "entry_id": r["entry_id"],
            "kind": kind,
            "date": r["date"],
            "description": r["description"],
            "account": acct,
            "amount": round(float(amount or 0), 2),
            "created_by": r["created_by"],
        })
    return entries


def delete_capital_withdrawal(entry_id):
    eid = int(entry_id)
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT entry_id FROM journal_entries WHERE entry_id=%s AND tenant_id=%s AND source_type='capital_withdrawal'",
                (eid, tid)
            )
            if not cur.fetchone():
                raise ValueError("Entry not found.")
            cur.execute("DELETE FROM journal_entries WHERE entry_id=%s AND tenant_id=%s", (eid, tid))
        conn.commit()


def update_account_name(account_id, name):
    aid = int(account_id)
    name = str(name).strip()
    if not name:
        raise ValueError("Name is required.")
    acct = get_account(aid)
    if not acct:
        raise ValueError("Account not found.")
    if acct.get("is_system"):
        raise ValueError("System accounts cannot be renamed.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE accounts SET name = %s WHERE account_id = %s AND tenant_id = %s",
                        (name, aid, _tid()))


def next_expense_code():
    """Next available expense account code in the 5xxx+ range."""
    codes = [int(a["code"]) for a in get_accounts_by_type("expense")
             if str(a["code"]).isdigit()]
    base = max(codes) if codes else 5000
    code = base + 10
    existing = {str(a["code"]) for a in get_all_accounts()}
    while str(code) in existing:
        code += 10
    return str(code)


# ── Opening balances, journal sync, balance sheet ────────────────────

def get_books_start():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT date FROM journal_entries WHERE source_type='opening' AND tenant_id=%s LIMIT 1",
                        (_tid(),))
            r = cur.fetchone()
    if not r or not r.get("date"):
        return None
    d = r["date"]
    return d.date() if isinstance(d, datetime) else d


def get_opening_balances():
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT a.code, a.type, jl.debit, jl.credit
                FROM journal_entries je
                JOIN journal_lines jl ON jl.entry_id = je.entry_id
                JOIN accounts a ON a.account_id = jl.account_id
                WHERE je.source_type = 'opening' AND je.tenant_id = %s
            """, (_tid(),))
            raw = cur.fetchall()
    out = {}
    for r in raw:
        amount = float(r["debit"]) if r["type"] in _DEBIT_NORMAL_TYPES else float(r["credit"])
        out[str(r["code"])] = round(amount, 2)
    return out


def count_operational_before(start_date):
    """Count sales/purchases/supplier-payments/customer-payments dated before
    start_date -- i.e. exactly the transactions that would stop being posted to
    the journal if start_date became the new books-start date (see _after_start
    in sync_journal_from_operations). Used to warn before a destructive change."""
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)

    def _before(val):
        if val is None:
            return False
        d = val.date() if isinstance(val, datetime) else val
        return d < start_date

    count = 0
    for inv in get_all_invoices():
        if _before(inv.get("created_at")):
            count += 1
    for p in get_all_purchase_invoices():
        if _before(p.get("created_at")):
            count += 1
    for e in get_supplier_ledger_entries():
        if e.get("type") != "credit" or e.get("purchase_id") not in (None, ""):
            continue
        if _before(e.get("created_at")):
            count += 1
    for e in get_credit_ledger():
        if e.get("type") != "credit":
            continue
        if _before(e.get("created_at")):
            count += 1
    return count


def set_opening_balances(start_date, balances, created_by="system"):
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    seed_chart_of_accounts()  # auto-heal a missing/partial chart of accounts
    acc = {str(a["code"]): a for a in get_all_accounts()}
    total_assets = 0.0
    total_liab = 0.0
    lines = []
    for code, amount in balances.items():
        amount = round(float(amount or 0), 2)
        if amount == 0:
            continue
        a = acc.get(str(code))
        if not a or str(code) == "3000":
            continue
        if a["type"] in _DEBIT_NORMAL_TYPES:
            lines.append((a["account_id"], amount, 0.0))
            total_assets += amount
        else:
            lines.append((a["account_id"], 0.0, amount))
            total_liab += amount
    capital = round(total_assets - total_liab, 2)
    cap_id = acc["3000"]["account_id"]
    if capital >= 0:
        lines.append((cap_id, 0.0, capital))
    else:
        lines.append((cap_id, -capital, 0.0))
    if len(lines) < 2:
        raise ValueError("Enter at least one opening balance.")
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            # Replace opening + clear auto-synced operational entries so the next
            # sync re-posts only transactions on/after the new start date.
            cur.execute("""DELETE FROM journal_entries
                           WHERE tenant_id = %s AND source_type IN ('opening','sale','purchase',
                                                 'supplier_payment','customer_payment')""", (tid,))
            cur.execute(
                """INSERT INTO journal_entries (tenant_id, date, description, source_type, source_id, created_by)
                   VALUES (%s,%s,%s,'opening',NULL,%s)""",
                (tid, datetime.combine(start_date, datetime.min.time()), "Opening balances", created_by))
            eid = cur.lastrowid
            cur.executemany(
                "INSERT INTO journal_lines (tenant_id, entry_id, account_id, debit, credit) VALUES (%s,%s,%s,%s,%s)",
                [(tid, eid, aid, d, c) for aid, d, c in lines])
            return eid


def clear_opening_balances():
    """Remove opening balances + auto-synced entries (next view re-syncs fresh)."""
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """DELETE FROM journal_entries WHERE tenant_id = %s AND source_type IN
                   ('opening','sale','purchase','supplier_payment','customer_payment')""",
                (_tid(),))


def _existing_journal_sources():
    out = set()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT source_type, source_id FROM journal_entries "
                        "WHERE source_id IS NOT NULL AND tenant_id = %s", (_tid(),))
            for r in cur.fetchall():
                out.add((str(r["source_type"]), str(int(r["source_id"]))))
    return out


def _realized_credit_invoice_ids():
    """Credit invoices considered paid: a customer's payments cover them oldest
    first. Returns (realized_set, breakdown) where breakdown maps invoice_id ->
    {"cash": x, "bank": y} showing which account(s) actually received the money,
    based on the payment_method recorded on the covering credit_ledger entries."""
    payments_by_cust = {}
    for e in get_credit_ledger():
        if e.get("type") == "credit":
            cid = normalize_customer_id(e.get("customer_id"))
            if cid is not None:
                pm = str(e.get("payment_method") or "cash").strip().lower()
                pm = pm if pm == "bank" else "cash"
                payments_by_cust.setdefault(cid, []).append({
                    "remaining": float(e.get("amount") or 0),
                    "method": pm,
                    "entry_id": int(e["entry_id"]),
                })
    credit_by_cust = {}
    for inv in get_all_invoices():
        if str(inv.get("payment_method") or "").strip().lower() == "credit":
            cid = normalize_customer_id(inv.get("customer_id"))
            if cid is not None:
                credit_by_cust.setdefault(cid, []).append(inv)

    realized = set()
    breakdown = {}
    for cid, invs in credit_by_cust.items():
        invs.sort(key=lambda x: int(x["invoice_id"]))
        payments = sorted(payments_by_cust.get(cid, []), key=lambda p: p["entry_id"])
        ppos = 0
        for inv in invs:
            need = round(float(inv.get("total") or 0), 2)
            alloc = {"cash": 0.0, "bank": 0.0}
            while need > 0.0001 and ppos < len(payments):
                pay = payments[ppos]
                take = min(need, pay["remaining"])
                alloc[pay["method"]] = round(alloc[pay["method"]] + take, 2)
                pay["remaining"] = round(pay["remaining"] - take, 2)
                need = round(need - take, 2)
                if pay["remaining"] <= 0.0001:
                    ppos += 1
            if need <= 0.0001:
                iid = int(inv["invoice_id"])
                realized.add(iid)
                breakdown[iid] = alloc
            else:
                break
    return realized, breakdown


def _has_stale_ar_entries():
    """True if any journal line ever posted to Accounts Receivable (1100).
    AR is no longer used by this posting model (see sync_journal_from_operations);
    its presence means the derived journal was built by an earlier model and
    needs a one-time rebuild."""
    accts = {a["account_id"]: a for a in get_all_accounts()}
    ar_ids = {aid for aid, a in accts.items() if str(a["code"]) == "1100"}
    if not ar_ids:
        return False
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT 1 FROM journal_lines WHERE tenant_id = %s AND account_id IN %s LIMIT 1",
                (tid, tuple(ar_ids)))
            return cur.fetchone() is not None


def sync_journal_from_operations():
    """Generate journal entries for sales, purchases, supplier payments and
    customer payments not yet posted. A sale's Inventory/COGS moves immediately
    (physical stock left regardless of payment status), but Sales revenue for
    a credit sale is only recognised as the customer actually pays -- partial
    or full, whenever it's recorded. There is no Accounts Receivable account:
    outstanding customer balances live entirely in the Credit Ledger. Idempotent.
    Returns count posted."""
    seed_chart_of_accounts()  # auto-heal a missing/partial chart of accounts
    acc = {str(a["code"]): a["account_id"] for a in get_all_accounts()}
    CASH, BANK, INV, AP = acc["1000"], acc["1010"], acc["1200"], acc["2000"]
    TAX, SALES, COGS = acc["2100"], acc["4000"], acc["5000"]

    if _has_stale_ar_entries():
        # An earlier posting model used Accounts Receivable; wipe the derived
        # journal so it can be rebuilt below under the current model. Opening
        # balances and manual/expense entries are untouched. "customer_charge"
        # is a retired source_type from an interim model -- swept up here too
        # so no orphaned AR lines survive the rebuild.
        with _conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """DELETE FROM journal_entries WHERE tenant_id = %s AND source_type IN
                       ('sale','purchase','supplier_payment','customer_payment','customer_charge')""",
                    (_tid(),))

    existing = _existing_journal_sources()
    books_start = get_books_start()

    def _after_start(val):
        if books_start is None:
            return True
        d = val.date() if isinstance(val, datetime) else val
        return d is None or d >= books_start

    pending = []
    # Cash/bank sales post in full immediately: Dr Cash/Bank, Cr Sales (+Tax),
    # Dr COGS/Cr Inventory. A credit sale only moves the stock side at sale
    # time (Dr COGS/Cr Inventory -- the goods physically left regardless of
    # payment); its revenue is recognised later, per payment, below.
    for inv in get_all_invoices():
        sid = int(inv["invoice_id"])
        if ("sale", str(sid)) in existing or not _after_start(inv.get("created_at")):
            continue
        is_credit = str(inv.get("payment_method") or "").strip().lower() == "credit"
        items = get_invoice_items(sid)
        cogs = round(sum(float(it["purchase_price"] or 0) * int(it["quantity"] or 0) for it in items), 2)
        lines = []
        if is_credit:
            if cogs > 0:
                lines.append((COGS, cogs, 0.0))
                lines.append((INV, 0.0, cogs))
        else:
            net = round(float(inv.get("subtotal") or 0) - float(inv.get("discount_total") or 0), 2)
            tax = round(float(inv.get("tax_amount") or 0), 2)
            delivery = round(float(inv.get("delivery_charges") or 0), 2)
            total = round(float(inv.get("total") or 0), 2)
            pm = str(inv.get("payment_method") or "cash").strip().lower()
            recv_acct = BANK if pm == "bank" else CASH
            lines.append((recv_acct, total, 0.0))
            lines.append((SALES, 0.0, net + delivery))
            if tax > 0:
                lines.append((TAX, 0.0, tax))
            if cogs > 0:
                lines.append((COGS, cogs, 0.0))
                lines.append((INV, 0.0, cogs))
        if lines:
            pending.append((inv.get("created_at"), f"Sale INV-{sid}", "sale", sid, lines))

    _pay_acct = {"credit": AP, "cash": CASH, "bank": BANK}
    for p in get_all_purchase_invoices():
        pid = int(p["purchase_id"])
        if ("purchase", str(pid)) in existing or not _after_start(p.get("created_at")):
            continue
        total = round(float(p.get("total_amount") or 0), 2)
        if total <= 0:
            continue
        pm = str(p.get("payment_method") or "credit").strip().lower()
        credit_acct = _pay_acct.get(pm, AP)
        pending.append((p.get("created_at"), f"Purchase PINV-{pid}", "purchase", pid,
                        [(INV, total, 0.0), (credit_acct, 0.0, total)]))

    for e in get_supplier_ledger_entries():
        if e.get("type") != "credit":
            continue
        if e.get("purchase_id") not in (None, ""):
            continue  # immediate purchase payment, already credited in the purchase entry
        eid = int(e["entry_id"])
        if ("supplier_payment", str(eid)) in existing or not _after_start(e.get("created_at")):
            continue
        amt = round(float(e.get("amount") or 0), 2)
        if amt <= 0:
            continue
        pm = (e.get("payment_method") or "cash").strip().lower()
        cash_acct = BANK if pm == "bank" else CASH
        pending.append((e.get("created_at"), "Supplier payment", "supplier_payment", eid,
                        [(AP, amt, 0.0), (cash_acct, 0.0, amt)]))

    # Customer payments -> Dr Cash/Bank, Cr Sales. Revenue for a credit sale
    # (or a manual charge) is recognised right here, exactly as it's actually
    # collected -- partial or full -- instead of at sale time.
    for e in get_credit_ledger():
        if e.get("type") != "credit":
            continue
        eid = int(e["entry_id"])
        if ("customer_payment", str(eid)) in existing or not _after_start(e.get("created_at")):
            continue
        amt = round(float(e.get("amount") or 0), 2)
        if amt <= 0:
            continue
        pm = str(e.get("payment_method") or "cash").strip().lower()
        recv_acct = BANK if pm == "bank" else CASH
        pending.append((e.get("created_at"), "Customer payment", "customer_payment", eid,
                        [(recv_acct, amt, 0.0), (SALES, 0.0, amt)]))

    if not pending:
        return 0
    tid = _tid()
    with _conn() as conn:
        with conn.cursor() as cur:
            for when, desc, stype, sid, lines in pending:
                td = round(sum(d for _, d, _ in lines), 2)
                tc = round(sum(c for _, _, c in lines), 2)
                if td != tc:
                    continue
                cur.execute(
                    """INSERT INTO journal_entries (tenant_id, date, description, source_type, source_id, created_by)
                       VALUES (%s,%s,%s,%s,%s,'system')""", (tid, when, desc, stype, sid))
                eid = cur.lastrowid
                cur.executemany(
                    "INSERT INTO journal_lines (tenant_id, entry_id, account_id, debit, credit) VALUES (%s,%s,%s,%s,%s)",
                    [(tid, eid, aid, d, c) for aid, d, c in lines])
    return len(pending)


def get_balance_sheet():
    rows, _ = get_trial_balance()
    sections = {"asset": [], "liability": [], "equity": []}
    total = {"asset": 0.0, "liability": 0.0, "equity": 0.0}
    income = 0.0
    expense = 0.0
    for r in rows:
        atype = r["account"]["type"]
        bal = r["balance"]
        if atype in sections:
            sections[atype].append({
                "name": r["account"]["name"],
                "account_id": r["account"]["account_id"],
                "balance": bal,
            })
            total[atype] += bal
        elif atype == "income":
            income += bal
        elif atype == "expense":
            expense += bal
    net_income = round(income - expense, 2)
    equity_rows = list(sections["equity"])
    equity_rows.append({"name": "Net Income (current)", "balance": net_income})
    total_equity = round(total["equity"] + net_income, 2)
    return {
        "assets": sections["asset"],
        "liabilities": sections["liability"],
        "equity": equity_rows,
        "total_assets": round(total["asset"], 2),
        "total_liabilities": round(total["liability"], 2),
        "total_equity": total_equity,
        "total_liab_equity": round(total["liability"] + total_equity, 2),
        "net_income": net_income,
    }


# ── Partners / investor equity ───────────────────────────────────────

def get_all_partners(include_inactive=False):
    with _conn() as conn:
        with conn.cursor() as cur:
            if include_inactive:
                cur.execute("SELECT * FROM partners WHERE tenant_id = %s ORDER BY partner_id", (_tid(),))
            else:
                cur.execute("SELECT * FROM partners WHERE tenant_id = %s AND is_active = 1 ORDER BY partner_id",
                            (_tid(),))
            rows = cur.fetchall()
    for r in rows:
        r["share_pct"] = float(r["share_pct"] or 0)
        r["is_active"] = bool(r["is_active"])
    return rows


def get_partner(partner_id):
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM partners WHERE partner_id = %s AND tenant_id = %s",
                        (int(partner_id), _tid()))
            r = cur.fetchone()
    if r:
        r["share_pct"] = float(r["share_pct"] or 0)
        r["is_active"] = bool(r["is_active"])
    return r


def add_partner(name, share_pct):
    name = str(name).strip()
    if not name:
        raise ValueError("Partner name is required.")
    share = round(float(share_pct or 0), 2)
    if share < 0 or share > 100:
        raise ValueError("Share % must be between 0 and 100.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO partners (tenant_id, name, share_pct, is_active) VALUES (%s,%s,%s,1)",
                        (_tid(), name, share))
            return cur.lastrowid


def update_partner(partner_id, name, share_pct, is_active=True):
    name = str(name).strip()
    if not name:
        raise ValueError("Partner name is required.")
    share = round(float(share_pct or 0), 2)
    if share < 0 or share > 100:
        raise ValueError("Share % must be between 0 and 100.")
    if not get_partner(partner_id):
        raise ValueError("Partner not found.")
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute("UPDATE partners SET name=%s, share_pct=%s, is_active=%s WHERE partner_id=%s AND tenant_id=%s",
                        (name, share, 1 if is_active else 0, int(partner_id), _tid()))


def add_partner_transaction(partner_id, txn_type, amount, note="", txn_date=None,
                            created_by="system"):
    pid = int(partner_id)
    if txn_type not in ("capital", "drawing"):
        raise ValueError("Type must be 'capital' or 'drawing'.")
    amount = round(float(amount or 0), 2)
    if amount <= 0:
        raise ValueError("Amount must be positive.")
    if not get_partner(pid):
        raise ValueError("Partner not found.")
    when = txn_date or date.today()
    if isinstance(when, str):
        when = date.fromisoformat(when)
    when_dt = datetime.combine(when, datetime.min.time())
    with _conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO partner_transactions (tenant_id, partner_id, type, amount, note, date)
                   VALUES (%s,%s,%s,%s,%s,%s)""",
                (_tid(), pid, txn_type, amount, (note or "").strip(), when_dt))
            txn_id = cur.lastrowid
    acc = {str(a["code"]): a["account_id"] for a in get_all_accounts()}
    if txn_type == "capital":
        lines = [{"account_id": acc["1000"], "debit": amount},
                 {"account_id": acc["3000"], "credit": amount}]
        desc, stype = "Partner capital contribution", "partner_capital"
    else:
        lines = [{"account_id": acc["3100"], "debit": amount},
                 {"account_id": acc["1000"], "credit": amount}]
        desc, stype = "Partner drawing", "partner_drawing"
    post_journal(desc, lines, source_type=stype, source_id=txn_id,
                 created_by=created_by, entry_date=when_dt)
    return txn_id


def get_partner_transactions(partner_id=None):
    with _conn() as conn:
        with conn.cursor() as cur:
            if partner_id is not None:
                cur.execute("SELECT * FROM partner_transactions WHERE partner_id = %s AND tenant_id = %s ORDER BY txn_id DESC",
                            (int(partner_id), _tid()))
            else:
                cur.execute("SELECT * FROM partner_transactions WHERE tenant_id = %s ORDER BY txn_id DESC",
                            (_tid(),))
            return cur.fetchall()


def get_partner_equity():
    partners = get_all_partners()
    net_income = get_balance_sheet()["net_income"]
    txns = get_partner_transactions()
    cap_by, draw_by = {}, {}
    for t in txns:
        p = int(t["partner_id"])
        amt = float(t["amount"] or 0)
        if t["type"] == "capital":
            cap_by[p] = cap_by.get(p, 0.0) + amt
        elif t["type"] == "drawing":
            draw_by[p] = draw_by.get(p, 0.0) + amt
    rows = []
    for p in partners:
        pid = p["partner_id"]
        cap = round(cap_by.get(pid, 0.0), 2)
        draw = round(draw_by.get(pid, 0.0), 2)
        profit_share = round(net_income * (p["share_pct"] / 100.0), 2)
        equity = round(cap + profit_share - draw, 2)
        rows.append({"partner": p, "capital": cap, "share_pct": p["share_pct"],
                     "profit_share": profit_share, "drawings": draw, "equity": equity})
    totals = {
        "capital": round(sum(r["capital"] for r in rows), 2),
        "profit_share": round(sum(r["profit_share"] for r in rows), 2),
        "drawings": round(sum(r["drawings"] for r in rows), 2),
        "equity": round(sum(r["equity"] for r in rows), 2),
        "share_pct": round(sum(p["share_pct"] for p in partners), 2),
        "net_income": net_income,
    }
    return rows, totals
